from typing import Any, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.database import get_db
from app.dependencies import get_current_active_user, require_admin, require_supervisor_or_admin
from app.models.cargo import Cargo
from app.models.criterio_exclusion import CriterioExclusion
from app.models.factor_riesgo import CategoriaFactorRiesgo, FactorRiesgo
from app.models.profesiograma import (
    Profesiograma,
    ProfesiogramaEstado,
    ProfesiogramaExamen,
    ProfesiogramaFactor,
    ProfesiogramaControlESIAE,
    ProfesiogramaIntervencion,
)
from app.models.profesiograma_inmunizacion import ProfesiogramaInmunizacion
from app.models.inmunizacion import Inmunizacion
from app.models.tipo_examen import TipoExamen
from app.models.user import User
from app.schemas.criterio_exclusion import (
    CriterioExclusion as CriterioExclusionSchema,
    CriterioExclusionCreate,
    CriterioExclusionUpdate,
)
from app.schemas.factor_riesgo import (
    FactorRiesgo as FactorRiesgoSchema,
    FactorRiesgoCreate,
    FactorRiesgoUpdate,
)
from app.schemas.profesiograma import (
    Profesiograma as ProfesiogramaSchema,
    ProfesiogramaCreate,
    ProfesiogramaDuplicateRequest,
    ProfesiogramaDuplicateResult,
    ProfesiogramaEmoSuggestion,
    ProfesiogramaEmoJustificacionRequest,
    ProfesiogramaUpdate,
)
from app.schemas.tipo_examen import TipoExamen as TipoExamenSchema, TipoExamenCreate, TipoExamenUpdate
from app.schemas.inmunizacion import Inmunizacion as InmunizacionSchema, InmunizacionCreate, InmunizacionUpdate
from app.schemas.common import MessageResponse, PaginatedResponse
from app.services.emo_periodicidad import (
    compute_stats_for_cargo,
    compute_matrix_risk_summary_from_inputs,
    generate_justificacion_periodicidad_emo_from_db,
    generate_justificacion_periodicidad_emo,
    suggest_periodicidad_emo_meses,
    suggest_periodicidad_emo_meses_from_db,
)
from app.services.profesiograma_pdf import generate_profesiograma_report_pdf


router = APIRouter()


@router.get("/catalogos/factores-riesgo", response_model=List[FactorRiesgoSchema])
def list_factores_riesgo(
    activo: Optional[bool] = Query(None),
    q: Optional[str] = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    query = db.query(FactorRiesgo)
    if activo is not None:
        query = query.filter(FactorRiesgo.activo == activo)
    if q:
        like = f"%{q}%"
        query = query.filter((FactorRiesgo.nombre.ilike(like)) | (FactorRiesgo.codigo.ilike(like)))
    return query.order_by(FactorRiesgo.nombre.asc()).all()


@router.get(
    "/catalogos/factores-riesgo/search",
    response_model=PaginatedResponse[FactorRiesgoSchema],
)
def search_factores_riesgo(
    q: Optional[str] = Query(None),
    activo: Optional[bool] = Query(None),
    categoria: Optional[CategoriaFactorRiesgo] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    query = db.query(FactorRiesgo)
    if activo is not None:
        query = query.filter(FactorRiesgo.activo == activo)
    if categoria is not None:
        query = query.filter(FactorRiesgo.categoria == categoria.value)
    if q:
        like = f"%{q}%"
        query = query.filter((FactorRiesgo.nombre.ilike(like)) | (FactorRiesgo.codigo.ilike(like)))

    total = query.count()
    skip = (page - 1) * size
    items = query.order_by(FactorRiesgo.nombre.asc()).offset(skip).limit(size).all()

    pages = (total + size - 1) // size if total else 0
    has_next = page < pages
    has_prev = page > 1 and pages > 0
    return {
        "items": items,
        "total": total,
        "page": page,
        "size": size,
        "pages": pages,
        "has_next": has_next,
        "has_prev": has_prev,
    }


@router.post("/catalogos/factores-riesgo", response_model=FactorRiesgoSchema, status_code=status.HTTP_201_CREATED)
def create_factor_riesgo(
    payload: FactorRiesgoCreate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    exists = db.query(FactorRiesgo).filter(FactorRiesgo.codigo == payload.codigo).first()
    if exists:
        raise HTTPException(status_code=400, detail="Ya existe un FactorRiesgo con ese codigo")
    factor = FactorRiesgo(**payload.model_dump())
    db.add(factor)
    db.commit()
    db.refresh(factor)
    return factor


@router.put("/catalogos/factores-riesgo/{factor_id}", response_model=FactorRiesgoSchema)
def update_factor_riesgo(
    factor_id: int,
    payload: FactorRiesgoUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    factor = db.query(FactorRiesgo).filter(FactorRiesgo.id == factor_id).first()
    if not factor:
        raise HTTPException(status_code=404, detail="FactorRiesgo no encontrado")
    data = payload.model_dump(exclude_unset=True)
    if "codigo" in data:
        exists = (
            db.query(FactorRiesgo)
            .filter(FactorRiesgo.codigo == data["codigo"], FactorRiesgo.id != factor_id)
            .first()
        )
        if exists:
            raise HTTPException(status_code=400, detail="Ya existe un FactorRiesgo con ese codigo")
    for k, v in data.items():
        setattr(factor, k, v)
    db.commit()
    db.refresh(factor)
    return factor


@router.delete("/catalogos/factores-riesgo/{factor_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_factor_riesgo(
    factor_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    factor = db.query(FactorRiesgo).filter(FactorRiesgo.id == factor_id).first()
    if not factor:
        raise HTTPException(status_code=404, detail="FactorRiesgo no encontrado")
    try:
        db.delete(factor)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"No se puede eliminar el factor de riesgo: {str(e)}")
    return None


@router.get("/catalogos/tipos-examen", response_model=List[TipoExamenSchema])
def list_tipos_examen(
    activo: Optional[bool] = Query(None),
    q: Optional[str] = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    query = db.query(TipoExamen)
    if activo is not None:
        query = query.filter(TipoExamen.activo == activo)
    if q:
        like = f"%{q}%"
        query = query.filter(TipoExamen.nombre.ilike(like))
    return query.order_by(TipoExamen.nombre.asc()).all()


@router.get(
    "/catalogos/tipos-examen/search",
    response_model=PaginatedResponse[TipoExamenSchema],
)
def search_tipos_examen(
    q: Optional[str] = Query(None),
    activo: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    query = db.query(TipoExamen)
    if activo is not None:
        query = query.filter(TipoExamen.activo == activo)
    if q:
        like = f"%{q}%"
        query = query.filter(TipoExamen.nombre.ilike(like))

    total = query.count()
    skip = (page - 1) * size
    items = query.order_by(TipoExamen.nombre.asc()).offset(skip).limit(size).all()

    pages = (total + size - 1) // size if total else 0
    has_next = page < pages
    has_prev = page > 1 and pages > 0
    return {
        "items": items,
        "total": total,
        "page": page,
        "size": size,
        "pages": pages,
        "has_next": has_next,
        "has_prev": has_prev,
    }


@router.post("/catalogos/tipos-examen", response_model=TipoExamenSchema, status_code=status.HTTP_201_CREATED)
def create_tipo_examen(
    payload: TipoExamenCreate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    exists = db.query(TipoExamen).filter(TipoExamen.nombre == payload.nombre).first()
    if exists:
        raise HTTPException(status_code=400, detail="Ya existe un TipoExamen con ese nombre")
    tipo = TipoExamen(**payload.model_dump())
    db.add(tipo)
    db.commit()
    db.refresh(tipo)
    return tipo


@router.delete("/catalogos/tipos-examen/{tipo_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_tipo_examen(
    tipo_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    tipo = db.query(TipoExamen).filter(TipoExamen.id == tipo_id).first()
    if not tipo:
        raise HTTPException(status_code=404, detail="TipoExamen no encontrado")
    
    # Check for related records if necessary, or let DB constraints handle it
    # For now, we proceed with deletion
    try:
        db.delete(tipo)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"No se puede eliminar el tipo de examen: {str(e)}")
    
    return None


@router.put("/catalogos/tipos-examen/{tipo_id}", response_model=TipoExamenSchema)
def update_tipo_examen(
    tipo_id: int,
    payload: TipoExamenUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    tipo = db.query(TipoExamen).filter(TipoExamen.id == tipo_id).first()
    if not tipo:
        raise HTTPException(status_code=404, detail="TipoExamen no encontrado")
    data = payload.model_dump(exclude_unset=True)
    if "nombre" in data:
        exists = (
            db.query(TipoExamen)
            .filter(TipoExamen.nombre == data["nombre"], TipoExamen.id != tipo_id)
            .first()
        )
        if exists:
            raise HTTPException(status_code=400, detail="Ya existe un TipoExamen con ese nombre")
    for k, v in data.items():
        setattr(tipo, k, v)
    db.commit()
    db.refresh(tipo)
    return tipo


@router.get("/catalogos/criterios-exclusion", response_model=List[CriterioExclusionSchema])
def list_criterios_exclusion(
    q: Optional[str] = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    query = db.query(CriterioExclusion)
    if q:
        like = f"%{q}%"
        query = query.filter(CriterioExclusion.nombre.ilike(like))
    return query.order_by(CriterioExclusion.nombre.asc()).all()


@router.get(
    "/catalogos/criterios-exclusion/search",
    response_model=PaginatedResponse[CriterioExclusionSchema],
)
def search_criterios_exclusion(
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    query = db.query(CriterioExclusion)
    if q:
        like = f"%{q}%"
        query = query.filter(CriterioExclusion.nombre.ilike(like))

    total = query.count()
    skip = (page - 1) * size
    items = query.order_by(CriterioExclusion.nombre.asc()).offset(skip).limit(size).all()

    pages = (total + size - 1) // size if total else 0
    has_next = page < pages
    has_prev = page > 1 and pages > 0
    return {
        "items": items,
        "total": total,
        "page": page,
        "size": size,
        "pages": pages,
        "has_next": has_next,
        "has_prev": has_prev,
    }


@router.post("/catalogos/criterios-exclusion", response_model=CriterioExclusionSchema, status_code=status.HTTP_201_CREATED)
def create_criterio_exclusion(
    payload: CriterioExclusionCreate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    exists = db.query(CriterioExclusion).filter(CriterioExclusion.nombre == payload.nombre).first()
    if exists:
        raise HTTPException(status_code=400, detail="Ya existe un CriterioExclusion con ese nombre")
    criterio = CriterioExclusion(**payload.model_dump())
    db.add(criterio)
    db.commit()
    db.refresh(criterio)
    return criterio


@router.put("/catalogos/criterios-exclusion/{criterio_id}", response_model=CriterioExclusionSchema)
def update_criterio_exclusion(
    criterio_id: int,
    payload: CriterioExclusionUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    criterio = db.query(CriterioExclusion).filter(CriterioExclusion.id == criterio_id).first()
    if not criterio:
        raise HTTPException(status_code=404, detail="CriterioExclusion no encontrado")
    data = payload.model_dump(exclude_unset=True)
    if "nombre" in data:
        exists = (
            db.query(CriterioExclusion)
            .filter(CriterioExclusion.nombre == data["nombre"], CriterioExclusion.id != criterio_id)
            .first()
        )
        if exists:
            raise HTTPException(status_code=400, detail="Ya existe un CriterioExclusion con ese nombre")
    for k, v in data.items():
        setattr(criterio, k, v)
    db.commit()
    db.refresh(criterio)
    return criterio


@router.delete("/catalogos/criterios-exclusion/{criterio_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_criterio_exclusion(
    criterio_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    criterio = db.query(CriterioExclusion).filter(CriterioExclusion.id == criterio_id).first()
    if not criterio:
        raise HTTPException(status_code=404, detail="CriterioExclusion no encontrado")
    try:
        db.delete(criterio)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"No se puede eliminar el criterio de exclusión: {str(e)}")
    return None


@router.get("/catalogos/inmunizaciones", response_model=List[InmunizacionSchema])
def list_inmunizaciones(
    activo: Optional[bool] = Query(None),
    q: Optional[str] = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    query = db.query(Inmunizacion)
    if activo is not None:
        query = query.filter(Inmunizacion.activo == activo)
    if q:
        like = f"%{q}%"
        query = query.filter(Inmunizacion.nombre.ilike(like))
    return query.order_by(Inmunizacion.nombre.asc()).all()


@router.get(
    "/catalogos/inmunizaciones/search",
    response_model=PaginatedResponse[InmunizacionSchema],
)
def search_inmunizaciones(
    q: Optional[str] = Query(None),
    activo: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> Any:
    query = db.query(Inmunizacion)
    if activo is not None:
        query = query.filter(Inmunizacion.activo == activo)
    if q:
        like = f"%{q}%"
        query = query.filter(Inmunizacion.nombre.ilike(like))

    total = query.count()
    skip = (page - 1) * size
    items = query.order_by(Inmunizacion.nombre.asc()).offset(skip).limit(size).all()

    pages = (total + size - 1) // size if total else 0
    has_next = page < pages
    has_prev = page > 1 and pages > 0
    return {
        "items": items,
        "total": total,
        "page": page,
        "size": size,
        "pages": pages,
        "has_next": has_next,
        "has_prev": has_prev,
    }


@router.post("/catalogos/inmunizaciones", response_model=InmunizacionSchema, status_code=status.HTTP_201_CREATED)
def create_inmunizacion(
    payload: InmunizacionCreate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    exists = db.query(Inmunizacion).filter(Inmunizacion.nombre == payload.nombre).first()
    if exists:
        raise HTTPException(status_code=400, detail="Ya existe una Inmunización con ese nombre")
    inmun = Inmunizacion(**payload.model_dump())
    db.add(inmun)
    db.commit()
    db.refresh(inmun)
    return inmun


@router.put("/catalogos/inmunizaciones/{inmun_id}", response_model=InmunizacionSchema)
def update_inmunizacion(
    inmun_id: int,
    payload: InmunizacionUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    inmun = db.query(Inmunizacion).filter(Inmunizacion.id == inmun_id).first()
    if not inmun:
        raise HTTPException(status_code=404, detail="Inmunización no encontrada")
    data = payload.model_dump(exclude_unset=True)
    if "nombre" in data:
        exists = (
            db.query(Inmunizacion)
            .filter(Inmunizacion.nombre == data["nombre"], Inmunizacion.id != inmun_id)
            .first()
        )
        if exists:
            raise HTTPException(status_code=400, detail="Ya existe una Inmunización con ese nombre")
    for k, v in data.items():
        setattr(inmun, k, v)
    db.commit()
    db.refresh(inmun)
    return inmun


@router.delete("/catalogos/inmunizaciones/{inmun_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_inmunizacion(
    inmun_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    inmun = db.query(Inmunizacion).filter(Inmunizacion.id == inmun_id).first()
    if not inmun:
        raise HTTPException(status_code=404, detail="Inmunización no encontrada")
    
    try:
        db.delete(inmun)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"No se puede eliminar la inmunización: {str(e)}")
    
    return None


def _serialize_profesiograma(p: Profesiograma) -> ProfesiogramaSchema:
    criterios_ids = [c.id for c in getattr(p, "criterios_exclusion", [])]
    factores = list(getattr(p, "profesiograma_factores", []) or [])
    return ProfesiogramaSchema(
        id=p.id,
        cargo_id=p.cargo_id,
        version=p.version,
        estado=p.estado,
        empresa=p.empresa,
        departamento=p.departamento,
        codigo_cargo=p.codigo_cargo,
        numero_trabajadores_expuestos=p.numero_trabajadores_expuestos,
        fecha_elaboracion=p.fecha_elaboracion,
        validado_por=p.validado_por,
        proxima_revision=p.proxima_revision,
        elaborado_por=p.elaborado_por,
        revisado_por=p.revisado_por,
        aprobado_por=p.aprobado_por,
        fecha_aprobacion=p.fecha_aprobacion,
        vigencia_meses=p.vigencia_meses,
        posicion_predominante=p.posicion_predominante,
        descripcion_actividades=p.descripcion_actividades,
        periodicidad_emo_meses=p.periodicidad_emo_meses,
        justificacion_periodicidad_emo=p.justificacion_periodicidad_emo,
        fecha_ultima_revision=p.fecha_ultima_revision,
        nivel_riesgo_cargo=p.nivel_riesgo_cargo,
        creado_por=p.creado_por,
        fecha_creacion=p.fecha_creacion,
        modificado_por=p.modificado_por,
        fecha_modificacion=p.fecha_modificacion,
        factores=factores,
        profesiograma_factores=factores,
        examenes=list(getattr(p, "examenes", []) or []),
        inmunizaciones=list(getattr(p, "inmunizaciones", []) or []),
        criterios_exclusion_ids=criterios_ids,
    )


@router.get("/cargos/{cargo_id}", response_model=List[ProfesiogramaSchema])
def list_profesiogramas_by_cargo(
    cargo_id: int,
    estado: Optional[ProfesiogramaEstado] = Query(None),
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    cargo = db.query(Cargo).filter(Cargo.id == cargo_id).first()
    if not cargo:
        raise HTTPException(status_code=404, detail="Cargo no encontrado")
    query = db.query(Profesiograma).filter(Profesiograma.cargo_id == cargo_id)
    if estado:
        query = query.filter(Profesiograma.estado == estado)
    profes = query.order_by(Profesiograma.fecha_creacion.desc()).all()
    return [_serialize_profesiograma(p) for p in profes]


@router.get("/cargos/{cargo_id}/emo/sugerencia", response_model=ProfesiogramaEmoSuggestion)
def suggest_emo_periodicidad_for_cargo(
    cargo_id: int,
    periodicidad_emo_meses: Optional[int] = Query(None),
    formato: str = Query("breve"),
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    cargo = db.query(Cargo).filter(Cargo.id == cargo_id).first()
    if not cargo:
        raise HTTPException(status_code=404, detail="Cargo no encontrado")

    stats = compute_stats_for_cargo(db, cargo_id)
    recomendada = suggest_periodicidad_emo_meses_from_db(db, cargo_id)
    periodicidad_para_borrador = periodicidad_emo_meses or recomendada
    if periodicidad_para_borrador not in (6, 12, 24, 36):
        raise HTTPException(status_code=400, detail="periodicidad_emo_meses debe ser 6, 12, 24 o 36")
    borrador = generate_justificacion_periodicidad_emo_from_db(
        db,
        cargo_id,
        periodicidad_para_borrador,
        formato=formato,
    )
    return ProfesiogramaEmoSuggestion(
        cargo_id=cargo_id,
        periodicidad_sugerida=recomendada,
        periodicidad_borrador=periodicidad_para_borrador,
        numero_trabajadores_expuestos=stats.total_trabajadores_activos,
        menores_21=stats.menores_21,
        antiguedad_menor_2_anios=stats.antiguedad_menor_2_anios,
        sin_fecha_ingreso=stats.sin_fecha_ingreso,
        justificacion_periodicidad_emo_borrador=borrador,
    )


@router.post("/cargos/{cargo_id}/emo/justificacion", response_model=ProfesiogramaEmoSuggestion)
def build_emo_justificacion_for_cargo(
    cargo_id: int,
    payload: ProfesiogramaEmoJustificacionRequest,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    cargo = db.query(Cargo).filter(Cargo.id == cargo_id).first()
    if not cargo:
        raise HTTPException(status_code=404, detail="Cargo no encontrado")
    if payload.periodicidad_emo_meses not in (6, 12, 24, 36):
        raise HTTPException(status_code=400, detail="periodicidad_emo_meses debe ser 6, 12, 24 o 36")

    stats = compute_stats_for_cargo(db, cargo_id)
    factores = [(f.factor_riesgo_id, f.nd, f.ne, f.nc) for f in (payload.factores or [])]
    matrix_override = compute_matrix_risk_summary_from_inputs(db, factores)
    recomendada = suggest_periodicidad_emo_meses_from_db(db, cargo_id, matrix_override=matrix_override)
    borrador = generate_justificacion_periodicidad_emo_from_db(
        db,
        cargo_id,
        payload.periodicidad_emo_meses,
        matrix_override=matrix_override,
        matrix_label="configuración actual",
        formato=payload.formato,
    )
    return ProfesiogramaEmoSuggestion(
        cargo_id=cargo_id,
        periodicidad_sugerida=recomendada,
        periodicidad_borrador=payload.periodicidad_emo_meses,
        numero_trabajadores_expuestos=stats.total_trabajadores_activos,
        menores_21=stats.menores_21,
        antiguedad_menor_2_anios=stats.antiguedad_menor_2_anios,
        sin_fecha_ingreso=stats.sin_fecha_ingreso,
        justificacion_periodicidad_emo_borrador=borrador,
    )


@router.get("/{profesiograma_id}", response_model=ProfesiogramaSchema)
def get_profesiograma(
    profesiograma_id: int,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    p = db.query(Profesiograma).filter(Profesiograma.id == profesiograma_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Profesiograma no encontrado")
    return _serialize_profesiograma(p)


@router.delete("/{profesiograma_id}", response_model=MessageResponse)
def delete_profesiograma(
    profesiograma_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.query(Profesiograma).filter(Profesiograma.id == profesiograma_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Profesiograma no encontrado")
    try:
        db.delete(p)
        db.commit()
        return MessageResponse(message="Profesiograma eliminado exitosamente")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"No se puede eliminar el profesiograma: {str(e)}")


@router.post("/{profesiograma_id}/duplicate", response_model=List[ProfesiogramaDuplicateResult])
def duplicate_profesiograma(
    profesiograma_id: int,
    payload: ProfesiogramaDuplicateRequest,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Duplica un profesiograma existente a uno o más cargos destino.
    Crea copias completas incluyendo factores, exámenes, criterios e inmunizaciones.
    """
    # Obtener profesiograma fuente con todas sus relaciones
    source = db.query(Profesiograma).filter(Profesiograma.id == profesiograma_id).first()
    if not source:
        raise HTTPException(status_code=404, detail="Profesiograma fuente no encontrado")

    # Verificar que los cargos destino existen
    target_cargos = db.query(Cargo).filter(Cargo.id.in_(payload.cargo_ids)).all()
    cargo_map = {c.id: c for c in target_cargos}

    missing_cargos = set(payload.cargo_ids) - set(cargo_map.keys())
    if missing_cargos:
        raise HTTPException(
            status_code=400,
            detail=f"Los siguientes cargos no existen: {list(missing_cargos)}"
        )

    results = []

    for cargo_id in payload.cargo_ids:
        cargo = cargo_map[cargo_id]

        # Determinar la siguiente versión para este cargo
        existing_versions = (
            db.query(Profesiograma.version)
            .filter(Profesiograma.cargo_id == cargo_id)
            .all()
        )
        version_numbers = []
        for (v,) in existing_versions:
            try:
                version_numbers.append(float(v))
            except (ValueError, TypeError):
                pass
        next_version = f"{max(version_numbers, default=0) + 1.0:.1f}"

        try:
            # Crear nuevo profesiograma
            new_p = Profesiograma(
                cargo_id=cargo_id,
                version=next_version,
                estado=payload.estado,
                empresa=source.empresa,
                departamento=source.departamento,
                codigo_cargo=cargo.codigo if hasattr(cargo, 'codigo') else source.codigo_cargo,
                numero_trabajadores_expuestos=source.numero_trabajadores_expuestos,
                fecha_elaboracion=None,  # Nueva fecha será establecida al aprobar
                validado_por=None,
                proxima_revision=None,
                elaborado_por=None,
                revisado_por=None,
                aprobado_por=None,
                fecha_aprobacion=None,
                vigencia_meses=source.vigencia_meses,
                posicion_predominante=source.posicion_predominante,
                descripcion_actividades=source.descripcion_actividades,
                periodicidad_emo_meses=source.periodicidad_emo_meses,
                justificacion_periodicidad_emo=source.justificacion_periodicidad_emo,
                fecha_ultima_revision=source.fecha_ultima_revision,
                nivel_riesgo_cargo=source.nivel_riesgo_cargo,
                creado_por=current_user.id,
                modificado_por=None,
            )
            db.add(new_p)
            db.flush()

            # Copiar factores con sus controles e intervenciones
            for src_factor in source.profesiograma_factores:
                new_factor = ProfesiogramaFactor(
                    profesiograma_id=new_p.id,
                    factor_riesgo_id=src_factor.factor_riesgo_id,
                    proceso=src_factor.proceso,
                    actividad=src_factor.actividad,
                    tarea=src_factor.tarea,
                    rutinario=src_factor.rutinario,
                    descripcion_peligro=src_factor.descripcion_peligro,
                    efectos_posibles=src_factor.efectos_posibles,
                    zona_lugar=src_factor.zona_lugar,
                    tipo_peligro=src_factor.tipo_peligro,
                    clasificacion_peligro=src_factor.clasificacion_peligro,
                    controles_existentes=src_factor.controles_existentes,
                    fuente=src_factor.fuente,
                    medio=src_factor.medio,
                    individuo=src_factor.individuo,
                    peor_consecuencia=src_factor.peor_consecuencia,
                    requisito_legal=src_factor.requisito_legal,
                    nivel_exposicion=src_factor.nivel_exposicion,
                    tiempo_exposicion_horas=src_factor.tiempo_exposicion_horas,
                    valor_medido=src_factor.valor_medido,
                    valor_limite_permisible=src_factor.valor_limite_permisible,
                    unidad_medida=src_factor.unidad_medida,
                    nd=src_factor.nd,
                    ne=src_factor.ne,
                    nc=src_factor.nc,
                    eliminacion=src_factor.eliminacion,
                    sustitucion=src_factor.sustitucion,
                    controles_ingenieria=src_factor.controles_ingenieria,
                    controles_administrativos=src_factor.controles_administrativos,
                    senalizacion=src_factor.senalizacion,
                    epp_requerido=src_factor.epp_requerido,
                    registrado_por=current_user.id,
                )
                db.add(new_factor)

                # Copiar controles ESIAE
                for ctrl in src_factor.controles_esiae:
                    db.add(ProfesiogramaControlESIAE(
                        profesiograma_id=new_p.id,
                        factor_riesgo_id=src_factor.factor_riesgo_id,
                        nivel=ctrl.nivel,
                        medida=ctrl.medida,
                        descripcion=ctrl.descripcion,
                        estado_actual=ctrl.estado_actual,
                        meta=ctrl.meta,
                    ))

                # Copiar intervenciones
                for interv in src_factor.intervenciones:
                    db.add(ProfesiogramaIntervencion(
                        profesiograma_id=new_p.id,
                        factor_riesgo_id=src_factor.factor_riesgo_id,
                        tipo_control=interv.tipo_control,
                        descripcion=interv.descripcion,
                        responsable=interv.responsable,
                        plazo=interv.plazo,
                    ))

            # Copiar exámenes
            for src_exam in source.examenes:
                db.add(ProfesiogramaExamen(
                    profesiograma_id=new_p.id,
                    tipo_examen_id=src_exam.tipo_examen_id,
                    tipo_evaluacion=src_exam.tipo_evaluacion,
                    periodicidad_meses=src_exam.periodicidad_meses,
                    justificacion_periodicidad=src_exam.justificacion_periodicidad,
                    obligatorio=src_exam.obligatorio,
                    orden_realizacion=src_exam.orden_realizacion,
                    normativa_base=src_exam.normativa_base,
                ))

            # Copiar inmunizaciones
            for src_inmun in source.inmunizaciones:
                db.add(ProfesiogramaInmunizacion(
                    profesiograma_id=new_p.id,
                    inmunizacion_id=src_inmun.inmunizacion_id,
                ))

            # Copiar criterios de exclusión
            new_p.criterios_exclusion = list(source.criterios_exclusion)

            db.flush()

            results.append(ProfesiogramaDuplicateResult(
                cargo_id=cargo_id,
                cargo_nombre=cargo.nombre_cargo,
                profesiograma_id=new_p.id,
                version=next_version,
                success=True,
                message=f"Profesiograma duplicado exitosamente al cargo '{cargo.nombre_cargo}'"
            ))

        except Exception as e:
            results.append(ProfesiogramaDuplicateResult(
                cargo_id=cargo_id,
                cargo_nombre=cargo.nombre_cargo,
                profesiograma_id=0,
                version="",
                success=False,
                message=f"Error al duplicar: {str(e)}"
            ))

    # Solo hacer commit si al menos uno fue exitoso
    if any(r.success for r in results):
        db.commit()

    return results


@router.post("/cargos/{cargo_id}", response_model=ProfesiogramaSchema, status_code=status.HTTP_201_CREATED)
def create_profesiograma(
    cargo_id: int,
    payload: ProfesiogramaCreate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    cargo = db.query(Cargo).filter(Cargo.id == cargo_id).first()
    if not cargo:
        raise HTTPException(status_code=404, detail="Cargo no encontrado")

    exists = (
        db.query(Profesiograma)
        .filter(Profesiograma.cargo_id == cargo_id, Profesiograma.version == payload.version)
        .first()
    )
    if exists:
        raise HTTPException(status_code=400, detail="Ya existe un profesiograma para ese cargo y versión")

    if payload.estado == ProfesiogramaEstado.ACTIVO:
        db.query(Profesiograma).filter(
            Profesiograma.cargo_id == cargo_id,
            Profesiograma.estado == ProfesiogramaEstado.ACTIVO,
        ).update(
            {
                Profesiograma.estado: ProfesiogramaEstado.INACTIVO,
                Profesiograma.modificado_por: current_user.id,
                Profesiograma.fecha_modificacion: datetime.utcnow(),
            }
        )

    stats = compute_stats_for_cargo(db, cargo_id)
    numero_trabajadores_expuestos = (
        payload.numero_trabajadores_expuestos
        if payload.numero_trabajadores_expuestos is not None
        else stats.total_trabajadores_activos
    )
    justificacion_periodicidad_emo = payload.justificacion_periodicidad_emo
    if payload.periodicidad_emo_meses > 12 and not (justificacion_periodicidad_emo and justificacion_periodicidad_emo.strip()):
        justificacion_periodicidad_emo = generate_justificacion_periodicidad_emo_from_db(
            db,
            cargo_id,
            payload.periodicidad_emo_meses,
            formato="breve",
        )

    p = Profesiograma(
        cargo_id=cargo_id,
        version=payload.version,
        estado=payload.estado,
        empresa=payload.empresa,
        departamento=payload.departamento,
        codigo_cargo=payload.codigo_cargo,
        numero_trabajadores_expuestos=numero_trabajadores_expuestos,
        fecha_elaboracion=payload.fecha_elaboracion,
        validado_por=payload.validado_por,
        proxima_revision=payload.proxima_revision,
        elaborado_por=payload.elaborado_por,
        revisado_por=payload.revisado_por,
        aprobado_por=payload.aprobado_por,
        fecha_aprobacion=payload.fecha_aprobacion,
        vigencia_meses=payload.vigencia_meses,
        posicion_predominante=payload.posicion_predominante,
        descripcion_actividades=payload.descripcion_actividades,
        periodicidad_emo_meses=payload.periodicidad_emo_meses,
        justificacion_periodicidad_emo=justificacion_periodicidad_emo,
        fecha_ultima_revision=payload.fecha_ultima_revision,
        nivel_riesgo_cargo=payload.nivel_riesgo_cargo,
        creado_por=current_user.id,
        modificado_por=None,
    )
    db.add(p)
    
    # Actualizar la periodicidad del cargo
    if payload.periodicidad_emo_meses:
        periodicidad_map = {
            6: "semestral",
            12: "anual",
            24: "bianual",
            36: "trianual"
        }
        cargo.periodicidad_emo = periodicidad_map.get(payload.periodicidad_emo_meses, "anual")
        db.add(cargo)
        
    db.flush()

    unidad_por_factor_id = {}
    if payload.factores:
        factor_ids = [f.factor_riesgo_id for f in payload.factores]
        factors = db.query(FactorRiesgo).filter(FactorRiesgo.id.in_(factor_ids)).all()
        for factor in factors:
            unidad_por_factor_id[factor.id] = getattr(factor, "unidad_medida", None) or getattr(
                factor, "simbolo_unidad", None
            )

    for f in payload.factores:
        unidad_final = f.unidad_medida
        if unidad_final is None or (isinstance(unidad_final, str) and not unidad_final.strip()):
            unidad_final = unidad_por_factor_id.get(f.factor_riesgo_id)
        assoc = ProfesiogramaFactor(
            profesiograma_id=p.id,
            factor_riesgo_id=f.factor_riesgo_id,
            proceso=f.proceso,
            actividad=f.actividad,
            tarea=f.tarea,
            rutinario=f.rutinario,
            descripcion_peligro=f.descripcion_peligro,
            efectos_posibles=f.efectos_posibles,
            zona_lugar=f.zona_lugar,
            tipo_peligro=f.tipo_peligro,
            clasificacion_peligro=f.clasificacion_peligro,
            controles_existentes=f.controles_existentes,
            fuente=f.fuente,
            medio=f.medio,
            individuo=f.individuo,
            peor_consecuencia=f.peor_consecuencia,
            requisito_legal=f.requisito_legal,
            nivel_exposicion=f.nivel_exposicion,
            tiempo_exposicion_horas=f.tiempo_exposicion_horas,
            valor_medido=f.valor_medido,
            valor_limite_permisible=f.valor_limite_permisible,
            unidad_medida=unidad_final,
            nd=f.nd,
            ne=f.ne,
            nc=f.nc,
            eliminacion=f.eliminacion,
            sustitucion=f.sustitucion,
            controles_ingenieria=f.controles_ingenieria,
            controles_administrativos=f.controles_administrativos,
            senalizacion=f.senalizacion,
            epp_requerido=f.epp_requerido,
            registrado_por=current_user.id,
        )
        db.add(assoc)
        for c in getattr(f, "controles_esiae", []) or []:
            db.add(
                ProfesiogramaControlESIAE(
                    profesiograma_id=p.id,
                    factor_riesgo_id=f.factor_riesgo_id,
                    nivel=c.nivel,
                    medida=c.medida,
                    descripcion=c.descripcion,
                    estado_actual=c.estado_actual,
                    meta=c.meta,
                )
            )
        for i in getattr(f, "intervenciones", []) or []:
            db.add(
                ProfesiogramaIntervencion(
                    profesiograma_id=p.id,
                    factor_riesgo_id=f.factor_riesgo_id,
                    tipo_control=i.tipo_control,
                    descripcion=i.descripcion,
                    responsable=i.responsable,
                    plazo=i.plazo,
                )
            )

    for e in payload.examenes:
        assoc_exam = ProfesiogramaExamen(
            profesiograma_id=p.id,
            tipo_examen_id=e.tipo_examen_id,
            tipo_evaluacion=e.tipo_evaluacion,
            periodicidad_meses=e.periodicidad_meses,
            justificacion_periodicidad=e.justificacion_periodicidad,
            obligatorio=e.obligatorio,
            orden_realizacion=e.orden_realizacion,
            normativa_base=e.normativa_base,
        )
        db.add(assoc_exam)

    for i in payload.inmunizaciones:
        inmun = db.query(Inmunizacion).filter(Inmunizacion.id == i.inmunizacion_id).first()
        if not inmun:
            raise HTTPException(status_code=400, detail=f"Inmunización no existe: {i.inmunizacion_id}")
        assoc_inmun = ProfesiogramaInmunizacion(
            profesiograma_id=p.id,
            inmunizacion_id=i.inmunizacion_id,
        )
        db.add(assoc_inmun)

    if payload.criterios_exclusion_ids:
        criterios = (
            db.query(CriterioExclusion)
            .filter(CriterioExclusion.id.in_(payload.criterios_exclusion_ids))
            .all()
        )
        if len(criterios) != len(set(payload.criterios_exclusion_ids)):
            raise HTTPException(status_code=400, detail="Uno o más criterios_exclusion_ids no existen")
        p.criterios_exclusion = criterios

    db.commit()
    db.refresh(p)
    return _serialize_profesiograma(p)


@router.put("/{profesiograma_id}", response_model=ProfesiogramaSchema)
def update_profesiograma(
    profesiograma_id: int,
    payload: ProfesiogramaUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    p = db.query(Profesiograma).filter(Profesiograma.id == profesiograma_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Profesiograma no encontrado")

    data = payload.model_dump(exclude_unset=True)
    stats = None
    periodicidad_emo = data.get("periodicidad_emo_meses")
    if periodicidad_emo is not None and periodicidad_emo > 12:
        just = data.get("justificacion_periodicidad_emo")
        if not (just and str(just).strip()):
            data["justificacion_periodicidad_emo"] = generate_justificacion_periodicidad_emo_from_db(
                db,
                p.cargo_id,
                periodicidad_emo,
                formato="breve",
            )
            
    # Actualizar la periodicidad del cargo si cambia en el profesiograma
    if periodicidad_emo:
        cargo = db.query(Cargo).filter(Cargo.id == p.cargo_id).first()
        if cargo:
            periodicidad_map = {
                6: "semestral",
                12: "anual",
                24: "bianual",
                36: "trianual"
            }
            cargo.periodicidad_emo = periodicidad_map.get(periodicidad_emo, "anual")
            db.add(cargo)
    if "numero_trabajadores_expuestos" in data and data.get("numero_trabajadores_expuestos") is None:
        stats = stats or compute_stats_for_cargo(db, p.cargo_id)
        data["numero_trabajadores_expuestos"] = stats.total_trabajadores_activos

    if "version" in data:
        exists = (
            db.query(Profesiograma)
            .filter(
                Profesiograma.cargo_id == p.cargo_id,
                Profesiograma.version == data["version"],
                Profesiograma.id != p.id,
            )
            .first()
        )
        if exists:
            raise HTTPException(status_code=400, detail="Ya existe un profesiograma para ese cargo y versión")

    factores = payload.factores if "factores" in data else None
    examenes = payload.examenes if "examenes" in data else None
    inmunizaciones = payload.inmunizaciones if "inmunizaciones" in data else None
    criterios_ids = payload.criterios_exclusion_ids if "criterios_exclusion_ids" in data else None

    data.pop("factores", None)
    data.pop("examenes", None)
    data.pop("inmunizaciones", None)
    data.pop("criterios_exclusion_ids", None)

    for k, v in data.items():
        setattr(p, k, v)
    p.modificado_por = current_user.id

    if factores is not None:
        unidad_por_factor_id = {}
        if factores:
            factor_ids = []
            for f in factores:
                f_data = f.model_dump() if hasattr(f, "model_dump") else f
                factor_id = f_data.get("factor_riesgo_id")
                if factor_id is not None:
                    factor_ids.append(factor_id)
            if factor_ids:
                factors = db.query(FactorRiesgo).filter(FactorRiesgo.id.in_(factor_ids)).all()
                for factor in factors:
                    unidad_por_factor_id[factor.id] = getattr(factor, "unidad_medida", None) or getattr(
                        factor, "simbolo_unidad", None
                    )

        db.query(ProfesiogramaFactor).filter(ProfesiogramaFactor.profesiograma_id == p.id).delete()
        for f in factores:
            f_data = f.model_dump() if hasattr(f, "model_dump") else f
            factor_id = f_data.get("factor_riesgo_id")
            unidad_final = f_data.get("unidad_medida")
            if unidad_final is None or (isinstance(unidad_final, str) and not str(unidad_final).strip()):
                unidad_final = unidad_por_factor_id.get(factor_id)
            assoc = ProfesiogramaFactor(
                profesiograma_id=p.id,
                factor_riesgo_id=factor_id,
                proceso=f_data.get("proceso"),
                actividad=f_data.get("actividad"),
                tarea=f_data.get("tarea"),
                rutinario=f_data.get("rutinario"),
                descripcion_peligro=f_data.get("descripcion_peligro"),
                efectos_posibles=f_data.get("efectos_posibles"),
                zona_lugar=f_data.get("zona_lugar"),
                tipo_peligro=f_data.get("tipo_peligro"),
                clasificacion_peligro=f_data.get("clasificacion_peligro"),
                controles_existentes=f_data.get("controles_existentes"),
                fuente=f_data.get("fuente"),
                medio=f_data.get("medio"),
                individuo=f_data.get("individuo"),
                peor_consecuencia=f_data.get("peor_consecuencia"),
                requisito_legal=f_data.get("requisito_legal"),
                nivel_exposicion=f_data.get("nivel_exposicion"),
                tiempo_exposicion_horas=f_data.get("tiempo_exposicion_horas"),
                valor_medido=f_data.get("valor_medido"),
                valor_limite_permisible=f_data.get("valor_limite_permisible"),
                unidad_medida=unidad_final,
                nd=f_data.get("nd"),
                ne=f_data.get("ne"),
                nc=f_data.get("nc"),
                eliminacion=f_data.get("eliminacion"),
                sustitucion=f_data.get("sustitucion"),
                controles_ingenieria=f_data.get("controles_ingenieria"),
                controles_administrativos=f_data.get("controles_administrativos"),
                senalizacion=f_data.get("senalizacion"),
                epp_requerido=f_data.get("epp_requerido"),
                registrado_por=current_user.id,
            )
            db.add(assoc)
            for c in f_data.get("controles_esiae") or []:
                db.add(
                    ProfesiogramaControlESIAE(
                        profesiograma_id=p.id,
                        factor_riesgo_id=factor_id,
                        nivel=c.get("nivel"),
                        medida=c.get("medida"),
                        descripcion=c.get("descripcion"),
                        estado_actual=c.get("estado_actual"),
                        meta=c.get("meta"),
                    )
                )
            for i in f_data.get("intervenciones") or []:
                db.add(
                    ProfesiogramaIntervencion(
                        profesiograma_id=p.id,
                        factor_riesgo_id=factor_id,
                        tipo_control=i.get("tipo_control"),
                        descripcion=i.get("descripcion"),
                        responsable=i.get("responsable"),
                        plazo=i.get("plazo"),
                    )
                )

    if examenes is not None:
        db.query(ProfesiogramaExamen).filter(ProfesiogramaExamen.profesiograma_id == p.id).delete()
        for e in examenes:
            e_data = e.model_dump() if hasattr(e, "model_dump") else e
            assoc_exam = ProfesiogramaExamen(
                profesiograma_id=p.id,
                tipo_examen_id=e_data.get("tipo_examen_id"),
                tipo_evaluacion=e_data.get("tipo_evaluacion"),
                periodicidad_meses=e_data.get("periodicidad_meses"),
                justificacion_periodicidad=e_data.get("justificacion_periodicidad"),
                obligatorio=e_data.get("obligatorio"),
                orden_realizacion=e_data.get("orden_realizacion"),
                normativa_base=e_data.get("normativa_base"),
            )
            db.add(assoc_exam)

    if inmunizaciones is not None:
        db.query(ProfesiogramaInmunizacion).filter(ProfesiogramaInmunizacion.profesiograma_id == p.id).delete()
        for i in inmunizaciones:
            i_data = i.model_dump() if hasattr(i, "model_dump") else i
            inmun_id = i_data.get("inmunizacion_id")
            inmun = db.query(Inmunizacion).filter(Inmunizacion.id == inmun_id).first()
            if not inmun:
                raise HTTPException(status_code=400, detail=f"Inmunización no existe: {inmun_id}")
            assoc_inmun = ProfesiogramaInmunizacion(
                profesiograma_id=p.id,
                inmunizacion_id=inmun_id,
            )
            db.add(assoc_inmun)

    if criterios_ids is not None:
        criterios = []
        if criterios_ids:
            criterios = db.query(CriterioExclusion).filter(CriterioExclusion.id.in_(criterios_ids)).all()
            if len(criterios) != len(set(criterios_ids)):
                raise HTTPException(status_code=400, detail="Uno o más criterios_exclusion_ids no existen")
        p.criterios_exclusion = criterios

    db.commit()
    db.refresh(p)
    return _serialize_profesiograma(p)


@router.get("/cargos/{cargo_id}/export/profesiograma.pdf")
def export_profesiograma_pdf_by_cargo(
    cargo_id: int,
    download: bool = Query(True),
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    cargo = db.query(Cargo).filter(Cargo.id == cargo_id).first()
    if not cargo:
        raise HTTPException(status_code=404, detail="Cargo no encontrado")

    p = (
        db.query(Profesiograma)
        .filter(Profesiograma.cargo_id == cargo_id, Profesiograma.estado == ProfesiogramaEstado.ACTIVO)
        .order_by(Profesiograma.fecha_creacion.desc())
        .first()
    )
    if not p:
        p = (
            db.query(Profesiograma)
            .filter(Profesiograma.cargo_id == cargo_id)
            .order_by(Profesiograma.fecha_creacion.desc())
            .first()
        )
    if not p:
        raise HTTPException(status_code=404, detail="No hay profesiogramas para ese cargo")

    pdf_bytes = generate_profesiograma_report_pdf(cargo, p)

    cargo_name = getattr(cargo, "nombre_cargo", "") or str(cargo_id)
    safe_cargo = "".join(ch if ch.isalnum() else "_" for ch in cargo_name).strip("_") or f"cargo_{cargo_id}"
    safe_version = "".join(ch if ch.isalnum() else "_" for ch in (p.version or "1_0")).strip("_")
    filename = f"Profesiograma_{safe_cargo}_{safe_version}_{datetime.now().strftime('%Y%m%d')}.pdf"

    disposition = "attachment" if download else "inline"
    headers = {"Content-Disposition": f'{disposition}; filename="{filename}"'}
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@router.get("/{profesiograma_id}/export/matriz.xlsx")
def export_profesiograma_matriz_excel(
    profesiograma_id: int,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    p = db.query(Profesiograma).filter(Profesiograma.id == profesiograma_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Profesiograma no encontrado")

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Matriz"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Cargo",
        "Versión",
        "Estado",
        "Factor",
        "Categoría",
        "Zona/Lugar",
        "Tipo",
        "Clasificación",
        "Proceso",
        "Actividad",
        "Tarea",
        "¿Rutinario?",
        "Descripción del peligro",
        "Efectos posibles",
        "Controles existentes",
        "Fuente",
        "Medio",
        "Individuo",
        "ND",
        "NE",
        "NP",
        "Interpretación NP",
        "NC",
        "NR",
        "Nivel Riesgo",
        "Color",
        "Acción",
        "Aceptabilidad",
        "Tiempo exposición (h/día)",
        "Valor medido",
        "Valor límite permisible",
        "Unidad",
        "Peor consecuencia",
        "Requisito legal",
        "Eliminación",
        "Sustitución",
        "Controles ingeniería",
        "Controles administrativos",
        "Señalización",
        "EPP requerido",
        "Registrado por",
        "Fecha registro",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    cargo_name = getattr(getattr(p, "cargo", None), "nombre", None) or str(p.cargo_id)

    for row_idx, pf in enumerate(list(getattr(p, "profesiograma_factores", []) or []), 2):
        factor = getattr(pf, "factor_riesgo", None)
        row_values = [
            cargo_name,
            p.version,
            p.estado.value if hasattr(p.estado, "value") else p.estado,
            getattr(factor, "nombre", None) or str(pf.factor_riesgo_id),
            getattr(factor, "categoria", None) or "",
            pf.zona_lugar or "",
            pf.tipo_peligro or "",
            pf.clasificacion_peligro or "",
            pf.proceso or "",
            pf.actividad or "",
            pf.tarea or "",
            "SÍ" if pf.rutinario else "NO",
            pf.descripcion_peligro or "",
            pf.efectos_posibles or "",
            pf.controles_existentes or "",
            pf.fuente or "",
            pf.medio or "",
            pf.individuo or "",
            pf.nd,
            pf.ne,
            pf.np,
            pf.interpretacion_np,
            pf.nc,
            pf.nr,
            pf.nivel_riesgo,
            pf.color_riesgo,
            pf.accion_riesgo,
            pf.aceptabilidad,
            float(pf.tiempo_exposicion_horas) if pf.tiempo_exposicion_horas is not None else None,
            pf.valor_medido or "",  # Ahora es string (puede ser numérico o texto)
            pf.valor_limite_permisible or "",  # Ahora es string (puede ser numérico o texto)
            pf.unidad_medida or "",
            pf.peor_consecuencia or "",
            pf.requisito_legal or "",
            pf.eliminacion or "",
            pf.sustitucion or "",
            pf.controles_ingenieria or "",
            pf.controles_administrativos or "",
            pf.senalizacion or "",
            pf.epp_requerido or "",
            pf.registrado_por,
            pf.fecha_registro.strftime("%Y-%m-%d %H:%M:%S") if pf.fecha_registro else "",
        ]
        for col_idx, value in enumerate(row_values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws_esiae = workbook.create_sheet("ESIAE")
    esiae_headers = [
        "Cargo",
        "Versión",
        "Factor",
        "Nivel",
        "Medida",
        "Descripción",
        "Estado actual",
        "Meta",
    ]
    for col, header in enumerate(esiae_headers, 1):
        cell = ws_esiae.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row_esiae = 2
    for pf in list(getattr(p, "profesiograma_factores", []) or []):
        factor = getattr(pf, "factor_riesgo", None)
        for c in list(getattr(pf, "controles_esiae", []) or []):
            values = [
                cargo_name,
                p.version,
                getattr(factor, "nombre", None) or str(pf.factor_riesgo_id),
                c.nivel,
                c.medida or "",
                c.descripcion or "",
                c.estado_actual or "",
                c.meta or "",
            ]
            for col_idx, value in enumerate(values, 1):
                ws_esiae.cell(row=row_esiae, column=col_idx, value=value)
            row_esiae += 1

    ws_int = workbook.create_sheet("Intervenciones")
    int_headers = [
        "Cargo",
        "Versión",
        "Factor",
        "Tipo de control",
        "Descripción",
        "Responsable",
        "Plazo",
    ]
    for col, header in enumerate(int_headers, 1):
        cell = ws_int.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row_int = 2
    for pf in list(getattr(p, "profesiograma_factores", []) or []):
        factor = getattr(pf, "factor_riesgo", None)
        for i in list(getattr(pf, "intervenciones", []) or []):
            values = [
                cargo_name,
                p.version,
                getattr(factor, "nombre", None) or str(pf.factor_riesgo_id),
                i.tipo_control,
                i.descripcion or "",
                i.responsable or "",
                i.plazo or "",
            ]
            for col_idx, value in enumerate(values, 1):
                ws_int.cell(row=row_int, column=col_idx, value=value)
            row_int += 1

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"matriz_gtc45_profesiograma_{profesiograma_id}_{timestamp}.xlsx"
    return StreamingResponse(
        BytesIO(buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _render_matriz_gtc45_markdown(p: Profesiograma) -> str:
    def md_table(headers: list[str], rows: list[list[str]]) -> str:
        lines = []
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join(["---"] * len(headers)) + "|")
        for r in rows:
            lines.append("| " + " | ".join(r) + " |")
        return "\n".join(lines)

    def nd_label(nd: int | None) -> str:
        mapping = {10: "MUY ALTO", 6: "ALTO", 2: "MEDIO", 0: "BAJO", -10: "BAJO"}
        return mapping.get(nd, "")

    def ne_label(ne: int | None) -> str:
        mapping = {4: "CONTINUO", 3: "FRECUENTE", 2: "OCASIONAL", 1: "ESPORÁDICO"}
        return mapping.get(ne, "")

    def nc_label(nc: int | None) -> str:
        mapping = {100: "MUY GRAVE", 60: "GRAVE", 25: "MODERADA", 10: "LEVE"}
        return mapping.get(nc, "")

    def si_no(value: bool | None) -> str:
        if value is None:
            return ""
        return "SÍ" if value else "NO"

    def clean(value) -> str:
        if value is None:
            return ""
        return str(value).replace("\r", "").strip()

    cargo_name = (
        getattr(getattr(p, "cargo", None), "nombre_cargo", None)
        or getattr(getattr(p, "cargo", None), "nombre", None)
        or str(p.cargo_id)
    )

    empresa_line = clean(p.empresa) if p.empresa else ""
    header_empresa = f"### Empresa: {empresa_line} - Evaluación Completa" if empresa_line else "### Empresa: - Evaluación Completa"

    lines: list[str] = []
    lines.append("# Matriz de Identificación de Peligros - GTC 45")
    lines.append(f"## Cargo: {cargo_name}")
    lines.append(header_empresa)
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## Información General")
    lines.append("")
    info_rows = [
        ["**Código del Cargo**", clean(p.codigo_cargo)],
        ["**Departamento**", clean(p.departamento)],
        ["**Número de Trabajadores Expuestos**", clean(p.numero_trabajadores_expuestos)],
        ["**Fecha de Elaboración**", clean(p.fecha_elaboracion)],
        ["**Validado por**", clean(p.validado_por)],
        ["**Próxima Revisión**", clean(p.proxima_revision)],
        ["**Versión**", clean(p.version)],
    ]
    lines.append(md_table(["Campo", "Descripción"], info_rows))
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## Escala de Valoración - Metodología GTC 45")
    lines.append("")
    lines.append("### Nivel de Deficiencia (ND)")
    lines.append("")
    lines.append(
        md_table(
            ["Nivel", "Valor", "Descripción"],
            [
                ["**Muy Alto**", "10", "Control inexistente o totalmente inadecuado"],
                ["**Alto**", "6", "Control presente pero deficiente"],
                ["**Medio**", "2", "Control presente aunque incompleto"],
                ["**Bajo**", "-10", "Controles establecidos y efectivos"],
            ],
        )
    )
    lines.append("")
    lines.append("### Nivel de Exposición (NE)")
    lines.append("")
    lines.append(
        md_table(
            ["Nivel", "Valor", "Descripción", "Frecuencia"],
            [
                ["**Continuo**", "4", "Exposición permanente toda la jornada", "8 horas/día"],
                ["**Frecuente**", "3", "Exposición diaria con tiempo limitado", "2-4 horas/día"],
                ["**Ocasional**", "2", "Exposición 1-2 veces por semana", "1-2 horas"],
                ["**Esporádico**", "1", "Exposición menor a 1 vez/mes", "<1 hora"],
            ],
        )
    )
    lines.append("")
    lines.append("### Nivel de Probabilidad (NP = ND × NE)")
    lines.append("")
    lines.append(
        md_table(
            ["Rango NP", "Interpretación", "Descripción"],
            [
                ["**24-100**", "**MUY ALTO**", "Evento ocurrirá sin dudas si no se interviene"],
                ["**10-23**", "**ALTO**", "Evento ocurrirá frecuentemente sin intervención"],
                ["**4-9**", "**MEDIO**", "Evento podría ocurrir bajo ciertas condiciones"],
                ["**1-3**", "**BAJO**", "Evento difícilmente ocurrirá"],
            ],
        )
    )
    lines.append("")
    lines.append("### Nivel de Consecuencia (NC)")
    lines.append("")
    lines.append(
        md_table(
            ["Nivel", "Valor", "Descripción", "Ejemplo"],
            [
                ["**Muy Grave**", "100", "Muerte, incapacidad permanente", "Muerte, invalidez total"],
                ["**Grave**", "60", "Incapacidad temporal >30 días o invalidez parcial", "Fractura exposición, quemadura severa"],
                ["**Moderada**", "25", "Incapacidad temporal 15-30 días", "Distensión muscular, contusión"],
                ["**Leve**", "10", "Incapacidad temporal <15 días", "Herida superficial, irritación"],
            ],
        )
    )
    lines.append("")
    lines.append("### Nivel de Riesgo (NR = NP × NC)")
    lines.append("")
    lines.append(
        md_table(
            ["Rango NR", "Nivel Riesgo", "Color", "Acción"],
            [
                ["**600-4800**", "**CRÍTICO**", "Rojo Intenso", "Suspender actividad inmediatamente"],
                ["**200-599**", "**ALTO**", "Rojo", "Intervención inmediata"],
                ["**50-199**", "**MEDIO**", "Amarillo", "Intervención a corto plazo (1-3 meses)"],
                ["**1-49**", "**BAJO**", "Verde", "Monitoreo y vigilancia periódica"],
            ],
        )
    )
    lines.append("")
    lines.append("### Aceptabilidad del Riesgo")
    lines.append("")
    lines.append(
        md_table(
            ["Nivel de Riesgo", "Aceptabilidad", "Acción Requerida"],
            [
                ["**Crítico (>600)**", "NO ACEPTABLE", "Suspender actividad, intervención inmediata"],
                ["**Alto (200-599)**", "NO ACEPTABLE", "Corrección urgente (1-2 semanas)"],
                ["**Medio (50-199)**", "CONDICIONALMENTE ACEPTABLE", "Mejoras en corto plazo"],
                ["**Bajo (1-49)**", "ACEPTABLE", "Mantener vigilancia"],
            ],
        )
    )
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append(f"## Matriz Completa de Peligros - {cargo_name}")
    lines.append("")

    factores = list(getattr(p, "profesiograma_factores", []) or [])
    for idx, pf in enumerate(factores, 1):
        factor = getattr(pf, "factor_riesgo", None)
        titulo = (getattr(factor, "nombre", None) or clean(pf.descripcion_peligro) or f"Factor {pf.factor_riesgo_id}").upper()
        lines.append(f"### PELIGRO {idx}: {titulo}")
        lines.append("")
        lines.append("#### Información del Proceso")
        lines.append("")
        lines.append(
            md_table(
                ["Aspecto", "Descripción"],
                [
                    ["**Proceso**", clean(pf.proceso)],
                    ["**Zona/Lugar**", clean(pf.zona_lugar)],
                    ["**Actividades**", clean(pf.actividad)],
                    ["**Tareas**", clean(pf.tarea)],
                    ["**¿Rutinario?**", si_no(pf.rutinario)],
                ],
            )
        )
        lines.append("")
        lines.append("#### Identificación del Peligro")
        lines.append("")
        lines.append(
            md_table(
                ["Aspecto", "Descripción"],
                [
                    ["**Peligro**", clean(getattr(factor, "nombre", None) or "")],
                    ["**Tipo**", clean(pf.tipo_peligro)],
                    ["**Descripción**", clean(pf.descripcion_peligro)],
                    ["**Efectos Posibles**", clean(pf.efectos_posibles)],
                    ["**Clasificación**", clean(pf.clasificacion_peligro)],
                ],
            )
        )
        lines.append("")
        lines.append("#### Análisis de Controles Existentes")
        lines.append("")
        lines.append(
            md_table(
                ["Aspecto", "Descripción"],
                [
                    ["**Controles Existentes**", clean(pf.controles_existentes)],
                    ["**Fuente**", clean(pf.fuente)],
                    ["**Medio**", clean(pf.medio)],
                    ["**Individuo**", clean(pf.individuo)],
                ],
            )
        )
        lines.append("")
        lines.append("#### Evaluación del Riesgo - Cálculo")
        lines.append("")
        lines.append(
            md_table(
                ["Factor", "Nivel", "Valor", "Justificación"],
                [
                    ["**ND (Nivel de Deficiencia)**", nd_label(pf.nd), clean(pf.nd), ""],
                    ["**NE (Nivel de Exposición)**", ne_label(pf.ne), clean(pf.ne), ""],
                    ["**NP (Nivel Probabilidad)**", clean(pf.interpretacion_np), clean(pf.np), ""],
                    ["**NC (Nivel Consecuencia)**", nc_label(pf.nc), clean(pf.nc), ""],
                    ["**NR (Nivel de Riesgo)**", clean(pf.nivel_riesgo), clean(pf.nr), clean(pf.color_riesgo)],
                ],
            )
        )
        lines.append("")
        lines.append("#### Aceptabilidad del Riesgo")
        lines.append("")
        lines.append(
            md_table(
                ["Aspecto", "Resultado"],
                [
                    ["**Aceptabilidad**", clean(pf.aceptabilidad)],
                    ["**Nro Expuestos**", clean(p.numero_trabajadores_expuestos)],
                    ["**Peor Consecuencia**", clean(pf.peor_consecuencia)],
                    ["**Requisito Legal Específico**", clean(pf.requisito_legal)],
                ],
            )
        )
        lines.append("")
        lines.append("#### Criterios para Establecer Controles (Jerarquía ESIAE)")
        lines.append("")
        esiae_rows: list[list[str]] = []
        controls = list(getattr(pf, "controles_esiae", []) or [])
        if controls:
            for c in controls:
                esiae_rows.append(
                    [
                        f"**{clean(c.nivel)}**",
                        clean(c.medida),
                        clean(c.descripcion),
                        clean(c.estado_actual),
                        clean(c.meta),
                    ]
                )
        else:
            esiae_rows = [
                ["**E - Eliminación**", "", clean(pf.eliminacion), "", ""],
                ["**S - Sustitución**", "", clean(pf.sustitucion), "", ""],
                ["**I - Ingeniería**", "", clean(pf.controles_ingenieria), "", ""],
                ["**A - Administrativa**", "", clean(pf.controles_administrativos), "", ""],
                ["**I - Ingeniería**", "Señalización", clean(pf.senalizacion), "", ""],
                ["**EPP**", "", clean(pf.epp_requerido), "", ""],
            ]
        lines.append(md_table(["Nivel", "Medida", "Descripción", "Estado Actual", "Meta"], esiae_rows))
        lines.append("")
        lines.append("#### Medidas de Intervención")
        lines.append("")
        interv_rows: list[list[str]] = []
        interv = list(getattr(pf, "intervenciones", []) or [])
        for i in interv:
            interv_rows.append([f"**{clean(i.tipo_control)}**", clean(i.descripcion), clean(i.responsable), clean(i.plazo)])
        if not interv_rows:
            interv_rows.append(["", "", "", ""])
        lines.append(md_table(["Tipo de Control", "Descripción", "Responsable", "Plazo"], interv_rows))
        lines.append("")
        lines.append("---")
        lines.append("")

    lines.append("## Resumen Ejecutivo de Riesgos")
    lines.append("")
    altos = [(i + 1, pf) for i, pf in enumerate(factores) if pf.nivel_riesgo == "ALTO"]
    moderados = [(i + 1, pf) for i, pf in enumerate(factores) if pf.nivel_riesgo in ("MEDIO-ALTO", "MEDIO")]
    lines.append("### Riesgos Críticos Identificados")
    lines.append("")
    crit_rows = []
    for n, pf in altos:
        factor = getattr(pf, "factor_riesgo", None)
        crit_rows.append(
            [
                str(n),
                clean(getattr(factor, "nombre", None) or f"Factor {pf.factor_riesgo_id}"),
                clean(pf.nivel_riesgo),
                clean(pf.nr),
                clean(pf.aceptabilidad),
                clean(pf.accion_riesgo),
            ]
        )
    if not crit_rows:
        crit_rows.append(["", "", "", "", "", ""])
    lines.append(md_table(["#", "Peligro", "Nivel Riesgo", "NR", "Estado", "Acción Inmediata"], crit_rows))
    lines.append("")
    lines.append("### Riesgos Moderados Identificados")
    lines.append("")
    mod_rows = []
    for n, pf in moderados:
        factor = getattr(pf, "factor_riesgo", None)
        mod_rows.append(
            [
                str(n),
                clean(getattr(factor, "nombre", None) or f"Factor {pf.factor_riesgo_id}"),
                clean(pf.nivel_riesgo),
                clean(pf.nr),
                clean(pf.aceptabilidad),
                clean(pf.accion_riesgo),
            ]
        )
    if not mod_rows:
        mod_rows.append(["", "", "", "", "", ""])
    lines.append(md_table(["#", "Peligro", "Nivel Riesgo", "NR", "Estado", "Acción"], mod_rows))
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## Validación del Documento")
    lines.append("")
    lines.append(f"**Elaborado por:** {clean(p.elaborado_por)}  ")
    lines.append(f"**Revisado por:** {clean(p.revisado_por)}  ")
    lines.append(f"**Aprobado por:** {clean(p.aprobado_por)}  ")
    lines.append(f"**Fecha de Aprobación:** {clean(p.fecha_aprobacion)}  ")
    lines.append(f"**Vigencia:** {clean(p.vigencia_meses)} meses (próxima revisión: {clean(p.proxima_revision)})")
    lines.append("")
    return "\n".join(lines).strip() + "\n"


@router.get("/{profesiograma_id}/export/matriz.md")
def export_profesiograma_matriz_markdown(
    profesiograma_id: int,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    p = db.query(Profesiograma).filter(Profesiograma.id == profesiograma_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Profesiograma no encontrado")

    content = _render_matriz_gtc45_markdown(p)
    buffer = BytesIO(content.encode("utf-8"))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"matriz_gtc45_profesiograma_{profesiograma_id}_{timestamp}.md"
    return StreamingResponse(
        buffer,
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
