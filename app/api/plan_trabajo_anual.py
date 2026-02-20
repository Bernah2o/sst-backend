import io
from datetime import datetime
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import weasyprint

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user, require_supervisor_or_admin
from app.models.user import User
from app.models.plan_trabajo_anual import (
    PlanTrabajoAnual, PlanTrabajoActividad, PlanTrabajoSeguimiento,
    CicloPhva, CategoriaActividad, EstadoPlan
)
from app.schemas.plan_trabajo_anual import (
    PlanTrabajoAnualCreate, PlanTrabajoAnualUpdate, PlanTrabajoAnualResponse,
    PlanTrabajoAnualDetailResponse,
    PlanTrabajoActividadCreate, PlanTrabajoActividadUpdate, PlanTrabajoActividadResponse,
    PlanTrabajoSeguimientoUpdate, PlanTrabajoSeguimientoResponse,
    DashboardIndicadores, MesIndicador
)
from app.services.plan_trabajo_template import get_plantilla_actividades

router = APIRouter()

NOMBRE_MESES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]


# =====================================================================
# PLANES
# =====================================================================

@router.get("/", response_model=List[PlanTrabajoAnualResponse])
def listar_planes(
    año: Optional[int] = Query(None),
    empresa_id: Optional[int] = Query(None),
    estado: Optional[EstadoPlan] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(PlanTrabajoAnual)
    if año:
        query = query.filter(PlanTrabajoAnual.año == año)
    if empresa_id:
        query = query.filter(PlanTrabajoAnual.empresa_id == empresa_id)
    if estado:
        query = query.filter(PlanTrabajoAnual.estado == estado)
    return query.order_by(PlanTrabajoAnual.año.desc()).all()


@router.post("/", response_model=PlanTrabajoAnualResponse, status_code=status.HTTP_201_CREATED)
def crear_plan(
    plan_data: PlanTrabajoAnualCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    # Verificar que no exista ya un plan para ese año y empresa
    existente = db.query(PlanTrabajoAnual).filter(
        PlanTrabajoAnual.año == plan_data.año,
        PlanTrabajoAnual.empresa_id == plan_data.empresa_id
    ).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Ya existe un Plan de Trabajo Anual para el año {plan_data.año}"
        )

    plan = PlanTrabajoAnual(
        **plan_data.model_dump(),
        created_by=current_user.id
    )
    db.add(plan)
    db.commit()
    db.refresh(plan)
    return plan


@router.post("/crear-desde-plantilla", response_model=PlanTrabajoAnualDetailResponse, status_code=status.HTTP_201_CREATED)
def crear_plan_desde_plantilla(
    año: int = Query(..., description="Año del plan de trabajo"),
    empresa_id: Optional[int] = Query(None),
    encargado_sgsst: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    # Verificar que no exista ya un plan para ese año y empresa
    existente = db.query(PlanTrabajoAnual).filter(
        PlanTrabajoAnual.año == año,
        PlanTrabajoAnual.empresa_id == empresa_id
    ).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Ya existe un Plan de Trabajo Anual para el año {año}"
        )

    # Crear el plan padre
    plan = PlanTrabajoAnual(
        año=año,
        empresa_id=empresa_id,
        encargado_sgsst=encargado_sgsst,
        created_by=current_user.id
    )
    db.add(plan)
    db.flush()

    # Crear las actividades desde la plantilla
    plantilla = get_plantilla_actividades(año)
    for act_data in plantilla:
        meses_programados = act_data.pop("meses_programados", [])
        costo = act_data.pop("costo", None)

        actividad = PlanTrabajoActividad(
            plan_id=plan.id,
            ciclo=CicloPhva(act_data["ciclo"]),
            categoria=CategoriaActividad(act_data["categoria"]),
            estandar=act_data.get("estandar"),
            descripcion=act_data["descripcion"],
            frecuencia=act_data.get("frecuencia"),
            responsable=act_data.get("responsable"),
            recurso_financiero=act_data.get("recurso_financiero", False),
            recurso_tecnico=act_data.get("recurso_tecnico", False),
            costo=costo,
            orden=act_data.get("orden", 0),
        )
        db.add(actividad)
        db.flush()

        # Crear seguimientos mensuales (12 meses)
        for mes in range(1, 13):
            seguimiento = PlanTrabajoSeguimiento(
                actividad_id=actividad.id,
                mes=mes,
                programada=(mes in meses_programados),
                ejecutada=False,
            )
            db.add(seguimiento)

    db.commit()
    db.refresh(plan)
    return plan


@router.get("/{plan_id}", response_model=PlanTrabajoAnualDetailResponse)
def obtener_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    plan = db.query(PlanTrabajoAnual).filter(PlanTrabajoAnual.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan no encontrado")
    return plan


@router.put("/{plan_id}", response_model=PlanTrabajoAnualResponse)
def actualizar_plan(
    plan_id: int,
    plan_data: PlanTrabajoAnualUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    plan = db.query(PlanTrabajoAnual).filter(PlanTrabajoAnual.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan no encontrado")

    for field, value in plan_data.model_dump(exclude_unset=True).items():
        setattr(plan, field, value)

    db.commit()
    db.refresh(plan)
    return plan


@router.delete("/{plan_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    plan = db.query(PlanTrabajoAnual).filter(PlanTrabajoAnual.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan no encontrado")
    db.delete(plan)
    db.commit()


# =====================================================================
# ACTIVIDADES
# =====================================================================

@router.get("/{plan_id}/actividades", response_model=List[PlanTrabajoActividadResponse])
def listar_actividades(
    plan_id: int,
    ciclo: Optional[CicloPhva] = Query(None),
    categoria: Optional[CategoriaActividad] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    plan = db.query(PlanTrabajoAnual).filter(PlanTrabajoAnual.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan no encontrado")

    query = db.query(PlanTrabajoActividad).filter(PlanTrabajoActividad.plan_id == plan_id)
    if ciclo:
        query = query.filter(PlanTrabajoActividad.ciclo == ciclo)
    if categoria:
        query = query.filter(PlanTrabajoActividad.categoria == categoria)

    return query.order_by(PlanTrabajoActividad.orden).all()


@router.post("/{plan_id}/actividades", response_model=PlanTrabajoActividadResponse, status_code=status.HTTP_201_CREATED)
def crear_actividad(
    plan_id: int,
    actividad_data: PlanTrabajoActividadCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    plan = db.query(PlanTrabajoAnual).filter(PlanTrabajoAnual.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan no encontrado")

    actividad = PlanTrabajoActividad(plan_id=plan_id, **actividad_data.model_dump())
    db.add(actividad)
    db.flush()

    # Crear los 12 seguimientos mensuales vacíos
    for mes in range(1, 13):
        db.add(PlanTrabajoSeguimiento(actividad_id=actividad.id, mes=mes))

    db.commit()
    db.refresh(actividad)
    return actividad


@router.put("/actividades/{actividad_id}", response_model=PlanTrabajoActividadResponse)
def actualizar_actividad(
    actividad_id: int,
    actividad_data: PlanTrabajoActividadUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    actividad = db.query(PlanTrabajoActividad).filter(PlanTrabajoActividad.id == actividad_id).first()
    if not actividad:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")

    for field, value in actividad_data.model_dump(exclude_unset=True).items():
        setattr(actividad, field, value)

    db.commit()
    db.refresh(actividad)
    return actividad


@router.delete("/actividades/{actividad_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_actividad(
    actividad_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    actividad = db.query(PlanTrabajoActividad).filter(PlanTrabajoActividad.id == actividad_id).first()
    if not actividad:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")
    db.delete(actividad)
    db.commit()


# =====================================================================
# SEGUIMIENTO MENSUAL (P/E)
# =====================================================================

@router.put("/actividades/{actividad_id}/seguimiento/{mes}", response_model=PlanTrabajoSeguimientoResponse)
def actualizar_seguimiento_mes(
    actividad_id: int,
    mes: int,
    seguimiento_data: PlanTrabajoSeguimientoUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El mes debe estar entre 1 y 12")

    actividad = db.query(PlanTrabajoActividad).filter(PlanTrabajoActividad.id == actividad_id).first()
    if not actividad:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")

    seguimiento = db.query(PlanTrabajoSeguimiento).filter(
        PlanTrabajoSeguimiento.actividad_id == actividad_id,
        PlanTrabajoSeguimiento.mes == mes
    ).first()

    if not seguimiento:
        seguimiento = PlanTrabajoSeguimiento(actividad_id=actividad_id, mes=mes)
        db.add(seguimiento)

    for field, value in seguimiento_data.model_dump(exclude_unset=True).items():
        setattr(seguimiento, field, value)

    db.commit()
    db.refresh(seguimiento)
    return seguimiento


# =====================================================================
# DASHBOARD / INDICADORES
# =====================================================================

@router.get("/{plan_id}/dashboard", response_model=DashboardIndicadores)
def obtener_dashboard(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    plan = db.query(PlanTrabajoAnual).filter(PlanTrabajoAnual.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan no encontrado")

    # Obtener todos los seguimientos del plan en una sola query
    seguimientos = (
        db.query(PlanTrabajoSeguimiento)
        .join(PlanTrabajoActividad)
        .filter(PlanTrabajoActividad.plan_id == plan_id)
        .all()
    )

    # Calcular indicadores por mes
    meses_data: dict = {m: {"programadas": 0, "ejecutadas": 0} for m in range(1, 13)}
    for s in seguimientos:
        if s.programada:
            meses_data[s.mes]["programadas"] += 1
        if s.ejecutada:
            meses_data[s.mes]["ejecutadas"] += 1

    meses_indicadores = []
    for mes_num in range(1, 13):
        prog = meses_data[mes_num]["programadas"]
        ejec = meses_data[mes_num]["ejecutadas"]
        pct = round((ejec / prog * 100), 1) if prog > 0 else 0.0
        meses_indicadores.append(MesIndicador(
            mes=mes_num,
            nombre_mes=NOMBRE_MESES[mes_num],
            programadas=prog,
            ejecutadas=ejec,
            porcentaje=pct,
        ))

    total_prog = sum(m.programadas for m in meses_indicadores)
    total_ejec = sum(m.ejecutadas for m in meses_indicadores)
    pct_global = round((total_ejec / total_prog * 100), 1) if total_prog > 0 else 0.0

    return DashboardIndicadores(
        plan_id=plan_id,
        año=plan.año,
        meta_porcentaje=plan.meta_porcentaje,
        total_programadas=total_prog,
        total_ejecutadas=total_ejec,
        porcentaje_global=pct_global,
        meses=meses_indicadores,
    )


# =====================================================================
# EXPORTAR A EXCEL
# =====================================================================

_MESES_ABREV = ['', 'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC']
_CICLO_LABELS_MAP = {
    'I_PLANEAR': 'I. PLANEAR', 'II_HACER': 'II. HACER',
    'III_VERIFICAR': 'III. VERIFICAR', 'IV_ACTUAR': 'IV. ACTUAR',
}
_CATEGORIA_LABELS_MAP = {
    'RECURSOS': 'RECURSOS',
    'GESTION_INTEGRAL': 'GESTIÓN INTEGRAL DEL SG-SST',
    'GESTION_SALUD': 'GESTIÓN DE LA SALUD',
    'GESTION_PELIGROS': 'GESTIÓN DE PELIGROS Y RIESGOS',
    'GESTION_AMENAZAS': 'GESTIÓN DE AMENAZAS',
    'VERIFICACION': 'VERIFICACIÓN DEL SG-SST',
    'MEJORAMIENTO': 'MEJORAMIENTO',
}
_CICLO_BG = {'I_PLANEAR': '1565C0', 'II_HACER': '2E7D32', 'III_VERIFICAR': 'E65100', 'IV_ACTUAR': '6A1B9A'}
_CICLO_LIGHT = {'I_PLANEAR': 'BBDEFB', 'II_HACER': 'C8E6C9', 'III_VERIFICAR': 'FFE0B2', 'IV_ACTUAR': 'E1BEE7'}
_CICLOS_ORDER = ['I_PLANEAR', 'II_HACER', 'III_VERIFICAR', 'IV_ACTUAR']
_CATS_BY_CICLO = {
    'I_PLANEAR': ['RECURSOS', 'GESTION_INTEGRAL'],
    'II_HACER': ['GESTION_SALUD', 'GESTION_PELIGROS', 'GESTION_AMENAZAS'],
    'III_VERIFICAR': ['VERIFICACION'],
    'IV_ACTUAR': ['MEJORAMIENTO'],
}


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _thin_border() -> Border:
    s = Side(style='thin', color='BDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)


def _apply(cell, fill=None, font=None, align=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    cell.border = _thin_border()


def _enum_val(v) -> str:
    return v.value if hasattr(v, 'value') else str(v)


@router.get("/{plan_id}/exportar/excel")
def exportar_plan_excel(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta el Plan de Trabajo Anual a Excel (.xlsx)."""
    plan = db.query(PlanTrabajoAnual).filter(PlanTrabajoAnual.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan no encontrado")

    actividades = (
        db.query(PlanTrabajoActividad)
        .filter(PlanTrabajoActividad.plan_id == plan_id)
        .order_by(PlanTrabajoActividad.orden)
        .all()
    )

    act_ids = [a.id for a in actividades]
    segs_list = (
        db.query(PlanTrabajoSeguimiento)
        .filter(PlanTrabajoSeguimiento.actividad_id.in_(act_ids))
        .all()
    ) if act_ids else []
    seg_dict: dict = {}
    for s in segs_list:
        seg_dict.setdefault(s.actividad_id, {})[s.mes] = s

    # Column layout: 4 fixed + 24 P/E (12×2) + 1 obs = 29 total
    COL_ESTANDAR, COL_DESC, COL_FREC, COL_RESP, COL_OBS, TOTAL_COLS = 1, 2, 3, 4, 29, 29

    def p_col(m): return 3 + 2 * m   # P column for month m
    def e_col(m): return 4 + 2 * m   # E column for month m

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Plan {plan.año}"

    # Column widths
    ws.column_dimensions['A'].width = 10   # Estándar
    ws.column_dimensions['B'].width = 48   # Descripción
    ws.column_dimensions['C'].width = 13   # Frecuencia
    ws.column_dimensions['D'].width = 20   # Responsable
    for m in range(1, 13):
        ws.column_dimensions[get_column_letter(p_col(m))].width = 3.5
        ws.column_dimensions[get_column_letter(e_col(m))].width = 3.5
    ws.column_dimensions[get_column_letter(COL_OBS)].width = 22

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    c = ws.cell(row=1, column=1, value=f'PLAN DE TRABAJO ANUAL SG-SST — {plan.año}')
    _apply(c, _make_fill('1A237E'), Font(color='FFFFFF', bold=True, size=14),
           Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 32

    # Row 2: meta info
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.cell(row=2, column=1, value=f'Código: {plan.codigo}  |  Versión: {plan.version}  |  Estado: {(plan.estado or "").upper()}')
    ws.merge_cells(start_row=2, start_column=9, end_row=2, end_column=TOTAL_COLS)
    encargado_txt = f'  |  Encargado: {plan.encargado_sgsst}' if plan.encargado_sgsst else ''
    ws.cell(row=2, column=9, value=f'Meta: {plan.meta_porcentaje}% cumplimiento mensual{encargado_txt}')
    meta_fill = _make_fill('E3F2FD')
    meta_font = Font(size=9)
    for col in range(1, TOTAL_COLS + 1):
        _apply(ws.cell(row=2, column=col), meta_fill, meta_font, Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 4  # spacer

    # Rows 4-5: column headers
    HDR1, HDR2 = 4, 5
    for col in [COL_ESTANDAR, COL_DESC, COL_FREC, COL_RESP, COL_OBS]:
        ws.merge_cells(start_row=HDR1, start_column=col, end_row=HDR2, end_column=col)
    labels = {COL_ESTANDAR: 'Estándar', COL_DESC: 'Descripción de la Actividad',
              COL_FREC: 'Frecuencia', COL_RESP: 'Responsable', COL_OBS: 'Observaciones'}
    for col, label in labels.items():
        ws.cell(row=HDR1, column=col, value=label)

    for m in range(1, 13):
        pc, ec = p_col(m), e_col(m)
        ws.merge_cells(start_row=HDR1, start_column=pc, end_row=HDR1, end_column=ec)
        ws.cell(row=HDR1, column=pc, value=_MESES_ABREV[m])
        ws.cell(row=HDR2, column=pc, value='P')
        ws.cell(row=HDR2, column=ec, value='E')

    hdr_fill = _make_fill('37474F')
    hdr_font = Font(color='FFFFFF', bold=True, size=9)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in [HDR1, HDR2]:
        for col in range(1, TOTAL_COLS + 1):
            _apply(ws.cell(row=row, column=col), hdr_fill, hdr_font, hdr_align)
    ws.row_dimensions[HDR1].height = 20
    ws.row_dimensions[HDR2].height = 14

    # Data rows
    acts_by_key: dict = {}
    for act in actividades:
        key = (_enum_val(act.ciclo), _enum_val(act.categoria))
        acts_by_key.setdefault(key, []).append(act)

    current_row = 6
    prev_ciclo = None
    for ciclo in _CICLOS_ORDER:
        for cat in _CATS_BY_CICLO.get(ciclo, []):
            acts = acts_by_key.get((ciclo, cat), [])
            if not acts:
                continue

            if ciclo != prev_ciclo:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=TOTAL_COLS)
                c = ws.cell(row=current_row, column=1, value=_CICLO_LABELS_MAP.get(ciclo, ciclo))
                _apply(c, _make_fill(_CICLO_BG[ciclo]),
                       Font(color='FFFFFF', bold=True, size=11),
                       Alignment(horizontal='center', vertical='center'))
                ws.row_dimensions[current_row].height = 22
                current_row += 1
                prev_ciclo = ciclo

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=TOTAL_COLS)
            c = ws.cell(row=current_row, column=1, value=_CATEGORIA_LABELS_MAP.get(cat, cat))
            _apply(c, _make_fill(_CICLO_LIGHT[ciclo]),
                   Font(bold=True, size=9),
                   Alignment(horizontal='left', vertical='center', indent=1))
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            for act in acts:
                segs = seg_dict.get(act.id, {})
                txt_align = Alignment(vertical='top', wrap_text=True)
                ctr_align = Alignment(horizontal='center', vertical='center')

                for col, val in [
                    (COL_ESTANDAR, act.estandar or ''),
                    (COL_DESC, act.descripcion),
                    (COL_FREC, act.frecuencia or ''),
                    (COL_RESP, act.responsable or ''),
                    (COL_OBS, act.observaciones or ''),
                ]:
                    _apply(ws.cell(row=current_row, column=col, value=val),
                           font=Font(size=8), align=txt_align)

                for m in range(1, 13):
                    s = segs.get(m)
                    prog = s.programada if s else False
                    ejec = s.ejecutada if s else False
                    p_val = '✓' if prog else ''
                    e_val = '✓' if ejec else ''

                    p_fill = _make_fill('A5D6A7') if (prog and ejec) else (_make_fill('FFF9C4') if prog else None)
                    e_fill = _make_fill('C8E6C9') if ejec else None

                    pc, ec = p_col(m), e_col(m)
                    _apply(ws.cell(row=current_row, column=pc, value=p_val),
                           fill=p_fill, font=Font(size=8, bold=True), align=ctr_align)
                    _apply(ws.cell(row=current_row, column=ec, value=e_val),
                           fill=e_fill, font=Font(size=8, bold=True, color='1B5E20'), align=ctr_align)

                ws.row_dimensions[current_row].height = 30
                current_row += 1

    ws.freeze_panes = ws.cell(row=6, column=5)  # freeze header + 4 fixed cols

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"plan_trabajo_anual_{plan.año}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =====================================================================
# EXPORTAR A PDF
# =====================================================================

@router.get("/{plan_id}/exportar/pdf")
def exportar_plan_pdf(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta el Plan de Trabajo Anual a PDF (A4 apaisado)."""
    plan = db.query(PlanTrabajoAnual).filter(PlanTrabajoAnual.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plan no encontrado")

    actividades = (
        db.query(PlanTrabajoActividad)
        .filter(PlanTrabajoActividad.plan_id == plan_id)
        .order_by(PlanTrabajoActividad.orden)
        .all()
    )

    act_ids = [a.id for a in actividades]
    segs_list = (
        db.query(PlanTrabajoSeguimiento)
        .filter(PlanTrabajoSeguimiento.actividad_id.in_(act_ids))
        .all()
    ) if act_ids else []
    seg_dict: dict = {}
    for s in segs_list:
        seg_dict.setdefault(s.actividad_id, {})[s.mes] = s

    MESES_ABREV_PDF = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    acts_by_key: dict = {}
    for act in actividades:
        key = (_enum_val(act.ciclo), _enum_val(act.categoria))
        acts_by_key.setdefault(key, []).append(act)

    # Month headers HTML
    month_ths = ''.join(
        f'<th colspan="2" class="mes-hdr">{MESES_ABREV_PDF[m]}</th>' for m in range(1, 13)
    )
    pe_ths = '<th class="pe">P</th><th class="pe">E</th>' * 12

    # Build table body
    body_rows = ''
    prev_ciclo = None
    for ciclo in _CICLOS_ORDER:
        for cat in _CATS_BY_CICLO.get(ciclo, []):
            acts = acts_by_key.get((ciclo, cat), [])
            if not acts:
                continue

            if ciclo != prev_ciclo:
                color = {'I_PLANEAR': '#1565C0', 'II_HACER': '#2E7D32',
                         'III_VERIFICAR': '#E65100', 'IV_ACTUAR': '#6A1B9A'}[ciclo]
                label = _CICLO_LABELS_MAP.get(ciclo, ciclo)
                body_rows += (
                    f'<tr><td colspan="29" class="ciclo-hdr" style="background:{color}">'
                    f'{label}</td></tr>\n'
                )
                prev_ciclo = ciclo

            light = {'I_PLANEAR': '#BBDEFB', 'II_HACER': '#C8E6C9',
                     'III_VERIFICAR': '#FFE0B2', 'IV_ACTUAR': '#E1BEE7'}[ciclo]
            cat_label = _CATEGORIA_LABELS_MAP.get(cat, cat)
            body_rows += (
                f'<tr><td colspan="29" class="cat-hdr" style="background:{light}">'
                f'{cat_label}</td></tr>\n'
            )

            for act in acts:
                segs = seg_dict.get(act.id, {})
                cells = ''
                for m in range(1, 13):
                    s = segs.get(m)
                    prog = s.programada if s else False
                    ejec = s.ejecutada if s else False
                    p_cls = 'pe-both' if (prog and ejec) else ('pe-prog' if prog else 'pe-empty')
                    e_cls = 'pe-exec' if ejec else 'pe-empty'
                    cells += (
                        f'<td class="pe {p_cls}">{"✓" if prog else ""}</td>'
                        f'<td class="pe {e_cls}">{"✓" if ejec else ""}</td>'
                    )

                body_rows += (
                    f'<tr>'
                    f'<td class="estandar">{act.estandar or ""}</td>'
                    f'<td class="desc">{act.descripcion}</td>'
                    f'<td class="frec">{act.frecuencia or ""}</td>'
                    f'<td class="resp">{act.responsable or ""}</td>'
                    f'{cells}'
                    f'<td class="obs">{act.observaciones or ""}</td>'
                    f'</tr>\n'
                )

    encargado_html = f'<span><b>Encargado SG-SST:</b> {plan.encargado_sgsst}</span>' if plan.encargado_sgsst else ''
    aprobado_html = f'<span><b>Aprobado por:</b> {plan.aprobado_por}</span>' if plan.aprobado_por else ''
    now = datetime.now().strftime('%d/%m/%Y %H:%M')

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Plan de Trabajo Anual SG-SST {plan.año}</title>
<style>
  @page {{ size: A4 landscape; margin: 8mm 6mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 6.5pt; color: #212121; }}
  h1 {{ text-align: center; font-size: 10.5pt; color: #1A237E; margin-bottom: 2mm; font-weight: bold; }}
  .meta {{ background: #E8EAF6; padding: 1.5mm 2mm; border-radius: 2px; margin-bottom: 2mm; }}
  .meta span {{ margin-right: 5mm; }}
  table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
  th, td {{ border: 0.3pt solid #9E9E9E; padding: 0.8mm; vertical-align: top; overflow: hidden; word-break: break-word; }}
  thead th {{ background: #37474F; color: #fff; font-weight: bold; text-align: center; vertical-align: middle; }}
  .mes-hdr {{ font-size: 5.5pt; }}
  .pe {{ text-align: center; font-weight: bold; font-size: 6pt; width: 4mm; }}
  .ciclo-hdr {{ color: #fff; font-weight: bold; font-size: 7.5pt; text-align: center; padding: 1.5mm; }}
  .cat-hdr {{ font-weight: bold; font-size: 6pt; padding: 1mm 2mm; }}
  .estandar {{ width: 9mm; font-size: 5.5pt; text-align: center; }}
  .desc {{ width: 50mm; }}
  .frec {{ width: 12mm; font-size: 5.5pt; text-align: center; }}
  .resp {{ width: 17mm; font-size: 5.5pt; }}
  .obs {{ width: 14mm; font-size: 5.5pt; }}
  .pe-prog {{ background: #FFF9C4; }}
  .pe-exec {{ background: #C8E6C9; color: #1B5E20; }}
  .pe-both {{ background: #A5D6A7; color: #1B5E20; }}
  .pe-empty {{ }}
  .footer {{ margin-top: 2mm; font-size: 5pt; color: #757575; text-align: right; }}
</style>
</head>
<body>
  <h1>PLAN DE TRABAJO ANUAL SG-SST &mdash; {plan.año}</h1>
  <div class="meta">
    <span><b>Código:</b> {plan.codigo}</span>
    <span><b>Versión:</b> {plan.version}</span>
    <span><b>Estado:</b> {(plan.estado or '').upper()}</span>
    <span><b>Meta:</b> {plan.meta_porcentaje}% cumplimiento mensual</span>
    {encargado_html}{aprobado_html}
  </div>
  <table>
    <thead>
      <tr>
        <th rowspan="2" style="width:9mm">Estándar</th>
        <th rowspan="2" style="width:50mm">Descripción de la Actividad</th>
        <th rowspan="2" style="width:12mm">Frecuencia</th>
        <th rowspan="2" style="width:17mm">Responsable</th>
        {month_ths}
        <th rowspan="2" style="width:14mm">Obs.</th>
      </tr>
      <tr>{pe_ths}</tr>
    </thead>
    <tbody>
{body_rows}
    </tbody>
  </table>
  <div class="footer">Generado: {now} &bull; Sistema SG-SST &bull; Código PL-SST-02</div>
</body>
</html>"""

    pdf_bytes = weasyprint.HTML(string=html).write_pdf(
        metadata={
            "title": f"Plan de Trabajo Anual SG-SST {plan.año}",
            "author": "Sistema SG-SST",
        },
        pdf_version=(1, 7),
        compress=True,
    )

    filename = f"plan_trabajo_anual_{plan.año}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
