import io
from datetime import datetime
from decimal import Decimal
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import weasyprint

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user, require_supervisor_or_admin
from app.models.user import User
from app.models.presupuesto_sst import (
    PresupuestoSST,
    PresupuestoCategoria,
    PresupuestoItem,
    PresupuestoMensual,
)
from app.schemas.presupuesto_sst import (
    PresupuestoSSTCreate,
    PresupuestoSSTUpdate,
    PresupuestoSSTResponse,
    PresupuestoSSTDetailResponse,
    PresupuestoItemCreate,
    PresupuestoItemUpdate,
    PresupuestoItemResponse,
    PresupuestoMensualUpdate,
    PresupuestoMensualResponse,
)
from app.services.presupuesto_template import (
    CATEGORIAS_ORDER,
    CATEGORIA_LABELS,
    get_default_items,
)

router = APIRouter()

MESES_ABREV = [
    "", "ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
    "JUL", "AGOS", "SEPT", "OCT", "NOV", "DIC",
]
MESES_NOMBRE = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

# Colores para Excel (mismo estilo que plan_trabajo_anual.py)
CAT_COLORS = {
    "MEDICINA_PREVENTIVA":  "1565C0",
    "HIGIENE_INDUSTRIAL":   "2E7D32",
    "SEGURIDAD_INDUSTRIAL": "E65100",
    "CAPACITACION":         "6A1B9A",
    "INFRAESTRUCTURA":      "37474F",
}
CAT_LIGHT_COLORS = {
    "MEDICINA_PREVENTIVA":  "E3F2FD",
    "HIGIENE_INDUSTRIAL":   "E8F5E9",
    "SEGURIDAD_INDUSTRIAL": "FFF3E0",
    "CAPACITACION":         "F3E5F5",
    "INFRAESTRUCTURA":      "ECEFF1",
}


# ─────────────────────────────────────────────
# Helpers Excel
# ─────────────────────────────────────────────

def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply(cell, fill=None, font=None, align=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def _fmt_money(val) -> str:
    v = float(val or 0)
    return f"${v:,.0f}"


def _pct(num, den) -> float:
    return round((float(num) / float(den)) * 100, 1) if den and float(den) != 0 else 0.0


# ─────────────────────────────────────────────
# CRUD — Presupuesto
# ─────────────────────────────────────────────

@router.get("/", response_model=List[PresupuestoSSTResponse])
def listar_presupuestos(
    año: Optional[int] = Query(None),
    empresa_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(PresupuestoSST)
    if año:
        query = query.filter(PresupuestoSST.año == año)
    if empresa_id:
        query = query.filter(PresupuestoSST.empresa_id == empresa_id)
    return query.order_by(PresupuestoSST.año.desc()).all()


@router.post(
    "/crear-desde-plantilla",
    response_model=PresupuestoSSTDetailResponse,
    status_code=status.HTTP_201_CREATED,
)
def crear_desde_plantilla(
    año: int = Query(..., description="Año del presupuesto"),
    empresa_id: Optional[int] = Query(None),
    encargado_sgsst: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    existente = db.query(PresupuestoSST).filter(
        PresupuestoSST.año == año,
        PresupuestoSST.empresa_id == empresa_id,
    ).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Ya existe un Presupuesto SST para el año {año}",
        )

    presupuesto = PresupuestoSST(
        año=año,
        empresa_id=empresa_id,
        encargado_sgsst=encargado_sgsst,
        created_by=current_user.id,
    )
    db.add(presupuesto)
    db.flush()

    for orden_cat, cat_str in enumerate(CATEGORIAS_ORDER):
        categoria = PresupuestoCategoria(
            presupuesto_id=presupuesto.id,
            categoria=cat_str,
            orden=orden_cat,
        )
        db.add(categoria)
        db.flush()

        for orden_item, nombre in enumerate(get_default_items(cat_str)):
            item = PresupuestoItem(
                categoria_id=categoria.id,
                actividad=nombre,
                es_default=True,
                orden=orden_item,
            )
            db.add(item)
            db.flush()

            for mes in range(1, 13):
                db.add(PresupuestoMensual(
                    item_id=item.id,
                    mes=mes,
                    proyectado=Decimal("0"),
                    ejecutado=Decimal("0"),
                ))

    db.commit()
    db.refresh(presupuesto)
    return presupuesto


@router.get("/{presupuesto_id}", response_model=PresupuestoSSTDetailResponse)
def obtener_presupuesto(
    presupuesto_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    p = db.query(PresupuestoSST).filter(PresupuestoSST.id == presupuesto_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    return p


@router.put("/{presupuesto_id}", response_model=PresupuestoSSTResponse)
def actualizar_presupuesto(
    presupuesto_id: int,
    data: PresupuestoSSTUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    p = db.query(PresupuestoSST).filter(PresupuestoSST.id == presupuesto_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(p, field, value)
    p.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(p)
    return p


@router.delete("/{presupuesto_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_presupuesto(
    presupuesto_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    p = db.query(PresupuestoSST).filter(PresupuestoSST.id == presupuesto_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
    db.delete(p)
    db.commit()


# ─────────────────────────────────────────────
# Items
# ─────────────────────────────────────────────

@router.post(
    "/{presupuesto_id}/categorias/{categoria}/items",
    response_model=PresupuestoItemResponse,
    status_code=status.HTTP_201_CREATED,
)
def agregar_item(
    presupuesto_id: int,
    categoria: str,
    data: PresupuestoItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    cat_obj = db.query(PresupuestoCategoria).filter(
        PresupuestoCategoria.presupuesto_id == presupuesto_id,
        PresupuestoCategoria.categoria == categoria.upper(),
    ).first()
    if not cat_obj:
        raise HTTPException(status_code=404, detail="Categoría no encontrada")

    item = PresupuestoItem(
        categoria_id=cat_obj.id,
        actividad=data.actividad,
        es_default=False,
        orden=data.orden,
    )
    db.add(item)
    db.flush()

    for mes in range(1, 13):
        db.add(PresupuestoMensual(
            item_id=item.id,
            mes=mes,
            proyectado=Decimal("0"),
            ejecutado=Decimal("0"),
        ))

    db.commit()
    db.refresh(item)
    return item


@router.put("/items/{item_id}", response_model=PresupuestoItemResponse)
def actualizar_item(
    item_id: int,
    data: PresupuestoItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    item = db.query(PresupuestoItem).filter(PresupuestoItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    item.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)
    return item


@router.delete("/items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    item = db.query(PresupuestoItem).filter(PresupuestoItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")
    db.delete(item)
    db.commit()


# ─────────────────────────────────────────────
# Montos mensuales
# ─────────────────────────────────────────────

@router.put("/items/{item_id}/mensual/{mes}", response_model=PresupuestoMensualResponse)
def actualizar_mensual(
    item_id: int,
    mes: int,
    data: PresupuestoMensualUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=400, detail="El mes debe estar entre 1 y 12")

    item = db.query(PresupuestoItem).filter(PresupuestoItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")

    mensual = db.query(PresupuestoMensual).filter(
        PresupuestoMensual.item_id == item_id,
        PresupuestoMensual.mes == mes,
    ).first()

    if not mensual:
        mensual = PresupuestoMensual(item_id=item_id, mes=mes, proyectado=0, ejecutado=0)
        db.add(mensual)

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(mensual, field, value)
    mensual.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(mensual)
    return mensual


# ─────────────────────────────────────────────
# Exportar Excel
# ─────────────────────────────────────────────

@router.get("/{presupuesto_id}/exportar/excel")
def exportar_excel(
    presupuesto_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta el Presupuesto SST a Excel (.xlsx)."""
    p = db.query(PresupuestoSST).filter(PresupuestoSST.id == presupuesto_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")

    # Recargar categorías e ítems
    categorias = (
        db.query(PresupuestoCategoria)
        .filter(PresupuestoCategoria.presupuesto_id == presupuesto_id)
        .order_by(PresupuestoCategoria.orden)
        .all()
    )

    item_ids = []
    for cat in categorias:
        for it in cat.items:
            item_ids.append(it.id)

    montos_list = (
        db.query(PresupuestoMensual)
        .filter(PresupuestoMensual.item_id.in_(item_ids))
        .all()
    ) if item_ids else []

    monto_dict: dict = {}
    for m in montos_list:
        monto_dict.setdefault(m.item_id, {})[m.mes] = m

    # ── Layout de columnas ──
    # A=1: Actividad | B=2: TotalProy | C=3: TotalEjec | D=4: %Ejec
    # E=5: PorEjecutar | F=6: %PorEj | G-Z (7-30): 12×2 meses
    COL_ACT = 1
    COL_TPROY = 2
    COL_TEJEC = 3
    COL_PCT_EJ = 4
    COL_POR_EJ = 5
    COL_PCT_POR = 6
    TOTAL_COLS = 30

    def proy_col(m): return 5 + 2 * m   # m=1 → 7 (G)
    def ejec_col(m): return 6 + 2 * m   # m=1 → 8 (H)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Presupuesto {p.año}"

    # Anchos de columna
    ws.column_dimensions["A"].width = 45  # Actividad
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 8
    for m in range(1, 13):
        ws.column_dimensions[get_column_letter(proy_col(m))].width = 11
        ws.column_dimensions[get_column_letter(ejec_col(m))].width = 11

    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    # ── Fila 1: Título ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    c = ws.cell(row=1, column=1, value=f"CONSOLIDADO GENERAL PRESUPUESTO SG-SST — {p.año}")
    _apply(c, _make_fill("1A237E"), Font(color="FFFFFF", bold=True, size=13), center)
    ws.row_dimensions[1].height = 28

    # ── Fila 2: meta info ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.cell(row=2, column=1,
            value=f"Código: {p.codigo}  |  Versión: {p.version}")
    ws.merge_cells(start_row=2, start_column=11, end_row=2, end_column=TOTAL_COLS)
    enc = f"  |  Encargado: {p.encargado_sgsst}" if p.encargado_sgsst else ""
    ws.cell(row=2, column=11, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}{enc}")
    meta_fill = _make_fill("E3F2FD")
    for col in range(1, TOTAL_COLS + 1):
        _apply(ws.cell(row=2, column=col), meta_fill, Font(size=9), left)
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 4  # spacer

    # ── Filas 4-5: Encabezados ──
    HDR1, HDR2 = 4, 5

    # Encabezados fijos (merge en ambas filas)
    fixed_headers = {
        COL_ACT:    "ACTIVIDADES",
        COL_TPROY:  "PRESUPUESTO\nPROYECTADO",
        COL_TEJEC:  "PRESUPUESTO\nEJECUTADO",
        COL_PCT_EJ: "%",
        COL_POR_EJ: "PRESUPUESTO\nPOR EJECUTAR",
        COL_PCT_POR: "%",
    }
    for col, label in fixed_headers.items():
        ws.merge_cells(start_row=HDR1, start_column=col, end_row=HDR2, end_column=col)
        ws.cell(row=HDR1, column=col, value=label)

    # Encabezados de mes (merge HDR1 cols Proy+Ejec, en HDR2 P / E)
    for m in range(1, 13):
        pc, ec = proy_col(m), ejec_col(m)
        ws.merge_cells(start_row=HDR1, start_column=pc, end_row=HDR1, end_column=ec)
        ws.cell(row=HDR1, column=pc, value=MESES_ABREV[m])
        ws.cell(row=HDR2, column=pc, value="PROY")
        ws.cell(row=HDR2, column=ec, value="EJEC")

    hdr_fill = _make_fill("37474F")
    hdr_font = Font(color="FFFFFF", bold=True, size=8)
    for row in [HDR1, HDR2]:
        for col in range(1, TOTAL_COLS + 1):
            _apply(ws.cell(row=row, column=col), hdr_fill, hdr_font, center, border)
    ws.row_dimensions[HDR1].height = 22
    ws.row_dimensions[HDR2].height = 14

    # ── Datos ──
    current_row = 6

    grand_proy = Decimal("0")
    grand_ejec = Decimal("0")

    for cat in categorias:
        cat_str = cat.categoria
        cat_label = CATEGORIA_LABELS.get(cat_str, cat_str)
        bg_dark = CAT_COLORS.get(cat_str, "37474F")
        bg_light = CAT_LIGHT_COLORS.get(cat_str, "F5F5F5")

        # Fila de categoría
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=TOTAL_COLS
        )
        c = ws.cell(row=current_row, column=1, value=cat_label)
        _apply(c, _make_fill(bg_dark), Font(color="FFFFFF", bold=True, size=10), left, border)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        cat_proy = Decimal("0")
        cat_ejec = Decimal("0")

        for item in cat.items:
            montos = monto_dict.get(item.id, {})

            total_proy = sum(Decimal(str(montos[m].proyectado)) for m in range(1, 13) if m in montos)
            total_ejec = sum(Decimal(str(montos[m].ejecutado)) for m in range(1, 13) if m in montos)
            por_ejecutar = total_proy - total_ejec
            pct_ejec = _pct(total_ejec, total_proy)
            pct_por = _pct(por_ejecutar, total_proy)

            cat_proy += total_proy
            cat_ejec += total_ejec

            item_fill = _make_fill(bg_light)
            num_font = Font(size=8)
            txt_font = Font(size=8)

            ws.cell(row=current_row, column=COL_ACT, value=item.actividad)
            _apply(ws.cell(row=current_row, column=COL_ACT), item_fill, txt_font, left, border)

            ws.cell(row=current_row, column=COL_TPROY, value=float(total_proy))
            ws.cell(row=current_row, column=COL_TPROY).number_format = '$#,##0'
            _apply(ws.cell(row=current_row, column=COL_TPROY), item_fill, num_font, right_align, border)

            ws.cell(row=current_row, column=COL_TEJEC, value=float(total_ejec))
            ws.cell(row=current_row, column=COL_TEJEC).number_format = '$#,##0'
            _apply(ws.cell(row=current_row, column=COL_TEJEC), item_fill, num_font, right_align, border)

            ws.cell(row=current_row, column=COL_PCT_EJ, value=pct_ejec / 100)
            ws.cell(row=current_row, column=COL_PCT_EJ).number_format = '0%'
            _apply(ws.cell(row=current_row, column=COL_PCT_EJ), item_fill, num_font, center, border)

            ws.cell(row=current_row, column=COL_POR_EJ, value=float(por_ejecutar))
            ws.cell(row=current_row, column=COL_POR_EJ).number_format = '$#,##0'
            _apply(ws.cell(row=current_row, column=COL_POR_EJ), item_fill, num_font, right_align, border)

            ws.cell(row=current_row, column=COL_PCT_POR, value=pct_por / 100)
            ws.cell(row=current_row, column=COL_PCT_POR).number_format = '0%'
            _apply(ws.cell(row=current_row, column=COL_PCT_POR), item_fill, num_font, center, border)

            for m in range(1, 13):
                s = montos.get(m)
                pv = float(s.proyectado) if s else 0.0
                ev = float(s.ejecutado) if s else 0.0

                pc, ec = proy_col(m), ejec_col(m)

                proy_fill = _make_fill("FFF9C4") if pv > 0 else item_fill
                ejec_fill = _make_fill("C8E6C9") if ev > 0 else item_fill

                ws.cell(row=current_row, column=pc, value=pv if pv else None)
                if pv:
                    ws.cell(row=current_row, column=pc).number_format = '$#,##0'
                _apply(ws.cell(row=current_row, column=pc), proy_fill, num_font, right_align, border)

                ws.cell(row=current_row, column=ec, value=ev if ev else None)
                if ev:
                    ws.cell(row=current_row, column=ec).number_format = '$#,##0'
                _apply(ws.cell(row=current_row, column=ec), ejec_fill, num_font, right_align, border)

            ws.row_dimensions[current_row].height = 16
            current_row += 1

        # Fila subtotal por categoría
        por_cat = cat_proy - cat_ejec
        ws.cell(row=current_row, column=COL_ACT, value=f"TOTAL {cat_label.upper()}")
        _apply(ws.cell(row=current_row, column=COL_ACT),
               _make_fill(bg_dark), Font(color="FFFFFF", bold=True, size=8), left, border)

        for col, val in [
            (COL_TPROY, float(cat_proy)),
            (COL_TEJEC, float(cat_ejec)),
            (COL_POR_EJ, float(por_cat)),
        ]:
            ws.cell(row=current_row, column=col, value=val)
            ws.cell(row=current_row, column=col).number_format = '$#,##0'
            _apply(ws.cell(row=current_row, column=col),
                   _make_fill(bg_dark), Font(color="FFFFFF", bold=True, size=8), right_align, border)

        pct_c = _pct(cat_ejec, cat_proy)
        ws.cell(row=current_row, column=COL_PCT_EJ, value=pct_c / 100)
        ws.cell(row=current_row, column=COL_PCT_EJ).number_format = '0%'
        _apply(ws.cell(row=current_row, column=COL_PCT_EJ),
               _make_fill(bg_dark), Font(color="FFFFFF", bold=True, size=8), center, border)

        pct_pc = _pct(por_cat, cat_proy)
        ws.cell(row=current_row, column=COL_PCT_POR, value=pct_pc / 100)
        ws.cell(row=current_row, column=COL_PCT_POR).number_format = '0%'
        _apply(ws.cell(row=current_row, column=COL_PCT_POR),
               _make_fill(bg_dark), Font(color="FFFFFF", bold=True, size=8), center, border)

        # Rellenar el resto de columnas de mes vacías para el subtotal
        for col in range(proy_col(1), TOTAL_COLS + 1):
            cell = ws.cell(row=current_row, column=col)
            if cell.value is None:
                _apply(cell, _make_fill(bg_dark), border=border)
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        grand_proy += cat_proy
        grand_ejec += cat_ejec

    # Fila Gran Total
    grand_por = grand_proy - grand_ejec
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=1)
    ws.cell(row=current_row, column=COL_ACT, value="GRAN TOTAL PRESUPUESTO SST")
    _apply(ws.cell(row=current_row, column=COL_ACT),
           _make_fill("1A237E"), Font(color="FFFFFF", bold=True, size=9), left, border)

    for col, val in [
        (COL_TPROY, float(grand_proy)),
        (COL_TEJEC, float(grand_ejec)),
        (COL_POR_EJ, float(grand_por)),
    ]:
        ws.cell(row=current_row, column=col, value=val)
        ws.cell(row=current_row, column=col).number_format = '$#,##0'
        _apply(ws.cell(row=current_row, column=col),
               _make_fill("1A237E"), Font(color="FFFFFF", bold=True, size=9), right_align, border)

    pct_g = _pct(grand_ejec, grand_proy)
    ws.cell(row=current_row, column=COL_PCT_EJ, value=pct_g / 100)
    ws.cell(row=current_row, column=COL_PCT_EJ).number_format = '0%'
    _apply(ws.cell(row=current_row, column=COL_PCT_EJ),
           _make_fill("1A237E"), Font(color="FFFFFF", bold=True, size=9), center, border)

    pct_gp = _pct(grand_por, grand_proy)
    ws.cell(row=current_row, column=COL_PCT_POR, value=pct_gp / 100)
    ws.cell(row=current_row, column=COL_PCT_POR).number_format = '0%'
    _apply(ws.cell(row=current_row, column=COL_PCT_POR),
           _make_fill("1A237E"), Font(color="FFFFFF", bold=True, size=9), center, border)

    for col in range(proy_col(1), TOTAL_COLS + 1):
        cell = ws.cell(row=current_row, column=col)
        if cell.value is None:
            _apply(cell, _make_fill("1A237E"), border=border)

    ws.row_dimensions[current_row].height = 18

    ws.freeze_panes = ws.cell(row=6, column=2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"presupuesto_sst_{p.año}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────
# Exportar PDF
# ─────────────────────────────────────────────

@router.get("/{presupuesto_id}/exportar/pdf")
def exportar_pdf(
    presupuesto_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta el Presupuesto SST a PDF."""
    p = db.query(PresupuestoSST).filter(PresupuestoSST.id == presupuesto_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")

    categorias = (
        db.query(PresupuestoCategoria)
        .filter(PresupuestoCategoria.presupuesto_id == presupuesto_id)
        .order_by(PresupuestoCategoria.orden)
        .all()
    )

    item_ids = []
    for cat in categorias:
        for it in cat.items:
            item_ids.append(it.id)

    montos_list = (
        db.query(PresupuestoMensual)
        .filter(PresupuestoMensual.item_id.in_(item_ids))
        .all()
    ) if item_ids else []

    monto_dict: dict = {}
    for m in montos_list:
        monto_dict.setdefault(m.item_id, {})[m.mes] = m

    # ── Construir HTML ──
    def fmt_money(val) -> str:
        v = float(val or 0)
        if v == 0:
            return "—"
        return f"${v:,.0f}"

    def pct_str(num, den) -> str:
        if not den or float(den) == 0:
            return "0%"
        return f"{round(float(num) / float(den) * 100, 1)}%"

    rows_html = ""
    grand_proy = Decimal("0")
    grand_ejec = Decimal("0")

    for cat in categorias:
        cat_str = cat.categoria
        cat_label = CATEGORIA_LABELS.get(cat_str, cat_str)
        bg_dark = CAT_COLORS.get(cat_str, "37474F")
        bg_light = CAT_LIGHT_COLORS.get(cat_str, "F5F5F5")

        # Fila de categoría
        rows_html += f"""
        <tr>
          <td colspan="30" style="background:#{bg_dark};color:#fff;font-weight:bold;
              padding:5px 8px;font-size:9pt;">{cat_label}</td>
        </tr>"""

        cat_proy = Decimal("0")
        cat_ejec = Decimal("0")

        for item in cat.items:
            montos = monto_dict.get(item.id, {})
            total_proy = sum(Decimal(str(montos[m].proyectado)) for m in range(1, 13) if m in montos)
            total_ejec = sum(Decimal(str(montos[m].ejecutado)) for m in range(1, 13) if m in montos)
            por_ej = total_proy - total_ejec

            cat_proy += total_proy
            cat_ejec += total_ejec

            month_cells = ""
            for m in range(1, 13):
                s = montos.get(m)
                pv = float(s.proyectado) if s else 0.0
                ev = float(s.ejecutado) if s else 0.0
                pv_bg = f"background:#FFF9C4;" if pv > 0 else f"background:#{bg_light};"
                ev_bg = f"background:#C8E6C9;" if ev > 0 else f"background:#{bg_light};"
                month_cells += f'<td style="{pv_bg}text-align:right;">{fmt_money(pv) if pv else "—"}</td>'
                month_cells += f'<td style="{ev_bg}text-align:right;">{fmt_money(ev) if ev else "—"}</td>'

            rows_html += f"""
            <tr style="background:#{bg_light};">
              <td style="padding-left:10px;">{item.actividad}</td>
              <td style="text-align:right;">{fmt_money(total_proy)}</td>
              <td style="text-align:right;">{fmt_money(total_ejec)}</td>
              <td style="text-align:center;">{pct_str(total_ejec, total_proy)}</td>
              <td style="text-align:right;">{fmt_money(por_ej)}</td>
              <td style="text-align:center;">{pct_str(por_ej, total_proy)}</td>
              {month_cells}
            </tr>"""

        por_cat = cat_proy - cat_ejec
        grand_proy += cat_proy
        grand_ejec += cat_ejec

        rows_html += f"""
        <tr style="background:#{bg_dark};color:#fff;font-weight:bold;">
          <td>TOTAL {cat_label.upper()}</td>
          <td style="text-align:right;">{fmt_money(cat_proy)}</td>
          <td style="text-align:right;">{fmt_money(cat_ejec)}</td>
          <td style="text-align:center;">{pct_str(cat_ejec, cat_proy)}</td>
          <td style="text-align:right;">{fmt_money(por_cat)}</td>
          <td style="text-align:center;">{pct_str(por_cat, cat_proy)}</td>
          {"".join(f"<td></td><td></td>" for _ in range(12))}
        </tr>"""

    grand_por = grand_proy - grand_ejec

    mes_headers = "".join(
        f'<th colspan="2" style="text-align:center;">{MESES_ABREV[m]}</th>'
        for m in range(1, 13)
    )
    mes_subheaders = "".join(
        '<th>P</th><th>E</th>'
        for _ in range(12)
    )

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  @page {{ size: A4 landscape; margin: 10mm; }}
  body {{ font-family: Arial, sans-serif; font-size: 7.5pt; }}
  h2 {{ text-align:center; font-size:11pt; margin:0 0 4px 0; color:#1A237E; }}
  .meta {{ text-align:center; font-size:8pt; color:#555; margin-bottom:6px; }}
  table {{ width:100%; border-collapse:collapse; table-layout:fixed; }}
  th, td {{ border:1px solid #ccc; padding:2px 3px; font-size:7pt; overflow:hidden; }}
  th {{ background:#37474F; color:#fff; text-align:center; }}
  td:first-child {{ width:160px; }}
  .total-row {{ background:#1A237E; color:#fff; font-weight:bold; }}
  .footer {{ text-align:right; font-size:7pt; color:#888; margin-top:6px; }}
</style>
</head>
<body>
<h2>CONSOLIDADO GENERAL PRESUPUESTO SG-SST — {p.año}</h2>
<p class="meta">Código: {p.codigo} | Versión: {p.version}
  {"| Encargado: " + p.encargado_sgsst if p.encargado_sgsst else ""}
  | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}
</p>
<table>
  <thead>
    <tr>
      <th rowspan="2">ACTIVIDADES</th>
      <th rowspan="2">PRESUP. PROYECTADO</th>
      <th rowspan="2">PRESUP. EJECUTADO</th>
      <th rowspan="2">%</th>
      <th rowspan="2">POR EJECUTAR</th>
      <th rowspan="2">%</th>
      {mes_headers}
    </tr>
    <tr>{mes_subheaders}</tr>
  </thead>
  <tbody>
    {rows_html}
    <tr class="total-row">
      <td>GRAN TOTAL PRESUPUESTO SST</td>
      <td style="text-align:right;">{fmt_money(grand_proy)}</td>
      <td style="text-align:right;">{fmt_money(grand_ejec)}</td>
      <td style="text-align:center;">{pct_str(grand_ejec, grand_proy)}</td>
      <td style="text-align:right;">{fmt_money(grand_por)}</td>
      <td style="text-align:center;">{pct_str(grand_por, grand_proy)}</td>
      {"".join(f"<td></td><td></td>" for _ in range(12))}
    </tr>
  </tbody>
</table>
<p class="footer">Documento generado automáticamente — Sistema de Gestión SST</p>
</body>
</html>"""

    pdf_bytes = weasyprint.HTML(string=html).write_pdf()
    filename = f"presupuesto_sst_{p.año}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
