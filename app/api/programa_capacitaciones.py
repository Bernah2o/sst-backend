import io
from datetime import datetime
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import weasyprint

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user, require_supervisor_or_admin
from app.models.user import User
from app.models.programa_capacitaciones import (
    ProgramaCapacitaciones,
    CapacitacionActividad,
    CapacitacionSeguimiento,
    CapacitacionIndicadorMensual,
)
from app.schemas.programa_capacitaciones import (
    ProgramaCapacitacionesCreate,
    ProgramaCapacitacionesUpdate,
    ProgramaCapacitacionesResponse,
    ProgramaCapacitacionesDetailResponse,
    CapacitacionActividadCreate,
    CapacitacionActividadUpdate,
    CapacitacionActividadResponse,
    CapacitacionSeguimientoUpdate,
    CapacitacionSeguimientoResponse,
    IndicadorMensualUpdate,
    IndicadorMensualResponse,
    DashboardCapacitaciones,
    KpiIndicador,
    KpiMesData,
)
from app.services.capacitacion_template import (
    get_plantilla_actividades,
    INDICADOR_INFO,
    CICLO_LABELS,
    NOMBRE_MESES,
)

router = APIRouter()

MESES_ABREV = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

_CICLOS_ORDER = ['I_PLANEAR', 'II_HACER', 'III_VERIFICAR', 'IV_ACTUAR']

_CICLO_BG = {
    'I_PLANEAR': '1565C0',
    'II_HACER': '2E7D32',
    'III_VERIFICAR': 'E65100',
    'IV_ACTUAR': '6A1B9A',
}

_CICLO_LIGHT = {
    'I_PLANEAR': 'BBDEFB',
    'II_HACER': 'C8E6C9',
    'III_VERIFICAR': 'FFE0B2',
    'IV_ACTUAR': 'E1BEE7',
}

_CICLO_COLOR_HEX = {
    'I_PLANEAR': '#1565C0',
    'II_HACER': '#2E7D32',
    'III_VERIFICAR': '#E65100',
    'IV_ACTUAR': '#6A1B9A',
}

_CICLO_LIGHT_HEX = {
    'I_PLANEAR': '#BBDEFB',
    'II_HACER': '#C8E6C9',
    'III_VERIFICAR': '#FFE0B2',
    'IV_ACTUAR': '#E1BEE7',
}

TIPO_ORDER = ['CUMPLIMIENTO', 'COBERTURA', 'EFICACIA']


# =====================================================================
# HELPERS
# =====================================================================

def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _thin_border() -> Border:
    s = Side(style='thin', color='BDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)


def _apply(cell, fill=None, font=None, align=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    cell.border = _thin_border()


def _get_programa_or_404(db: Session, programa_id: int) -> ProgramaCapacitaciones:
    programa = db.query(ProgramaCapacitaciones).filter(
        ProgramaCapacitaciones.id == programa_id
    ).first()
    if not programa:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Programa de capacitaciones no encontrado")
    return programa


def _compute_kpis(db: Session, programa: ProgramaCapacitaciones) -> List[KpiIndicador]:
    """Construye los 3 KpiIndicador desde los datos de indicadores_mensuales."""
    indicadores_db = db.query(CapacitacionIndicadorMensual).filter(
        CapacitacionIndicadorMensual.programa_id == programa.id
    ).all()

    ind_map: dict = {}
    for ind in indicadores_db:
        ind_map.setdefault(ind.tipo_indicador, {})[ind.mes] = ind

    metas = {
        'CUMPLIMIENTO': programa.meta_cumplimiento,
        'COBERTURA': programa.meta_cobertura,
        'EFICACIA': programa.meta_eficacia,
    }

    kpis = []
    for tipo in TIPO_ORDER:
        info = INDICADOR_INFO[tipo]
        meta = metas[tipo]
        meses_data = []
        total_num = 0.0
        total_den = 0.0

        # Análisis trimestrales
        analisis_t = {1: None, 2: None, 3: None, 4: None}
        for mes in range(1, 13):
            ind = ind_map.get(tipo, {}).get(mes)
            num = ind.numerador if ind else 0.0
            den = ind.denominador if ind else 0.0
            val = round((num / den) * 100, 1) if den > 0 else 0.0
            total_num += num
            total_den += den
            trimestre = (mes - 1) // 3 + 1
            if mes in (3, 6, 9, 12) and ind and ind.analisis_trimestral:
                analisis_t[trimestre] = ind.analisis_trimestral
            meses_data.append(KpiMesData(
                mes=mes,
                nombre_mes=NOMBRE_MESES[mes],
                valor=val,
                meta=meta,
                numerador=num,
                denominador=den,
                cumple=val >= meta,
            ))

        valor_global = round((total_num / total_den) * 100, 1) if total_den > 0 else 0.0
        kpis.append(KpiIndicador(
            tipo=tipo,
            nombre=info['nombre'],
            formula=info['formula'],
            meta=meta,
            frecuencia=info['frecuencia'],
            meses=meses_data,
            valor_global=valor_global,
            cumple_global=valor_global >= meta,
            analisis_t1=analisis_t[1],
            analisis_t2=analisis_t[2],
            analisis_t3=analisis_t[3],
            analisis_t4=analisis_t[4],
        ))
    return kpis


# =====================================================================
# PROGRAMAS (CRUD)
# =====================================================================

@router.get("/", response_model=List[ProgramaCapacitacionesResponse])
def listar_programas(
    año: Optional[int] = Query(None),
    empresa_id: Optional[int] = Query(None),
    estado: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(ProgramaCapacitaciones)
    if año:
        query = query.filter(ProgramaCapacitaciones.año == año)
    if empresa_id:
        query = query.filter(ProgramaCapacitaciones.empresa_id == empresa_id)
    if estado:
        query = query.filter(ProgramaCapacitaciones.estado == estado)
    return query.order_by(ProgramaCapacitaciones.año.desc()).all()


@router.post("/", response_model=ProgramaCapacitacionesResponse, status_code=status.HTTP_201_CREATED)
def crear_programa(
    data: ProgramaCapacitacionesCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    existente = db.query(ProgramaCapacitaciones).filter(
        ProgramaCapacitaciones.año == data.año,
        ProgramaCapacitaciones.empresa_id == data.empresa_id,
    ).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Ya existe un Programa de Capacitaciones para el año {data.año}"
        )
    programa = ProgramaCapacitaciones(**data.model_dump(), created_by=current_user.id)
    db.add(programa)
    db.commit()
    db.refresh(programa)
    return programa


@router.post("/crear-desde-plantilla", response_model=ProgramaCapacitacionesDetailResponse, status_code=status.HTTP_201_CREATED)
def crear_desde_plantilla(
    año: int = Query(..., description="Año del programa"),
    empresa_id: Optional[int] = Query(None),
    encargado_sgsst: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    """Crea un Programa de Capacitaciones con las ~35 actividades estándar del PR-SST-01."""
    existente = db.query(ProgramaCapacitaciones).filter(
        ProgramaCapacitaciones.año == año,
        ProgramaCapacitaciones.empresa_id == empresa_id,
    ).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Ya existe un Programa de Capacitaciones para el año {año}"
        )

    programa = ProgramaCapacitaciones(
        año=año,
        empresa_id=empresa_id,
        encargado_sgsst=encargado_sgsst,
        created_by=current_user.id,
    )
    db.add(programa)
    db.flush()

    # Crear actividades con 12 seguimientos cada una
    for act_data in get_plantilla_actividades(año):
        meses_programados = act_data.pop("meses_programados", [])
        actividad = CapacitacionActividad(programa_id=programa.id, **act_data)
        db.add(actividad)
        db.flush()
        for mes in range(1, 13):
            db.add(CapacitacionSeguimiento(
                actividad_id=actividad.id,
                mes=mes,
                programada=(mes in meses_programados),
                ejecutada=False,
            ))

    # Inicializar 36 filas de indicadores: 3 tipos × 12 meses
    metas = {
        'CUMPLIMIENTO': programa.meta_cumplimiento,
        'COBERTURA': programa.meta_cobertura,
        'EFICACIA': programa.meta_eficacia,
    }
    for tipo in TIPO_ORDER:
        for mes in range(1, 13):
            db.add(CapacitacionIndicadorMensual(
                programa_id=programa.id,
                tipo_indicador=tipo,
                mes=mes,
                numerador=0.0,
                denominador=0.0,
                valor_porcentaje=0.0,
                meta=metas[tipo],
            ))

    db.commit()
    db.refresh(programa)
    return programa


@router.get("/{programa_id}", response_model=ProgramaCapacitacionesDetailResponse)
def obtener_programa(
    programa_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return _get_programa_or_404(db, programa_id)


@router.put("/{programa_id}", response_model=ProgramaCapacitacionesResponse)
def actualizar_programa(
    programa_id: int,
    data: ProgramaCapacitacionesUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    programa = _get_programa_or_404(db, programa_id)
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(programa, field, value)
    programa.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(programa)
    return programa


@router.delete("/{programa_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_programa(
    programa_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    programa = _get_programa_or_404(db, programa_id)
    db.delete(programa)
    db.commit()


# =====================================================================
# ACTIVIDADES
# =====================================================================

@router.get("/{programa_id}/actividades", response_model=List[CapacitacionActividadResponse])
def listar_actividades(
    programa_id: int,
    ciclo: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_programa_or_404(db, programa_id)
    query = db.query(CapacitacionActividad).filter(
        CapacitacionActividad.programa_id == programa_id
    )
    if ciclo:
        query = query.filter(CapacitacionActividad.ciclo == ciclo)
    return query.order_by(CapacitacionActividad.orden).all()


@router.post("/{programa_id}/actividades", response_model=CapacitacionActividadResponse, status_code=status.HTTP_201_CREATED)
def crear_actividad(
    programa_id: int,
    data: CapacitacionActividadCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    _get_programa_or_404(db, programa_id)
    actividad = CapacitacionActividad(programa_id=programa_id, **data.model_dump())
    db.add(actividad)
    db.flush()
    for mes in range(1, 13):
        db.add(CapacitacionSeguimiento(actividad_id=actividad.id, mes=mes))
    db.commit()
    db.refresh(actividad)
    return actividad


@router.put("/actividades/{actividad_id}", response_model=CapacitacionActividadResponse)
def actualizar_actividad(
    actividad_id: int,
    data: CapacitacionActividadUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    actividad = db.query(CapacitacionActividad).filter(
        CapacitacionActividad.id == actividad_id
    ).first()
    if not actividad:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(actividad, field, value)
    actividad.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(actividad)
    return actividad


@router.delete("/actividades/{actividad_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_actividad(
    actividad_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin),
):
    actividad = db.query(CapacitacionActividad).filter(
        CapacitacionActividad.id == actividad_id
    ).first()
    if not actividad:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Actividad no encontrada")
    db.delete(actividad)
    db.commit()


# =====================================================================
# SEGUIMIENTO MENSUAL (P/E)
# =====================================================================

@router.put("/actividades/{actividad_id}/seguimiento/{mes}", response_model=CapacitacionSeguimientoResponse)
def actualizar_seguimiento(
    actividad_id: int,
    mes: int,
    data: CapacitacionSeguimientoUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El mes debe estar entre 1 y 12")

    seg = db.query(CapacitacionSeguimiento).filter(
        CapacitacionSeguimiento.actividad_id == actividad_id,
        CapacitacionSeguimiento.mes == mes,
    ).first()
    if not seg:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Seguimiento no encontrado")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(seg, field, value)
    seg.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(seg)
    return seg


# =====================================================================
# INDICADORES KPI MENSUALES
# =====================================================================

@router.get("/{programa_id}/indicadores", response_model=List[IndicadorMensualResponse])
def listar_indicadores_mensuales(
    programa_id: int,
    tipo_indicador: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_programa_or_404(db, programa_id)
    query = db.query(CapacitacionIndicadorMensual).filter(
        CapacitacionIndicadorMensual.programa_id == programa_id
    )
    if tipo_indicador:
        query = query.filter(CapacitacionIndicadorMensual.tipo_indicador == tipo_indicador)
    return query.order_by(
        CapacitacionIndicadorMensual.tipo_indicador,
        CapacitacionIndicadorMensual.mes
    ).all()


@router.put("/{programa_id}/indicadores/{tipo_indicador}/{mes}", response_model=IndicadorMensualResponse)
def actualizar_indicador_mensual(
    programa_id: int,
    tipo_indicador: str,
    mes: int,
    data: IndicadorMensualUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El mes debe estar entre 1 y 12")
    if tipo_indicador not in TIPO_ORDER:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Tipo inválido. Use: {TIPO_ORDER}")

    ind = db.query(CapacitacionIndicadorMensual).filter(
        CapacitacionIndicadorMensual.programa_id == programa_id,
        CapacitacionIndicadorMensual.tipo_indicador == tipo_indicador,
        CapacitacionIndicadorMensual.mes == mes,
    ).first()
    if not ind:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Indicador mensual no encontrado")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(ind, field, value)

    # Recalcular porcentaje
    if ind.denominador and ind.denominador > 0:
        ind.valor_porcentaje = round((ind.numerador / ind.denominador) * 100, 2)
    else:
        ind.valor_porcentaje = 0.0

    ind.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(ind)
    return ind


# =====================================================================
# DASHBOARD
# =====================================================================

@router.get("/{programa_id}/dashboard", response_model=DashboardCapacitaciones)
def obtener_dashboard(
    programa_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    programa = _get_programa_or_404(db, programa_id)

    actividades = db.query(CapacitacionActividad).filter(
        CapacitacionActividad.programa_id == programa_id
    ).all()
    act_ids = [a.id for a in actividades]

    seguimientos = db.query(CapacitacionSeguimiento).filter(
        CapacitacionSeguimiento.actividad_id.in_(act_ids)
    ).all() if act_ids else []

    total_programadas = sum(1 for s in seguimientos if s.programada)
    total_ejecutadas = sum(1 for s in seguimientos if s.ejecutada)
    porcentaje = round((total_ejecutadas / total_programadas) * 100, 1) if total_programadas > 0 else 0.0

    kpis = _compute_kpis(db, programa)

    return DashboardCapacitaciones(
        programa_id=programa_id,
        año=programa.año,
        total_actividades=len(actividades),
        actividades_programadas=total_programadas,
        actividades_ejecutadas=total_ejecutadas,
        porcentaje_cumplimiento=porcentaje,
        kpis=kpis,
    )


# =====================================================================
# EXPORTAR A EXCEL
# =====================================================================

@router.get("/{programa_id}/exportar/excel")
def exportar_excel(
    programa_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta el Programa de Capacitaciones a Excel (.xlsx)."""
    programa = _get_programa_or_404(db, programa_id)

    actividades = db.query(CapacitacionActividad).filter(
        CapacitacionActividad.programa_id == programa_id
    ).order_by(CapacitacionActividad.orden).all()

    act_ids = [a.id for a in actividades]
    segs_list = db.query(CapacitacionSeguimiento).filter(
        CapacitacionSeguimiento.actividad_id.in_(act_ids)
    ).all() if act_ids else []
    seg_dict: dict = {}
    for s in segs_list:
        seg_dict.setdefault(s.actividad_id, {})[s.mes] = s

    # Columnas fijas: N°, Actividad, Encargado, Recursos, Horas = 5
    # Meses P/E: 12 × 2 = 24 columnas
    # Consolidado: P, E, % = 3 columnas
    # Total = 32 columnas
    FIXED = 5
    TOTAL_COLS = FIXED + 24 + 3

    def p_col(m): return FIXED + (m - 1) * 2 + 1
    def e_col(m): return FIXED + (m - 1) * 2 + 2
    COL_TOT_P = FIXED + 24 + 1
    COL_TOT_E = FIXED + 24 + 2
    COL_PCT = FIXED + 24 + 3

    wb = openpyxl.Workbook()

    # ── Hoja 1: Cronograma ────────────────────────────────────────
    ws = wb.active
    ws.title = f"Cronograma {programa.año}"

    # Anchos de columna
    ws.column_dimensions['A'].width = 4   # N°
    ws.column_dimensions['B'].width = 48  # Actividad
    ws.column_dimensions['C'].width = 25  # Encargado
    ws.column_dimensions['D'].width = 20  # Recursos
    ws.column_dimensions['E'].width = 6   # Horas
    for m in range(1, 13):
        ws.column_dimensions[get_column_letter(p_col(m))].width = 3.5
        ws.column_dimensions[get_column_letter(e_col(m))].width = 3.5
    ws.column_dimensions[get_column_letter(COL_TOT_P)].width = 6
    ws.column_dimensions[get_column_letter(COL_TOT_E)].width = 6
    ws.column_dimensions[get_column_letter(COL_PCT)].width = 8

    # Fila 1: Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    c = ws.cell(row=1, column=1, value=f'PROGRAMA DE CAPACITACIONES PR-SST-01 — {programa.año}')
    _apply(c, _make_fill('0D47A1'), Font(color='FFFFFF', bold=True, size=14),
           Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 32

    # Fila 2: Metadatos
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.cell(row=2, column=1, value=f'Código: {programa.codigo}  |  Versión: {programa.version}  |  Estado: {(programa.estado or "").upper()}')
    ws.merge_cells(start_row=2, start_column=11, end_row=2, end_column=TOTAL_COLS)
    enc_txt = f'  |  Encargado: {programa.encargado_sgsst}' if programa.encargado_sgsst else ''
    ws.cell(row=2, column=11, value=f'Metas — Cumplimiento: {programa.meta_cumplimiento}%  Cobertura: {programa.meta_cobertura}%  Eficacia: {programa.meta_eficacia}%{enc_txt}')
    meta_fill = _make_fill('E3F2FD')
    meta_font = Font(size=9)
    for col in range(1, TOTAL_COLS + 1):
        _apply(ws.cell(row=2, column=col), meta_fill, meta_font, Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 4  # espaciado

    # Filas 4-5: Encabezados
    HDR1, HDR2 = 4, 5
    for col in [1, 2, 3, 4, 5, COL_TOT_P, COL_TOT_E, COL_PCT]:
        ws.merge_cells(start_row=HDR1, start_column=col, end_row=HDR2, end_column=col)
    labels = {1: 'N°', 2: 'Actividad', 3: 'Encargado', 4: 'Recursos', 5: 'Horas',
              COL_TOT_P: 'Total P', COL_TOT_E: 'Total E', COL_PCT: '% Cum.'}
    for col, label in labels.items():
        ws.cell(row=HDR1, column=col, value=label)

    for m in range(1, 13):
        pc, ec = p_col(m), e_col(m)
        ws.merge_cells(start_row=HDR1, start_column=pc, end_row=HDR1, end_column=ec)
        ws.cell(row=HDR1, column=pc, value=MESES_ABREV[m])
        ws.cell(row=HDR2, column=pc, value='P')
        ws.cell(row=HDR2, column=ec, value='E')

    hdr_fill = _make_fill('37474F')
    hdr_font = Font(color='FFFFFF', bold=True, size=9)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in [HDR1, HDR2]:
        for col in range(1, TOTAL_COLS + 1):
            _apply(ws.cell(row=row, column=col), hdr_fill, hdr_font, hdr_align)
    ws.row_dimensions[HDR1].height = 20
    ws.row_dimensions[HDR2].height = 14

    # Filas de datos agrupadas por ciclo PHVA
    acts_by_ciclo: dict = {}
    for act in actividades:
        acts_by_ciclo.setdefault(act.ciclo, []).append(act)

    current_row = 6
    nro = 1
    for ciclo in _CICLOS_ORDER:
        acts = acts_by_ciclo.get(ciclo, [])
        if not acts:
            continue

        # Fila separadora de ciclo
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=TOTAL_COLS)
        c = ws.cell(row=current_row, column=1, value=CICLO_LABELS.get(ciclo, ciclo))
        _apply(c, _make_fill(_CICLO_BG[ciclo]),
               Font(color='FFFFFF', bold=True, size=11),
               Alignment(horizontal='center', vertical='center'))
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for act in acts:
            segs = seg_dict.get(act.id, {})
            txt_align = Alignment(vertical='top', wrap_text=True)
            ctr_align = Alignment(horizontal='center', vertical='center')

            total_p = sum(1 for m in range(1, 13) if segs.get(m) and segs[m].programada)
            total_e = sum(1 for m in range(1, 13) if segs.get(m) and segs[m].ejecutada)
            pct = round((total_e / total_p) * 100) if total_p > 0 else 0

            for col, val in [
                (1, nro),
                (2, act.nombre),
                (3, act.encargado or ''),
                (4, act.recursos or ''),
                (5, act.horas or 0),
                (COL_TOT_P, total_p),
                (COL_TOT_E, total_e),
                (COL_PCT, f'{pct}%'),
            ]:
                _apply(ws.cell(row=current_row, column=col, value=val),
                       font=Font(size=8), align=txt_align)

            for m in range(1, 13):
                s = segs.get(m)
                prog = s.programada if s else False
                ejec = s.ejecutada if s else False
                p_fill = _make_fill('A5D6A7') if (prog and ejec) else (_make_fill('FFF9C4') if prog else None)
                e_fill = _make_fill('C8E6C9') if ejec else None
                pc, ec = p_col(m), e_col(m)
                _apply(ws.cell(row=current_row, column=pc, value='✓' if prog else ''),
                       fill=p_fill, font=Font(size=8, bold=True), align=ctr_align)
                _apply(ws.cell(row=current_row, column=ec, value='✓' if ejec else ''),
                       fill=e_fill, font=Font(size=8, bold=True, color='1B5E20'), align=ctr_align)

            ws.row_dimensions[current_row].height = 30
            current_row += 1
            nro += 1

    ws.freeze_panes = ws.cell(row=6, column=6)

    # ── Hoja 2: Indicadores ───────────────────────────────────────
    ws2 = wb.create_sheet(title="Indicadores")
    ws2.column_dimensions['A'].width = 18
    for m in range(1, 13):
        ws2.column_dimensions[get_column_letter(m + 1)].width = 10

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    c = ws2.cell(row=1, column=1, value=f'INDICADORES — PROGRAMA DE CAPACITACIONES {programa.año}')
    _apply(c, _make_fill('0D47A1'), Font(color='FFFFFF', bold=True, size=13),
           Alignment(horizontal='center', vertical='center'))
    ws2.row_dimensions[1].height = 28

    # Encabezados de mes
    ws2.cell(row=2, column=1, value='Indicador')
    for m in range(1, 13):
        c = ws2.cell(row=2, column=m + 1, value=NOMBRE_MESES[m])
        _apply(c, _make_fill('37474F'), Font(color='FFFFFF', bold=True, size=8),
               Alignment(horizontal='center', vertical='center'))
    _apply(ws2.cell(row=2, column=1), _make_fill('37474F'),
           Font(color='FFFFFF', bold=True, size=8),
           Alignment(horizontal='center', vertical='center'))
    ws2.row_dimensions[2].height = 18

    indicadores_db = db.query(CapacitacionIndicadorMensual).filter(
        CapacitacionIndicadorMensual.programa_id == programa_id
    ).all()
    ind_map: dict = {}
    for ind in indicadores_db:
        ind_map.setdefault(ind.tipo_indicador, {})[ind.mes] = ind

    metas = {
        'CUMPLIMIENTO': programa.meta_cumplimiento,
        'COBERTURA': programa.meta_cobertura,
        'EFICACIA': programa.meta_eficacia,
    }
    ind_colors = {'CUMPLIMIENTO': '1565C0', 'COBERTURA': '2E7D32', 'EFICACIA': 'E65100'}

    row = 3
    for tipo in TIPO_ORDER:
        meta = metas[tipo]
        info = INDICADOR_INFO[tipo]
        color = ind_colors[tipo]
        # Fila valor %
        c = ws2.cell(row=row, column=1, value=f'{info["nombre"]} (meta {meta}%)')
        _apply(c, _make_fill(color), Font(color='FFFFFF', bold=True, size=9),
               Alignment(horizontal='left', vertical='center', indent=1))
        ws2.row_dimensions[row].height = 16
        for m in range(1, 13):
            ind = ind_map.get(tipo, {}).get(m)
            val = ind.valor_porcentaje if ind else 0.0
            cumple = val >= meta
            fill = _make_fill('C8E6C9') if cumple else _make_fill('FFCDD2')
            c = ws2.cell(row=row, column=m + 1, value=f'{val:.1f}%')
            _apply(c, fill, Font(size=9, bold=True), Alignment(horizontal='center', vertical='center'))
        row += 1

        # Fila meta
        c = ws2.cell(row=row, column=1, value=f'Meta {info["nombre"]}')
        _apply(c, _make_fill('ECEFF1'), Font(size=8, italic=True),
               Alignment(horizontal='left', vertical='center', indent=1))
        for m in range(1, 13):
            c = ws2.cell(row=row, column=m + 1, value=f'{meta:.0f}%')
            _apply(c, _make_fill('ECEFF1'), Font(size=8, italic=True),
                   Alignment(horizontal='center', vertical='center'))
        ws2.row_dimensions[row].height = 14
        row += 2  # espacio entre indicadores

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"programa_capacitaciones_{programa.año}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =====================================================================
# EXPORTAR A PDF
# =====================================================================

@router.get("/{programa_id}/exportar/pdf")
def exportar_pdf(
    programa_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta el Programa de Capacitaciones a PDF (cronograma + fichas de indicadores)."""
    programa = _get_programa_or_404(db, programa_id)

    actividades = db.query(CapacitacionActividad).filter(
        CapacitacionActividad.programa_id == programa_id
    ).order_by(CapacitacionActividad.orden).all()

    act_ids = [a.id for a in actividades]
    segs_list = db.query(CapacitacionSeguimiento).filter(
        CapacitacionSeguimiento.actividad_id.in_(act_ids)
    ).all() if act_ids else []
    seg_dict: dict = {}
    for s in segs_list:
        seg_dict.setdefault(s.actividad_id, {})[s.mes] = s

    indicadores_db = db.query(CapacitacionIndicadorMensual).filter(
        CapacitacionIndicadorMensual.programa_id == programa_id
    ).all()
    ind_map: dict = {}
    for ind in indicadores_db:
        ind_map.setdefault(ind.tipo_indicador, {})[ind.mes] = ind

    metas = {
        'CUMPLIMIENTO': programa.meta_cumplimiento,
        'COBERTURA': programa.meta_cobertura,
        'EFICACIA': programa.meta_eficacia,
    }

    # ── Página 1: Cronograma ────────────────────────────────────────
    month_ths = ''.join(
        f'<th colspan="2" class="mes-hdr">{MESES_ABREV[m]}</th>' for m in range(1, 13)
    )
    pe_ths = '<th class="pe">P</th><th class="pe">E</th>' * 12

    acts_by_ciclo: dict = {}
    for act in actividades:
        acts_by_ciclo.setdefault(act.ciclo, []).append(act)

    body_rows = ''
    nro = 1
    for ciclo in _CICLOS_ORDER:
        acts = acts_by_ciclo.get(ciclo, [])
        if not acts:
            continue
        color = _CICLO_COLOR_HEX.get(ciclo, '#333')
        label = CICLO_LABELS.get(ciclo, ciclo)
        body_rows += (
            f'<tr><td colspan="31" class="ciclo-hdr" style="background:{color}">'
            f'{label}</td></tr>\n'
        )
        for act in acts:
            segs = seg_dict.get(act.id, {})
            total_p = sum(1 for m in range(1, 13) if segs.get(m) and segs[m].programada)
            total_e = sum(1 for m in range(1, 13) if segs.get(m) and segs[m].ejecutada)
            pct = round((total_e / total_p) * 100) if total_p > 0 else 0
            cells = ''
            for m in range(1, 13):
                s = segs.get(m)
                prog = s.programada if s else False
                ejec = s.ejecutada if s else False
                p_cls = 'pe-both' if (prog and ejec) else ('pe-prog' if prog else 'pe-empty')
                e_cls = 'pe-exec' if ejec else 'pe-empty'
                cells += (
                    f'<td class="pe {p_cls}">{"&#10003;" if prog else ""}</td>'
                    f'<td class="pe {e_cls}">{"&#10003;" if ejec else ""}</td>'
                )
            body_rows += (
                f'<tr>'
                f'<td class="nro">{nro}</td>'
                f'<td class="act-nombre">{act.nombre}</td>'
                f'<td class="encargado">{act.encargado or ""}</td>'
                f'<td class="horas">{act.horas or 0}</td>'
                f'{cells}'
                f'<td class="tot">{total_p}</td>'
                f'<td class="tot">{total_e}</td>'
                f'<td class="pct">{pct}%</td>'
                f'</tr>\n'
            )
            nro += 1

    enc_html = f'<span><b>Encargado:</b> {programa.encargado_sgsst}</span>' if programa.encargado_sgsst else ''
    apr_html = f'<span><b>Aprobado:</b> {programa.aprobado_por}</span>' if programa.aprobado_por else ''
    now = datetime.now().strftime('%d/%m/%Y')

    cronograma_html = f"""
  <div class="page page-cronograma">
    <h1>SISTEMA DE GESTIÓN DE SEGURIDAD Y SALUD EN EL TRABAJO</h1>
    <h2>PROGRAMA DE CAPACITACIONES &mdash; {programa.año}</h2>
    <div class="meta-bar">
      <span><b>Código:</b> {programa.codigo}</span>
      <span><b>Versión:</b> {programa.version}</span>
      <span><b>Estado:</b> {(programa.estado or '').upper()}</span>
      {enc_html}{apr_html}
    </div>
    <div class="meta-bar kpi-bar">
      <span><b>Meta Cumplimiento:</b> {programa.meta_cumplimiento}%</span>
      <span><b>Meta Cobertura:</b> {programa.meta_cobertura}%</span>
      <span><b>Meta Eficacia:</b> {programa.meta_eficacia}%</span>
      <span><b>Alcance:</b> {programa.alcance or ''}</span>
    </div>
    <table>
      <thead>
        <tr>
          <th rowspan="2" class="nro">N°</th>
          <th rowspan="2" class="act-nombre">Actividades</th>
          <th rowspan="2" class="encargado">Encargado</th>
          <th rowspan="2" class="horas">Hrs</th>
          {month_ths}
          <th rowspan="2" class="tot">P</th>
          <th rowspan="2" class="tot">E</th>
          <th rowspan="2" class="pct">%</th>
        </tr>
        <tr>{pe_ths}</tr>
      </thead>
      <tbody>
{body_rows}
      </tbody>
    </table>
    <div class="footer">Generado: {now} &bull; {programa.codigo} &bull; PR-SST-01</div>
  </div>
"""

    # ── Páginas 2-4: Fichas de Indicadores ─────────────────────────
    ind_pages_html = ''
    ind_colors_hex = {
        'CUMPLIMIENTO': '#1565C0',
        'COBERTURA': '#2E7D32',
        'EFICACIA': '#E65100',
    }
    for tipo in TIPO_ORDER:
        info = INDICADOR_INFO[tipo]
        meta = metas[tipo]
        color = ind_colors_hex[tipo]

        # Tabla mensual
        rows_html = ''
        all_vals = []
        for mes in range(1, 13):
            ind = ind_map.get(tipo, {}).get(mes)
            num = ind.numerador if ind else 0.0
            den = ind.denominador if ind else 0.0
            val = round((num / den) * 100, 1) if den > 0 else 0.0
            all_vals.append(val)
            cumple = val >= meta
            semaforo = '&#128994;' if cumple else '&#128308;'  # 🟢 / 🔴
            rows_html += f"""
          <tr>
            <td class="mes-nombre">{NOMBRE_MESES[mes]}</td>
            <td class="num-cell">{num:.0f}</td>
            <td class="den-cell">{den:.0f}</td>
            <td class="val-cell {'val-cumple' if cumple else 'val-nocumple'}">{val:.1f}%</td>
            <td class="meta-cell">{meta:.0f}%</td>
            <td class="sem-cell">{semaforo}</td>
          </tr>"""

        # Gráfico SVG simplificado
        svg_bars = ''
        bar_w = 20
        bar_gap = 5
        chart_h = 80
        for i, val in enumerate(all_vals):
            bar_h = int((val / 100) * chart_h) if val > 0 else 2
            x = i * (bar_w + bar_gap) + 5
            y = chart_h - bar_h + 10
            fill_c = '#4CAF50' if val >= meta else '#EF5350'
            svg_bars += f'<rect x="{x}" y="{y}" width="{bar_w}" height="{bar_h}" fill="{fill_c}" rx="2"/>'
            svg_bars += f'<text x="{x + bar_w // 2}" y="{y - 2}" text-anchor="middle" font-size="5" fill="#333">{val:.0f}%</text>'
            lbl = MESES_ABREV[i + 1] if i + 1 <= 12 else ''
            svg_bars += f'<text x="{x + bar_w // 2}" y="{chart_h + 20}" text-anchor="middle" font-size="5" fill="#555">{lbl}</text>'
        meta_line_y = chart_h - int((meta / 100) * chart_h) + 10
        chart_w = 12 * (bar_w + bar_gap) + 10
        svg = f"""<svg width="{chart_w}" height="{chart_h + 30}" xmlns="http://www.w3.org/2000/svg">
          {svg_bars}
          <line x1="5" y1="{meta_line_y}" x2="{chart_w - 5}" y2="{meta_line_y}" stroke="#F44336" stroke-width="1" stroke-dasharray="4,3"/>
          <text x="{chart_w - 4}" y="{meta_line_y - 1}" text-anchor="end" font-size="5" fill="#F44336">Meta {meta:.0f}%</text>
        </svg>"""

        # Análisis trimestrales
        analisis_rows = ''
        trimestre_labels = ['Primer Trimestre', 'Segundo Trimestre', 'Tercer Trimestre', 'Cuarto Trimestre']
        mes_trim = [3, 6, 9, 12]
        for t_idx, mes_t in enumerate(mes_trim):
            ind_t = ind_map.get(tipo, {}).get(mes_t)
            texto = (ind_t.analisis_trimestral or '') if ind_t else ''
            analisis_rows += f"""
            <tr>
              <td class="trim-label"><b>{trimestre_labels[t_idx]}:</b></td>
              <td class="trim-text">{texto}</td>
              <td class="trim-si">SI &square;</td>
              <td class="trim-no">NO &square;</td>
            </tr>"""

        ind_pages_html += f"""
  <div class="page page-indicador">
    <div class="ind-header" style="background:{color}">
      <span class="ind-codigo">PROGRAMA DE CAPACITACIONES — FICHA TÉCNICA INDICADORES</span>
      <span class="ind-fecha">{now}</span>
    </div>
    <table class="ind-meta-table">
      <tr><td class="lbl">Indicador</td><td class="val ind-num-label">Indicador: {info['nombre']}</td></tr>
      <tr><td class="lbl">Nombre</td><td class="val"><b>{info['nombre']}</b></td></tr>
      <tr><td class="lbl">Interpretación</td><td class="val">{info['nombre']} de Actividades en el programa</td></tr>
      <tr><td class="lbl">Factor que mide</td><td class="val">{info['nombre']}</td></tr>
      <tr><td class="lbl">Periodicidad</td><td class="val">Semestral — revisión con Actividades ejecutadas dentro del cronograma</td></tr>
      <tr><td class="lbl">Fuente</td><td class="val">Plan de trabajo</td></tr>
      <tr><td class="lbl">Responsable</td><td class="val">SST</td></tr>
      <tr><td class="lbl">Personas que deben conocer</td><td class="val">Alta gerencia — RRHH</td></tr>
      <tr><td class="lbl">Fórmula</td><td class="val"><em>{info['formula']}</em></td></tr>
    </table>

    <div class="ind-body">
      <div class="ind-table-col">
        <table class="ind-data-table">
          <thead>
            <tr>
              <th>Mes</th>
              <th>Numerador</th>
              <th>Denominador</th>
              <th>% Obtenido</th>
              <th>Meta</th>
              <th>Cumple</th>
            </tr>
          </thead>
          <tbody>{rows_html}</tbody>
        </table>
      </div>
      <div class="ind-chart-col">
        <div class="chart-title">Análisis Tendencial</div>
        {svg}
      </div>
    </div>

    <table class="analisis-table">
      <thead>
        <tr>
          <th colspan="2">Análisis Tendencial</th>
          <th>Plan de Acción</th>
          <th>Plazo</th>
          <th>Responsable</th>
          <th>Acción correctiva?</th>
        </tr>
      </thead>
      <tbody>{analisis_rows}</tbody>
    </table>
  </div>
"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Programa de Capacitaciones {programa.año}</title>
<style>
  @page {{ size: A4 landscape; margin: 7mm 6mm; }}
  @page :first {{ size: A4 landscape; margin: 7mm 6mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 6.5pt; color: #212121; }}

  .page {{ page-break-after: always; }}
  .page:last-child {{ page-break-after: avoid; }}

  /* ── Cronograma ── */
  h1 {{ text-align: center; font-size: 9.5pt; color: #0D47A1; margin-bottom: 1mm; font-weight: bold; }}
  h2 {{ text-align: center; font-size: 11pt; color: #0D47A1; margin-bottom: 2mm; font-weight: bold; }}
  .meta-bar {{ background: #E3F2FD; padding: 1.2mm 2mm; margin-bottom: 1.5mm; border-radius: 2px; }}
  .meta-bar span {{ margin-right: 4mm; }}
  .kpi-bar {{ background: #E8F5E9; }}
  table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
  th, td {{ border: 0.3pt solid #9E9E9E; padding: 0.6mm; vertical-align: top; overflow: hidden; word-break: break-word; }}
  thead th {{ background: #37474F; color: #fff; font-weight: bold; text-align: center; vertical-align: middle; }}
  .mes-hdr {{ font-size: 5.5pt; }}
  .pe {{ text-align: center; font-weight: bold; font-size: 6pt; width: 4mm; }}
  .ciclo-hdr {{ color: #fff; font-weight: bold; font-size: 7.5pt; text-align: center; padding: 1.5mm; }}
  .nro {{ width: 5mm; text-align: center; font-size: 5.5pt; }}
  .act-nombre {{ width: 55mm; font-size: 6pt; }}
  .encargado {{ width: 18mm; font-size: 5.5pt; }}
  .horas {{ width: 6mm; text-align: center; font-size: 5.5pt; }}
  .tot {{ width: 5mm; text-align: center; font-size: 5.5pt; font-weight: bold; }}
  .pct {{ width: 7mm; text-align: center; font-size: 5.5pt; font-weight: bold; }}
  .pe-prog {{ background: #FFF9C4; }}
  .pe-exec {{ background: #C8E6C9; color: #1B5E20; }}
  .pe-both {{ background: #A5D6A7; color: #1B5E20; }}
  .footer {{ margin-top: 2mm; font-size: 5pt; color: #757575; text-align: right; }}

  /* ── Fichas Indicadores ── */
  .page-indicador {{ padding: 2mm; }}
  .ind-header {{ color: #fff; padding: 2mm 3mm; margin-bottom: 2mm; border-radius: 2px; display: flex; justify-content: space-between; }}
  .ind-codigo {{ font-size: 8pt; font-weight: bold; }}
  .ind-fecha {{ font-size: 7pt; }}
  .ind-num-label {{ font-weight: bold; }}
  .ind-meta-table {{ width: 100%; margin-bottom: 3mm; font-size: 7pt; }}
  .ind-meta-table .lbl {{ background: #ECEFF1; font-weight: bold; width: 35mm; padding: 1mm 2mm; }}
  .ind-meta-table .val {{ padding: 1mm 2mm; }}
  .ind-body {{ display: flex; gap: 3mm; margin-bottom: 3mm; }}
  .ind-table-col {{ flex: 0 0 65%; }}
  .ind-chart-col {{ flex: 1; text-align: center; }}
  .chart-title {{ font-size: 7pt; font-weight: bold; text-align: center; margin-bottom: 1mm; color: #37474F; }}
  .ind-data-table {{ font-size: 7pt; }}
  .ind-data-table thead th {{ font-size: 6.5pt; }}
  .mes-nombre {{ width: 20mm; }}
  .num-cell, .den-cell {{ text-align: right; width: 16mm; }}
  .val-cell {{ text-align: center; font-weight: bold; width: 16mm; }}
  .val-cumple {{ background: #C8E6C9; color: #1B5E20; }}
  .val-nocumple {{ background: #FFCDD2; color: #B71C1C; }}
  .meta-cell {{ text-align: center; width: 12mm; }}
  .sem-cell {{ text-align: center; width: 10mm; }}
  .analisis-table {{ font-size: 6.5pt; }}
  .analisis-table thead th {{ font-size: 6pt; }}
  .trim-label {{ width: 30mm; background: #ECEFF1; font-size: 6.5pt; }}
  .trim-text {{ }}
  .trim-si, .trim-no {{ width: 12mm; text-align: center; }}
</style>
</head>
<body>
{cronograma_html}
{ind_pages_html}
</body>
</html>"""

    pdf_bytes = weasyprint.HTML(string=html).write_pdf(
        metadata={
            "title": f"Programa de Capacitaciones {programa.año}",
            "author": "Sistema SG-SST",
        },
        pdf_version=(1, 7),
        compress=True,
    )

    filename = f"programa_capacitaciones_{programa.año}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
