from typing import Any, List, Dict, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, date, timedelta
import os
import uuid

from app.database import get_db
from app.dependencies import get_current_user, get_current_active_user, require_admin, require_supervisor_or_admin
from app.models.user import User, UserRole
from app.models.worker import Worker, WorkerContract
from app.models.worker_document import WorkerDocument, DocumentCategory
from app.models.worker_novedad import WorkerNovedad, NovedadType, NovedadStatus
from app.services.s3_storage import s3_service
from app.utils.storage import storage_manager
from app.config import settings
from app.models.occupational_exam import OccupationalExam
from app.models.seguimiento import Seguimiento, EstadoSeguimiento
from app.models.reinduction import ReinductionRecord
from app.schemas.worker import (
    Worker as WorkerSchema,
    WorkerCreate,
    WorkerUpdate,
    WorkerList,
    WorkerContract as WorkerContractSchema,
    WorkerContractCreate,
    WorkerContractUpdate,
    WorkerDocumentResponse,
    WorkerDocumentUpdate
)
from app.schemas.worker_novedad import (
    WorkerNovedadCreate,
    WorkerNovedadUpdate,
    WorkerNovedadResponse,
    WorkerNovedadList,
    WorkerNovedadApproval,
    WorkerNovedadStats
)
from app.schemas.worker_vacation import (
    WorkerVacation,
    WorkerVacationCreate,
    WorkerVacationUpdate,
    VacationBalance,
    VacationAvailability,
    VacationStats,
    VacationStatus,
    VacationRequestWithWorker
)
from app.schemas.occupational_exam import (
    OccupationalExamCreate,
    OccupationalExamUpdate,
    OccupationalExamResponse,
    OccupationalExamListResponse
)
from app.schemas.common import MessageResponse
from app.models.enrollment import Enrollment
from app.models.course import Course
from app.models.survey import Survey, UserSurvey
from app.models.evaluation import Evaluation, UserEvaluation
from app.models.reinduction import ReinductionRecord

router = APIRouter()


# ==================== ESTADÍSTICAS DE TRABAJADORES ====================
@router.get("/stats", response_model=Dict[str, Any])
@router.get("/stats{trailing_slash:path}", response_model=Dict[str, Any])
async def get_worker_stats(
    trailing_slash: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Obtener estadísticas generales de trabajadores:
    - Totales activos e inactivos
    - Distribución por modalidad de trabajo
    """

    total_workers = db.query(func.count(Worker.id)).scalar() or 0
    active_workers = db.query(func.count(Worker.id)).filter(Worker.is_active == True).scalar() or 0
    inactive_workers = db.query(func.count(Worker.id)).filter(Worker.is_active == False).scalar() or 0

    modality_rows = (
        db.query(Worker.work_modality, func.count(Worker.id))
        .group_by(Worker.work_modality)
        .all()
    )

    by_modality: Dict[str, int] = {}
    for modality, count in modality_rows:
        key = modality.value if hasattr(modality, "value") and modality is not None else (modality or "UNKNOWN")
        by_modality[str(key)] = count

    # Asegurar claves presentes aunque no existan registros
    default_modalities = [
        "ON_SITE",
        "REMOTE",
        "TELEWORK",
        "HOME_OFFICE",
        "MOBILE",
        "UNKNOWN",
    ]
    for m in default_modalities:
        by_modality.setdefault(m, 0)

    # Porcentajes por modalidad (basado en total_workers)
    percent_by_modality: Dict[str, float] = {}
    denom = total_workers if total_workers > 0 else 1
    for m, c in by_modality.items():
        percent_by_modality[m] = round((c / denom) * 100, 2)

    return {
        "total_workers": total_workers,
        "active_workers": active_workers,
        "inactive_workers": inactive_workers,
        "by_modality": by_modality,
        "percent_by_modality": percent_by_modality,
    }

@router.get("/", response_model=List[WorkerList])
@router.get("", response_model=List[WorkerList])
async def get_workers(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    search: str = Query(None, description="Buscar por nombre, documento o email"),
    is_active: bool = Query(None, description="Filtrar por estado activo"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Obtener lista de trabajadores con filtros opcionales
    """
    query = db.query(Worker)
    
    # Filtro de búsqueda
    if search:
        search_filter = or_(
            Worker.first_name.ilike(f"%{search}%"),
            Worker.last_name.ilike(f"%{search}%"),
            Worker.document_number.ilike(f"%{search}%"),
            Worker.email.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    # Filtro por estado activo
    if is_active is not None:
        query = query.filter(Worker.is_active == is_active)
    
    workers = query.offset(skip).limit(limit).all()
    return workers


@router.post("/", response_model=WorkerSchema)
@router.post("", response_model=WorkerSchema)
async def create_worker(
    worker_data: WorkerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Crear un nuevo trabajador
    """
    # Verificar si ya existe un trabajador con el mismo documento o email
    existing_worker = db.query(Worker).filter(
        or_(
            Worker.document_number == worker_data.document_number,
            Worker.email == worker_data.email
        )
    ).first()
    
    if existing_worker:
        if existing_worker.document_number == worker_data.document_number:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe un trabajador con este número de documento"
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe un trabajador con este email"
            )
    
    # Crear el trabajador
    worker = Worker(**worker_data.dict())
    db.add(worker)
    db.commit()
    db.refresh(worker)
    
    return worker


@router.get("/basic{trailing_slash:path}", response_model=List[WorkerList])
async def get_workers_basic(
    trailing_slash: str = "",
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    search: str = Query(None, description="Buscar por nombre, documento o email"),
    is_active: bool = Query(None, description="Filtrar por estado activo"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """
    Obtener lista básica de trabajadores (sin restricciones de rol)
    """
    query = db.query(Worker)
    
    # Filtro de búsqueda
    if search:
        search_filter = or_(
            Worker.first_name.ilike(f"%{search}%"),
            Worker.last_name.ilike(f"%{search}%"),
            Worker.document_number.ilike(f"%{search}%"),
            Worker.email.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    # Filtro por estado activo
    if is_active is not None:
        query = query.filter(Worker.is_active == is_active)
    
    workers = query.offset(skip).limit(limit).all()
    return workers


# ==================== ENDPOINT PARA USUARIO AUTENTICADO ====================

@router.get("/me", response_model=WorkerSchema)
async def get_current_worker(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
) -> Any:
    """Obtener información del worker del usuario autenticado"""
    worker = db.query(Worker).filter(Worker.user_id == current_user.id).first()
    
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encontró información de trabajador para este usuario"
        )
    
    return worker


@router.get("/{worker_id}", response_model=WorkerSchema)
async def get_worker(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Obtener un trabajador por ID
    """
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    return worker


@router.get("/{worker_id}/detailed-info")
async def get_worker_detailed_info(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Obtener información detallada de un trabajador incluyendo cursos, encuestas y evaluaciones
    """
    # Obtener información del trabajador
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    
    # Get course enrollments with progress
    enrollments = db.query(Enrollment).filter(Enrollment.user_id == worker.user_id).all()
    courses = []
    for enrollment in enrollments:
        course = enrollment.course
        
        courses.append({
            "enrollment_id": enrollment.id,
            "course_id": course.id,
            "course_title": course.title,
            "enrollment_date": enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else None,
            "completion_date": enrollment.completed_at.isoformat() if enrollment.completed_at else None,
            "status": enrollment.status,
            "progress_percentage": enrollment.progress or 0
        })
    
    # Obtener encuestas del trabajador
    surveys_query = db.query(
        UserSurvey.id.label('user_survey_id'),
        Survey.id.label('survey_id'),
        Survey.title.label('survey_title'),
        UserSurvey.completed_at,
        UserSurvey.status
    ).join(
        Survey, UserSurvey.survey_id == Survey.id
    ).filter(
        UserSurvey.user_id == worker.user_id
    ).all()
    
    surveys = []
    for survey in surveys_query:
        surveys.append({
            "user_survey_id": survey.user_survey_id,
            "survey_id": survey.survey_id,
            "survey_title": survey.survey_title,
            "status": survey.status,
            "completed_at": survey.completed_at.isoformat() if survey.completed_at else None
        })
    
    # Obtener evaluaciones del trabajador
    evaluations_query = db.query(
        UserEvaluation.id.label('user_evaluation_id'),
        Evaluation.id.label('evaluation_id'),
        Evaluation.title.label('evaluation_title'),
        UserEvaluation.score,
        UserEvaluation.max_points,
        UserEvaluation.passed,
        UserEvaluation.completed_at,
        UserEvaluation.status
    ).join(
        Evaluation, UserEvaluation.evaluation_id == Evaluation.id
    ).filter(
        UserEvaluation.user_id == worker.user_id
    ).all()
    
    evaluations = []
    for evaluation in evaluations_query:
        evaluations.append({
            "user_evaluation_id": evaluation.user_evaluation_id,
            "evaluation_id": evaluation.evaluation_id,
            "evaluation_title": evaluation.evaluation_title,
            "score": evaluation.score,
            "max_points": evaluation.max_points,
            "passed": evaluation.passed,
            "completed_at": evaluation.completed_at.isoformat() if evaluation.completed_at else None,
            "status": evaluation.status
        })
    
    return {
        "worker": {
            "id": worker.id,
            "full_name": f"{worker.first_name} {worker.last_name}",
            "document_number": worker.document_number,
            "email": worker.email,
            "position": worker.position,
            "department": worker.department,
            "is_active": worker.is_active,
            "user_id": worker.user_id
        },
        "courses": courses,
        "surveys": surveys,
        "evaluations": evaluations
    }


@router.put("/{worker_id}", response_model=WorkerSchema)
async def update_worker(
    worker_id: int,
    worker_data: WorkerUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Actualizar un trabajador
    """
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    
    # Verificar duplicados si se actualiza documento o email
    update_data = worker_data.dict(exclude_unset=True)
    
    if "document_number" in update_data or "email" in update_data:
        existing_worker = db.query(Worker).filter(
            Worker.id != worker_id,
            or_(
                Worker.document_number == update_data.get("document_number", worker.document_number),
                Worker.email == update_data.get("email", worker.email)
            )
        ).first()
        
        if existing_worker:
            if existing_worker.document_number == update_data.get("document_number"):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Ya existe otro trabajador con este número de documento"
                )
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Ya existe otro trabajador con este email"
                )
    
    # Verificar si se está actualizando la fecha de ingreso
    fecha_ingreso_changed = "fecha_de_ingreso" in update_data and update_data["fecha_de_ingreso"] != worker.fecha_de_ingreso
    
    # Actualizar campos
    for field, value in update_data.items():
        setattr(worker, field, value)
    
    db.commit()
    db.refresh(worker)
    
    # Si se actualizó la fecha de ingreso, regenerar registros de reinducción
    if fecha_ingreso_changed and worker.fecha_de_ingreso:
        try:
            from app.services.reinduction_service import ReinductionService
            from app.models.reinduction import ReinductionRecord, ReinductionStatus
            
            service = ReinductionService(db)
            
            # Eliminar registros existentes que no estén completados
            existing_records = db.query(ReinductionRecord).filter(
                ReinductionRecord.worker_id == worker_id,
                ReinductionRecord.status.in_([
                    ReinductionStatus.PENDING,
                    ReinductionStatus.SCHEDULED,
                    ReinductionStatus.IN_PROGRESS,
                    ReinductionStatus.OVERDUE
                ])
            ).all()
            
            for record in existing_records:
                db.delete(record)
            
            db.commit()
            
            # Regenerar registros faltantes
            service.generate_missing_reinduction_records(worker_id=worker_id)
            
            print(f"Registros de reinducción regenerados para trabajador {worker_id} debido a cambio en fecha de ingreso")
            
        except Exception as e:
            print(f"Error al regenerar registros de reinducción para trabajador {worker_id}: {str(e)}")
            # No fallar la actualización del trabajador por este error
    
    return worker


@router.delete("/{worker_id}", response_model=MessageResponse)
async def delete_worker(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
) -> Any:
    """
    Eliminar un trabajador (solo administradores)
    """
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    
    db.delete(worker)
    db.commit()
    
    return MessageResponse(message="Trabajador eliminado exitosamente")


# Endpoints para contratos
@router.post("/{worker_id}/contracts", response_model=WorkerContractSchema)
async def create_worker_contract(
    worker_id: int,
    contract_data: WorkerContractCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Crear un nuevo contrato para un trabajador
    """
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    
    contract = WorkerContract(
        worker_id=worker_id,
        **contract_data.dict()
    )
    db.add(contract)
    db.commit()
    db.refresh(contract)
    
    return contract


@router.get("/{worker_id}/contracts", response_model=List[WorkerContractSchema])
async def get_worker_contracts(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Obtener todos los contratos de un trabajador
    """
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    
    return worker.contracts


@router.put("/contracts/{contract_id}", response_model=WorkerContractSchema)
async def update_worker_contract(
    contract_id: int,
    contract_data: WorkerContractUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Actualizar un contrato de trabajador
    """
    contract = db.query(WorkerContract).filter(WorkerContract.id == contract_id).first()
    if not contract:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Contrato no encontrado"
        )
    
    update_data = contract_data.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(contract, field, value)
    
    db.commit()
    db.refresh(contract)
    
    return contract


@router.delete("/contracts/{contract_id}", response_model=MessageResponse)
async def delete_worker_contract(
    contract_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
) -> Any:
    """
    Eliminar un contrato de trabajador (solo administradores)
    """
    contract = db.query(WorkerContract).filter(WorkerContract.id == contract_id).first()
    if not contract:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Contrato no encontrado"
        )
    
    db.delete(contract)
    db.commit()
    
    return MessageResponse(message="Contrato eliminado exitosamente")


@router.get("/check-document/{document_number}", response_model=Dict[str, Any])
async def check_worker_document(
    document_number: str,
    db: Session = Depends(get_db)
) -> Any:
    """
    Verificar si un número de documento corresponde a un trabajador registrado
    y devolver sus datos completos
    """
    worker = db.query(Worker).filter(
        Worker.document_number == document_number,
        Worker.is_active == True
    ).first()
    
    if not worker:
        return {
            "exists": False,
            "worker": None
        }
    
    return {
        "exists": True,
        "is_registered": worker.is_registered,
        "assigned_role": worker.assigned_role,
        "worker": {
            "id": worker.id,
            "first_name": worker.first_name,
            "last_name": worker.last_name,
            "email": worker.email,
            "phone": worker.phone,
            "department": worker.department,
            "position": worker.position,
            "fecha_de_ingreso": worker.fecha_de_ingreso.isoformat() if worker.fecha_de_ingreso else None,
            "document_type": worker.document_type,
            "document_number": worker.document_number
        }
    }


@router.post("/validate-employee", response_model=Dict[str, Any])
async def validate_employee_credentials(
    credentials: Dict[str, str],
    db: Session = Depends(get_db)
) -> Any:
    """
    Validar que tanto el número de documento como el correo electrónico
    correspondan a un trabajador registrado por el administrador
    """
    document_number = credentials.get("document_number")
    email = credentials.get("email")
    
    if not document_number or not email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Se requiere número de documento y correo electrónico"
        )
    
    worker = db.query(Worker).filter(
        Worker.document_number == document_number,
        Worker.email == email,
        Worker.is_active == True
    ).first()
    
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encontró un trabajador activo con ese número de documento y correo electrónico. Solo los empleados registrados por el administrador pueden crear una cuenta."
        )
    
    # Check if worker is already registered with a verified account
    if worker.is_registered and worker.user_id:
        # Check if the existing user is already verified
        from app.models.user import User
        existing_user = db.query(User).filter(User.id == worker.user_id).first()
        if existing_user and existing_user.is_verified:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Este empleado ya tiene una cuenta registrada en el sistema. Puede iniciar sesión directamente."
            )
    
    return {
        "valid": True,
        "worker_id": worker.id,
        "assigned_role": worker.assigned_role,
        "full_name": worker.full_name
    }


# Endpoints para exámenes ocupacionales
@router.post("/{worker_id}/occupational-exams", response_model=OccupationalExamResponse)
async def create_occupational_exam(
    worker_id: int,
    exam_data: OccupationalExamCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Crear un nuevo examen ocupacional para un trabajador
    """
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    
    exam = OccupationalExam(
        worker_id=worker_id,
        **exam_data.dict(exclude={"worker_id"})
    )
    db.add(exam)
    db.commit()
    db.refresh(exam)
    
    # Crear seguimiento automáticamente si se asigna un programa
    if exam.programa and exam.programa.strip():
        # Verificar si ya existe un seguimiento activo para este programa
        existing_seguimiento = db.query(Seguimiento).filter(
            Seguimiento.worker_id == worker_id,
            Seguimiento.programa == exam.programa,
            Seguimiento.estado == EstadoSeguimiento.INICIADO
        ).first()
        
        if not existing_seguimiento:
            # Crear nuevo seguimiento automáticamente
            seguimiento = Seguimiento(
                worker_id=worker_id,
                programa=exam.programa,
                nombre_trabajador=worker.full_name,
                cedula=worker.document_number,
                cargo=worker.position,
                fecha_ingreso=worker.fecha_de_ingreso,
                estado=EstadoSeguimiento.INICIADO,
                # Copiar datos del examen ocupacional
                conclusiones_ocupacionales=exam.occupational_conclusions,
                conductas_ocupacionales_prevenir=exam.preventive_occupational_behaviors,
                recomendaciones_generales=exam.general_recommendations,
                observaciones_examen=exam.observations
            )
            db.add(seguimiento)
            db.commit()
    
    # Calcular fecha del próximo examen (1 año después)
    from datetime import timedelta
    next_exam_date = exam.exam_date + timedelta(days=365)
    
    # Mapear medical_aptitude_concept a result legacy
    result_mapping = {
        "apto": "apto",
        "apto_con_recomendaciones": "apto_con_restricciones",
        "no_apto": "no_apto"
    }
    
    # Crear respuesta enriquecida con datos del trabajador
    exam_dict = {
        "id": exam.id,
        "exam_type": exam.exam_type,
        "exam_date": exam.exam_date,
        "programa": exam.programa,
        "occupational_conclusions": exam.occupational_conclusions,
        "preventive_occupational_behaviors": exam.preventive_occupational_behaviors,
        "general_recommendations": exam.general_recommendations,
        "medical_aptitude_concept": exam.medical_aptitude_concept,
        "observations": exam.observations,
        "examining_doctor": exam.examining_doctor,
        "medical_center": exam.medical_center,
        "worker_id": exam.worker_id,
        "worker_name": worker.full_name if worker else None,
        "worker_document": worker.document_number if worker else None,
        "worker_position": worker.position if worker else None,
        "worker_hire_date": worker.fecha_de_ingreso.isoformat() if worker and worker.fecha_de_ingreso else None,
        "next_exam_date": next_exam_date.isoformat(),
        
        # Campos legacy para compatibilidad con el frontend
        "status": "realizado",
        "result": result_mapping.get(exam.medical_aptitude_concept, "pendiente"),
        "restrictions": exam.general_recommendations if exam.medical_aptitude_concept == "apto_con_recomendaciones" else None,
        
        "created_at": exam.created_at,
        "updated_at": exam.updated_at
    }
    
    return exam_dict


@router.get("/{worker_id}/occupational-exams", response_model=List[OccupationalExamResponse])
async def get_worker_occupational_exams(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Obtener todos los exámenes ocupacionales de un trabajador ordenados por fecha descendente
    """
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    
    # Ordenar exámenes por fecha descendente (más reciente primero)
    exams = db.query(OccupationalExam).filter(
        OccupationalExam.worker_id == worker_id
    ).order_by(OccupationalExam.exam_date.desc()).all()
    
    return exams


@router.get("/occupational-exams/{exam_id}", response_model=OccupationalExamResponse)
async def get_occupational_exam(
    exam_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Obtener un examen ocupacional por ID con información del trabajador
    """
    exam = db.query(OccupationalExam).join(Worker).filter(OccupationalExam.id == exam_id).first()
    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Examen ocupacional no encontrado"
        )
    
    # Calcular fecha del próximo examen (1 año después)
    from datetime import timedelta
    next_exam_date = exam.exam_date + timedelta(days=365)
    
    # Mapear medical_aptitude_concept a result legacy
    result_mapping = {
        "apto": "apto",
        "apto_con_recomendaciones": "apto_con_restricciones",
        "no_apto": "no_apto"
    }
    
    # Crear respuesta enriquecida con datos del trabajador
    worker = exam.worker
    exam_dict = {
        "id": exam.id,
        "exam_type": exam.exam_type,
        "exam_date": exam.exam_date,
        "programa": exam.programa,
        "occupational_conclusions": exam.occupational_conclusions,
        "preventive_occupational_behaviors": exam.preventive_occupational_behaviors,
        "general_recommendations": exam.general_recommendations,
        "medical_aptitude_concept": exam.medical_aptitude_concept,
        "observations": exam.observations,
        "examining_doctor": exam.examining_doctor,
        "medical_center": exam.medical_center,
        "worker_id": exam.worker_id,
        "worker_name": worker.full_name if worker else None,
        "worker_document": worker.document_number if worker else None,
        "worker_position": worker.position if worker else None,
        "worker_hire_date": worker.fecha_de_ingreso.isoformat() if worker and worker.fecha_de_ingreso else None,
        "next_exam_date": next_exam_date.isoformat(),
        
        # Campos legacy para compatibilidad con el frontend
        "status": "realizado",
        "result": result_mapping.get(exam.medical_aptitude_concept, "pendiente"),
        "restrictions": exam.general_recommendations if exam.medical_aptitude_concept == "apto_con_recomendaciones" else None,
        
        "created_at": exam.created_at,
        "updated_at": exam.updated_at
    }
    
    return exam_dict


@router.put("/occupational-exams/{exam_id}", response_model=OccupationalExamResponse)
async def update_occupational_exam(
    exam_id: int,
    exam_data: OccupationalExamUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Actualizar un examen ocupacional
    """
    exam = db.query(OccupationalExam).filter(OccupationalExam.id == exam_id).first()
    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Examen ocupacional no encontrado"
        )
    
    # Guardar el programa anterior para comparar
    old_programa = exam.programa
    
    update_data = exam_data.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(exam, field, value)
    
    db.commit()
    db.refresh(exam)
    
    # Crear seguimiento automáticamente si se asigna un programa nuevo
    if exam.programa and exam.programa.strip() and exam.programa != old_programa:
        worker = db.query(Worker).filter(Worker.id == exam.worker_id).first()
        if worker:
            # Verificar si ya existe un seguimiento activo para este programa
            existing_seguimiento = db.query(Seguimiento).filter(
                Seguimiento.worker_id == exam.worker_id,
                Seguimiento.programa == exam.programa,
                Seguimiento.estado == EstadoSeguimiento.INICIADO
            ).first()
            
            if not existing_seguimiento:
                # Crear nuevo seguimiento automáticamente
                seguimiento = Seguimiento(
                    worker_id=exam.worker_id,
                    programa=exam.programa,
                    nombre_trabajador=worker.full_name,
                    cedula=worker.document_number,
                    cargo=worker.position,
                    fecha_ingreso=worker.fecha_de_ingreso,
                    estado=EstadoSeguimiento.INICIADO,
                    # Copiar datos del examen ocupacional
                    conclusiones_ocupacionales=exam.occupational_conclusions,
                    conductas_ocupacionales_prevenir=exam.preventive_occupational_behaviors,
                    recomendaciones_generales=exam.general_recommendations,
                    observaciones_examen=exam.observations
                )
                db.add(seguimiento)
                db.commit()
    
    # Calcular fecha del próximo examen (1 año después)
    from datetime import timedelta
    next_exam_date = exam.exam_date + timedelta(days=365)
    
    # Obtener información del trabajador para la respuesta
    worker = db.query(Worker).filter(Worker.id == exam.worker_id).first()
    
    # Mapear medical_aptitude_concept a result legacy
    result_mapping = {
        "apto": "apto",
        "apto_con_recomendaciones": "apto_con_restricciones",
        "no_apto": "no_apto"
    }
    
    # Crear respuesta enriquecida con datos del trabajador
    exam_dict = {
        "id": exam.id,
        "exam_type": exam.exam_type,
        "exam_date": exam.exam_date,
        "programa": exam.programa,
        "occupational_conclusions": exam.occupational_conclusions,
        "preventive_occupational_behaviors": exam.preventive_occupational_behaviors,
        "general_recommendations": exam.general_recommendations,
        "medical_aptitude_concept": exam.medical_aptitude_concept,
        "observations": exam.observations,
        "examining_doctor": exam.examining_doctor,
        "medical_center": exam.medical_center,
        "worker_id": exam.worker_id,
        "worker_name": worker.full_name if worker else None,
        "worker_document": worker.document_number if worker else None,
        "worker_position": worker.position if worker else None,
        "worker_hire_date": worker.fecha_de_ingreso.isoformat() if worker and worker.fecha_de_ingreso else None,
        "next_exam_date": next_exam_date.isoformat(),
        
        # Campos legacy para compatibilidad con el frontend
        "status": "realizado",
        "result": result_mapping.get(exam.medical_aptitude_concept, "pendiente"),
        "restrictions": exam.general_recommendations if exam.medical_aptitude_concept == "apto_con_recomendaciones" else None,
        
        "created_at": exam.created_at,
        "updated_at": exam.updated_at
    }
    
    return exam_dict


@router.delete("/occupational-exams/{exam_id}", response_model=MessageResponse)
async def delete_occupational_exam(
    exam_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
) -> Any:
    """
    Eliminar un examen ocupacional (solo administradores)
    """
    exam = db.query(OccupationalExam).filter(OccupationalExam.id == exam_id).first()
    if not exam:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Examen ocupacional no encontrado"
        )
    
    # Eliminar archivo PDF de S3 Storage si existe
    if exam.pdf_file_path:
        try:
            # Extraer la clave del archivo desde la URL de S3
            if "s3.amazonaws.com" in exam.pdf_file_path:
                file_key = exam.pdf_file_path.split('.com/')[-1]
                result = s3_service.delete_file(file_key)
                if result["success"]:
                    logger.info(f"PDF eliminado de S3 Storage: {exam.pdf_file_path}")
                else:
                    logger.warning(f"Error al eliminar PDF de S3 Storage: {result.get('error', 'Error desconocido')}")
            else:
                logger.warning(f"URL de archivo no válida para S3 Storage: {exam.pdf_file_path}")
        except Exception as e:
            logger.warning(f"Error al eliminar PDF de S3 Storage: {str(e)}")
            # No fallar la eliminación del examen si no se puede eliminar el PDF
    
    db.delete(exam)
    db.commit()
    
    return MessageResponse(message="Examen ocupacional eliminado exitosamente")


@router.get("/export/excel")
async def export_workers_to_excel(
    search: str = Query(None, description="Buscar por nombre, documento o email"),
    is_active: bool = Query(None, description="Filtrar por estado activo"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> StreamingResponse:
    """
    Exportar trabajadores a Excel
    """
    # Obtener trabajadores con los mismos filtros que el endpoint principal
    query = db.query(Worker)
    
    # Filtro de búsqueda
    if search:
        search_filter = or_(
            Worker.first_name.ilike(f"%{search}%"),
            Worker.last_name.ilike(f"%{search}%"),
            Worker.document_number.ilike(f"%{search}%"),
            Worker.email.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    # Filtro por estado activo
    if is_active is not None:
        query = query.filter(Worker.is_active == is_active)
    
    workers = query.all()
    
    # Crear el archivo Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Trabajadores"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        "ID", "Nombre", "Apellido", "Tipo Documento", "Número Documento", 
        "Email", "Teléfono", "Fecha Nacimiento", "Género", "Ciudad", 
        "Departamento", "País", "Cargo", "Fecha Ingreso", "Tipo Contrato", 
        "Modalidad Trabajo", "Profesión", "Nivel Riesgo", "Ocupación", 
        "Salario IBC", "EPS", "AFP", "ARL", "Tipo Sangre", "Observaciones", 
        "Estado", "Rol Asignado", "Fecha Creación"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos de trabajadores
    for row, worker in enumerate(workers, 2):
        worksheet.cell(row=row, column=1, value=worker.id)
        worksheet.cell(row=row, column=2, value=worker.first_name)
        worksheet.cell(row=row, column=3, value=worker.last_name)
        worksheet.cell(row=row, column=4, value=worker.document_type)
        worksheet.cell(row=row, column=5, value=worker.document_number)
        worksheet.cell(row=row, column=6, value=worker.email)
        worksheet.cell(row=row, column=7, value=worker.phone)
        worksheet.cell(row=row, column=8, value=worker.birth_date.strftime("%Y-%m-%d") if worker.birth_date else "")
        worksheet.cell(row=row, column=9, value=worker.gender)
        worksheet.cell(row=row, column=10, value=worker.city)
        worksheet.cell(row=row, column=11, value=worker.department)
        worksheet.cell(row=row, column=12, value=worker.country)
        worksheet.cell(row=row, column=13, value=worker.position)
        worksheet.cell(row=row, column=14, value=worker.fecha_de_ingreso.strftime("%Y-%m-%d") if worker.fecha_de_ingreso else "")
        worksheet.cell(row=row, column=15, value=worker.contract_type)
        worksheet.cell(row=row, column=16, value=worker.work_modality)
        worksheet.cell(row=row, column=17, value=worker.profession)
        worksheet.cell(row=row, column=18, value=worker.risk_level)
        worksheet.cell(row=row, column=19, value=worker.occupation)
        worksheet.cell(row=row, column=20, value=float(worker.salary_ibc) if worker.salary_ibc else "")
        worksheet.cell(row=row, column=21, value=worker.eps)
        worksheet.cell(row=row, column=22, value=worker.afp)
        worksheet.cell(row=row, column=23, value=worker.arl)
        worksheet.cell(row=row, column=24, value=worker.blood_type)
        worksheet.cell(row=row, column=25, value=worker.observations)
        worksheet.cell(row=row, column=26, value="Activo" if worker.is_active else "Inactivo")
        worksheet.cell(row=row, column=27, value=worker.assigned_role)
        worksheet.cell(row=row, column=28, value=worker.created_at.strftime("%Y-%m-%d %H:%M:%S") if worker.created_at else "")
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generar nombre de archivo con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"trabajadores_{timestamp}.xlsx"
    
    # Retornar como respuesta de streaming
    return StreamingResponse(
        BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# ENDPOINTS DE ALMACENAMIENTO S3
# =============================================================================

@router.post("/{worker_id}/documents/upload-s3", response_model=Dict[str, Any])
async def upload_worker_document_s3(
    worker_id: int,
    file: UploadFile = File(...),
    document_type: str = Form("general"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Subir documento de empleado a S3.
    
    Args:
        worker_id: ID del trabajador
        file: Archivo a subir
        document_type: Tipo de documento (cedula, contrato, hoja_vida, etc.)
        current_user: Usuario actual
        db: Sesión de base de datos
    
    Returns:
        Información del archivo subido
    """
    from app.services.s3_storage import upload_employee_document
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Verificar permisos
    if not has_permission(current_user, "workers", "update"):
        raise HTTPException(status_code=403, detail="No tienes permisos para subir documentos")
    
    try:
        # Subir archivo a S3
        result = await upload_employee_document(worker_id, file, document_type)
        
        if result["success"]:
            logger.info(f"Documento subido exitosamente para trabajador {worker_id}: {result['file_key']}")
            return {
                "success": True,
                "message": "Documento subido exitosamente",
                "data": result
            }
        else:
            raise HTTPException(status_code=400, detail=f"Error al subir documento: {result['error']}")
            
    except Exception as e:
        logger.error(f"Error al subir documento S3 para trabajador {worker_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")


@router.post("/{worker_id}/medical-exams/upload-s3", response_model=Dict[str, Any])
async def upload_medical_exam_s3(
    worker_id: int,
    file: UploadFile = File(...),
    exam_type: str = Form("ocupacional"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Subir examen médico a S3.
    
    Args:
        worker_id: ID del trabajador
        file: Archivo del examen médico
        exam_type: Tipo de examen (ocupacional, ingreso, egreso, etc.)
        current_user: Usuario actual
        db: Sesión de base de datos
    
    Returns:
        Información del archivo subido
    """
    from app.services.s3_storage import upload_medical_exam
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Verificar permisos
    if not has_permission(current_user, "workers", "update"):
        raise HTTPException(status_code=403, detail="No tienes permisos para subir exámenes médicos")
    
    try:
        # Subir archivo a S3
        result = await upload_medical_exam(worker_id, file, exam_type)
        
        if result["success"]:
            logger.info(f"Examen médico subido exitosamente para trabajador {worker_id}: {result['file_key']}")
            return {
                "success": True,
                "message": "Examen médico subido exitosamente",
                "data": result
            }
        else:
            raise HTTPException(status_code=400, detail=f"Error al subir examen médico: {result['error']}")
            
    except Exception as e:
        logger.error(f"Error al subir examen médico S3 para trabajador {worker_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")


@router.get("/{worker_id}/files/s3", response_model=Dict[str, Any])
async def list_worker_files_s3(
    worker_id: int,
    folder_type: Optional[str] = Query(None, description="Tipo de carpeta: documentos o examenes_medicos"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Listar archivos de un trabajador en S3.
    
    Args:
        worker_id: ID del trabajador
        folder_type: Tipo de carpeta (documentos/examenes_medicos)
        current_user: Usuario actual
        db: Sesión de base de datos
    
    Returns:
        Lista de archivos del trabajador
    """
    from app.services.s3_storage import list_worker_files
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Verificar permisos
    if not has_permission(current_user, "workers", "read"):
        raise HTTPException(status_code=403, detail="No tienes permisos para ver archivos")
    
    try:
        # Listar archivos en S3
        result = list_worker_files(worker_id, folder_type)
        
        if result["success"]:
            return {
                "success": True,
                "message": "Archivos listados exitosamente",
                "data": result
            }
        else:
            raise HTTPException(status_code=400, detail=f"Error al listar archivos: {result['error']}")
            
    except Exception as e:
        logger.error(f"Error al listar archivos S3 para trabajador {worker_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")


@router.get("/files/s3/{file_key:path}/url", response_model=Dict[str, Any])
async def get_file_url_s3(
    file_key: str,
    expiration: int = Query(3600, description="Tiempo de expiración en segundos"),
    current_user: User = Depends(get_current_user)
):
    """
    Obtener URL firmada para acceder a un archivo en S3.
    
    Args:
        file_key: Clave del archivo en S3
        expiration: Tiempo de expiración en segundos
        current_user: Usuario actual
    
    Returns:
        URL firmada para acceder al archivo
    """
    from app.services.s3_storage import get_file_url
    
    # Verificar permisos
    if not has_permission(current_user, "workers", "read"):
        raise HTTPException(status_code=403, detail="No tienes permisos para acceder a archivos")
    
    try:
        # Obtener URL firmada
        url = get_file_url(file_key, expiration)
        
        if url:
            return {
                "success": True,
                "message": "URL generada exitosamente",
                "data": {
                    "file_key": file_key,
                    "url": url,
                    "expiration": expiration
                }
            }
        else:
            raise HTTPException(status_code=400, detail="Error al generar URL del archivo")
            
    except Exception as e:
        logger.error(f"Error al generar URL S3 para archivo {file_key}: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")


@router.delete("/files/s3/{file_key:path}", response_model=Dict[str, Any])
async def delete_file_s3(
    file_key: str,
    current_user: User = Depends(get_current_user)
):
    """
    Eliminar archivo de S3.
    
    Args:
        file_key: Clave del archivo en S3
        current_user: Usuario actual
    
    Returns:
        Resultado de la operación
    """
    from app.services.s3_storage import delete_file
    
    # Verificar permisos
    if not has_permission(current_user, "workers", "delete"):
        raise HTTPException(status_code=403, detail="No tienes permisos para eliminar archivos")
    
    try:
        # Eliminar archivo de S3
        result = delete_file(file_key)
        
        if result["success"]:
            logger.info(f"Archivo eliminado exitosamente de S3: {file_key}")
            return {
                "success": True,
                "message": "Archivo eliminado exitosamente",
                "data": result
            }
        else:
            raise HTTPException(status_code=400, detail=f"Error al eliminar archivo: {result['error']}")
            
    except Exception as e:
        logger.error(f"Error al eliminar archivo S3 {file_key}: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")





@router.put("/{worker_id}/vacations/{vacation_id}", response_model=WorkerVacation)
async def update_vacation_request(
    worker_id: int,
    vacation_id: int,
    vacation_data: WorkerVacationUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Actualizar una solicitud de vacaciones"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus, VacationBalance
    from app.schemas.worker_vacation import WorkerVacation as WorkerVacationSchema
    from datetime import datetime
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id, Worker.is_active == True).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Buscar la solicitud de vacaciones
    vacation = db.query(WorkerVacation).filter(
        WorkerVacation.id == vacation_id,
        WorkerVacation.worker_id == worker_id
    ).first()
    
    if not vacation:
        raise HTTPException(status_code=404, detail="Solicitud de vacaciones no encontrada")
    
    # Verificar permisos
    if not (current_user.role in [UserRole.ADMIN, UserRole.SUPERVISOR] or 
            (hasattr(worker, 'user_id') and worker.user_id == current_user.id)):
        raise HTTPException(status_code=403, detail="No tiene permisos para actualizar esta solicitud")
    
    # Solo se pueden actualizar solicitudes pendientes
    if vacation.status != VacationStatus.PENDING:
        raise HTTPException(
            status_code=400, 
            detail=f"Solo se pueden actualizar solicitudes pendientes. Estado actual: {vacation.status.value}"
        )
    
    # Actualizar campos si se proporcionan
    if vacation_data.start_date is not None:
        vacation.start_date = vacation_data.start_date
    
    if vacation_data.end_date is not None:
        vacation.end_date = vacation_data.end_date
    
    if vacation_data.comments is not None:
        vacation.comments = vacation_data.comments
    
    # Recalcular días si se cambiaron las fechas
    if vacation_data.start_date is not None or vacation_data.end_date is not None:
        from datetime import timedelta

        # Calcular días laborables (excluyendo fines de semana)
        current_date = vacation.start_date
        days = 0
        while current_date <= vacation.end_date:
            if current_date.weekday() < 5:  # Lunes a Viernes
                days += 1
            current_date += timedelta(days=1)

        # Validación por aniversario: las nuevas fechas deben estar dentro del periodo anual
        if not worker.fecha_de_ingreso:
            raise HTTPException(
                status_code=400,
                detail="El trabajador no tiene registrada la fecha de ingreso. Actualice la ficha del trabajador antes de actualizar la solicitud."
            )

        from datetime import date as _date
        def _add_years_safe(d: _date, years: int) -> _date:
            try:
                return d.replace(year=d.year + years)
            except ValueError:
                return d.replace(year=d.year + years, month=2, day=28)

        hire_date = worker.fecha_de_ingreso
        first_eligible = _add_years_safe(hire_date, 1)
        if vacation.start_date < first_eligible:
            raise HTTPException(
                status_code=400,
                detail=f"Aún no cumple un (1) año desde el ingreso ({hire_date}). Puede solicitar vacaciones a partir de {first_eligible}."
            )

        years_since = vacation.start_date.year - hire_date.year
        if (vacation.start_date.month, vacation.start_date.day) < (hire_date.month, hire_date.day):
            years_since -= 1
        period_start = _add_years_safe(hire_date, max(1, years_since))
        period_end = _add_years_safe(period_start, 1) - timedelta(days=1)

        if vacation.start_date < period_start or vacation.end_date > period_end:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Las fechas deben estar dentro del periodo anual de vacaciones por aniversario: "
                    f"desde {period_start} hasta {period_end}."
                )
            )

        # Validar conflictos de área (pendientes o aprobadas), excluyendo esta solicitud
        from sqlalchemy import or_
        conflicts = db.query(WorkerVacation).join(Worker, WorkerVacation.worker_id == Worker.id).filter(
            WorkerVacation.id != vacation_id,
            Worker.area_id == worker.area_id,
            or_(
                WorkerVacation.status == VacationStatus.PENDING,
                WorkerVacation.status == VacationStatus.APPROVED
            ),
            WorkerVacation.start_date <= vacation.end_date,
            WorkerVacation.end_date >= vacation.start_date
        ).all()
        if conflicts:
            raise HTTPException(
                status_code=400,
                detail="Las nuevas fechas se solapan con otra solicitud en tu área"
            )
        
        # Actualizar balance de vacaciones si cambió el número de días
        old_days = vacation.days_requested
        vacation.days_requested = days
        
        if old_days != days:
            vacation_balance = db.query(VacationBalance).filter(
                VacationBalance.worker_id == worker_id,
                VacationBalance.year == vacation.start_date.year
            ).first()
            
            if vacation_balance:
                # Ajustar días pendientes
                vacation_balance.pending_days = vacation_balance.pending_days - old_days + days
                
                # Verificar que no exceda los días disponibles
                if vacation_balance.available_days < 0:
                    raise HTTPException(
                        status_code=400,
                        detail=f"No tiene suficientes días disponibles. Disponibles: {vacation_balance.total_days - vacation_balance.used_days}, Solicitados: {days}"
                    )
    
    vacation.updated_at = datetime.utcnow()
    
    try:
        db.commit()
        db.refresh(vacation)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="Error al actualizar la solicitud de vacaciones")
    
    # Preparar respuesta
    vacation_dict = vacation.__dict__.copy()
    vacation_dict['worker_name'] = f"{worker.first_name} {worker.last_name}"
    
    if vacation.approved_by:
        approved_user = db.query(User).filter(User.id == vacation.approved_by).first()
        if approved_user:
            vacation_dict['approved_by_name'] = getattr(approved_user, 'full_name',
                                                      f"{approved_user.first_name} {approved_user.last_name}")
    
    return WorkerVacationSchema(**vacation_dict)


@router.delete("/{worker_id}/vacations/{vacation_id}")
async def delete_vacation_request(
    worker_id: int,
    vacation_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
) -> Any:
    """Eliminar una solicitud de vacaciones (solo administradores)"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus, VacationBalance
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id, Worker.is_active == True).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Buscar la solicitud de vacaciones
    vacation = db.query(WorkerVacation).filter(
        WorkerVacation.id == vacation_id,
        WorkerVacation.worker_id == worker_id
    ).first()
    
    if not vacation:
        raise HTTPException(status_code=404, detail="Solicitud de vacaciones no encontrada")
    
    # Actualizar balance de vacaciones antes de eliminar
    vacation_balance = db.query(VacationBalance).filter(
        VacationBalance.worker_id == worker_id,
        VacationBalance.year == vacation.start_date.year
    ).first()
    
    if vacation_balance:
        if vacation.status == VacationStatus.PENDING:
            # Devolver días pendientes
            vacation_balance.pending_days -= vacation.days_requested
        elif vacation.status == VacationStatus.APPROVED:
            # Devolver días usados
            vacation_balance.used_days -= vacation.days_requested
    
    try:
        # Eliminar la solicitud de vacaciones
        db.delete(vacation)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="Error al eliminar la solicitud de vacaciones")
    
    return {"message": "Solicitud de vacaciones eliminada exitosamente"}


@router.post("/vacations/fix-inactive-status")
async def fix_inactive_vacation_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Función temporal para corregir el estado is_active de solicitudes rechazadas/canceladas"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus
    
    # Verificar que el usuario sea administrador
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden ejecutar esta función")
    
    # Contar solicitudes rechazadas (ya no necesitamos actualizar is_active)
    rejected_count = db.query(WorkerVacation).filter(
        WorkerVacation.status == VacationStatus.REJECTED
    ).count()
    
    # Contar solicitudes canceladas (ya no necesitamos actualizar is_active)
    cancelled_count = db.query(WorkerVacation).filter(
        WorkerVacation.status == VacationStatus.CANCELLED
    ).count()
    
    db.commit()
    
    return {
        "message": "Estado is_active corregido exitosamente",
        "rejected_updated": rejected_count,
        "cancelled_updated": cancelled_count,
        "total_updated": rejected_count + cancelled_count
    }


# Endpoints para documentos de trabajadores
@router.post("/{worker_id}/documents", response_model=WorkerDocumentResponse)
async def upload_worker_document(
    worker_id: int,
    title: str = Form(...),
    category: str = Form(...),
    description: str = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Subir un documento para un trabajador específico"""
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Mapear categorías del frontend a los valores del enum
    category_mapping = {
        "Identificación": DocumentCategory.IDENTIFICATION,
        "identificacion": DocumentCategory.IDENTIFICATION,
        "Contrato": DocumentCategory.CONTRACT,
        "contrato": DocumentCategory.CONTRACT,
        "Médico": DocumentCategory.MEDICAL,
        "medico": DocumentCategory.MEDICAL,
        "Exámenes Médicos": DocumentCategory.MEDICAL,
        "examenes_medicos": DocumentCategory.MEDICAL,
        "Capacitación": DocumentCategory.TRAINING,
        "capacitacion": DocumentCategory.TRAINING,
        "Capacitaciones": DocumentCategory.TRAINING,
        "capacitaciones": DocumentCategory.TRAINING,
        "Certificación": DocumentCategory.CERTIFICATION,
        "certificacion": DocumentCategory.CERTIFICATION,
        "Certificados": DocumentCategory.CERTIFICATION,
        "certificados": DocumentCategory.CERTIFICATION,
        "Personal": DocumentCategory.PERSONAL,
        "personal": DocumentCategory.PERSONAL,
        "Otrosí": DocumentCategory.OTROSI,
        "otrosi": DocumentCategory.OTROSI,
        "Otro": DocumentCategory.OTHER,
        "otro": DocumentCategory.OTHER,
        "Otros": DocumentCategory.OTHER,
        "otros": DocumentCategory.OTHER
    }
    
    # Validar y convertir la categoría
    if category not in category_mapping:
        valid_categories = list(category_mapping.keys())
        raise HTTPException(
            status_code=400, 
            detail=f"Categoría no válida. Las categorías válidas son: {', '.join(valid_categories)}"
        )
    
    document_category = category_mapping[category]
    
    # Validar tipo de archivo
    allowed_types = [
        "application/pdf",
        "image/jpeg", "image/jpg", "image/png", "image/gif",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ]
    
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400, 
            detail="Tipo de archivo no permitido. Solo se permiten PDF, imágenes, Word y Excel."
        )
    
    # Generar nombre único para el archivo
    file_extension = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    
    try:
        from app.services.s3_storage import s3_service
        temp_result = await s3_service.upload_employee_document(worker_id, file, document_category.value if document_category else "general")
        if not temp_result.get("success"):
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error al subir archivo: {temp_result.get('error')}")
        file_url = temp_result["file_url"]
        
        # Crear registro en la base de datos
        db_document = WorkerDocument(
            worker_id=worker_id,
            title=title,
            description=description,
            category=document_category,
            file_name=file.filename,
            file_url=file_url,
            file_size=temp_result.get("size"),
            file_type=file.content_type,
            uploaded_by=current_user.id
        )
        
        db.add(db_document)
        db.commit()
        db.refresh(db_document)
        
        # Obtener información del usuario que subió el archivo
        uploader = db.query(User).filter(User.id == current_user.id).first()
        
        # Crear respuesta
        response_data = WorkerDocumentResponse(
            id=db_document.id,
            title=db_document.title,
            description=db_document.description,
            category=db_document.category,
            file_name=db_document.file_name,
            file_url=db_document.file_url,
            file_size=db_document.file_size,
            file_type=db_document.file_type,
            uploaded_by=db_document.uploaded_by,
            uploader_name=f"{uploader.first_name} {uploader.last_name}" if uploader else None,
            is_active=db_document.is_active,
            created_at=db_document.created_at,
            updated_at=db_document.updated_at
        )
        
        return response_data
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al subir el documento: {str(e)}")


@router.get("/{worker_id}/documents", response_model=List[WorkerDocumentResponse])
def get_worker_documents(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtener todos los documentos de un trabajador"""
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Obtener documentos activos
    documents = db.query(WorkerDocument).filter(
        WorkerDocument.worker_id == worker_id,
        WorkerDocument.is_active == True
    ).order_by(WorkerDocument.created_at.desc()).all()
    
    # Crear respuesta con información del usuario que subió cada documento
    response_documents = []
    for doc in documents:
        uploader = db.query(User).filter(User.id == doc.uploaded_by).first()
        response_data = WorkerDocumentResponse(
            id=doc.id,
            title=doc.title,
            description=doc.description,
            category=doc.category,
            file_name=doc.file_name,
            file_url=doc.file_url,
            file_size=doc.file_size,
            file_type=doc.file_type,
            uploaded_by=doc.uploaded_by,
            uploader_name=f"{uploader.first_name} {uploader.last_name}" if uploader else None,
            is_active=doc.is_active,
            created_at=doc.created_at,
            updated_at=doc.updated_at
        )
        response_documents.append(response_data)
    
    return response_documents


@router.get("/{worker_id}/documents/{document_id}", response_model=WorkerDocumentResponse)
def get_worker_document(
    worker_id: int,
    document_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtener un documento específico de un trabajador"""
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Obtener el documento
    document = db.query(WorkerDocument).filter(
        WorkerDocument.id == document_id,
        WorkerDocument.worker_id == worker_id,
        WorkerDocument.is_active == True
    ).first()
    
    if not document:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    # Obtener información del usuario que subió el documento
    uploader = db.query(User).filter(User.id == document.uploaded_by).first()
    
    response_data = WorkerDocumentResponse(
        id=document.id,
        title=document.title,
        description=document.description,
        category=document.category,
        file_name=document.file_name,
        file_url=document.file_url,
        file_size=document.file_size,
        file_type=document.file_type,
        uploaded_by=document.uploaded_by,
        uploader_name=f"{uploader.first_name} {uploader.last_name}" if uploader else None,
        is_active=document.is_active,
        created_at=document.created_at,
        updated_at=document.updated_at
    )
    
    return response_data


@router.put("/{worker_id}/documents/{document_id}", response_model=WorkerDocumentResponse)
def update_worker_document(
    worker_id: int,
    document_id: int,
    document_update: WorkerDocumentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Actualizar información de un documento"""
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Obtener el documento
    document = db.query(WorkerDocument).filter(
        WorkerDocument.id == document_id,
        WorkerDocument.worker_id == worker_id
    ).first()
    
    if not document:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    # Actualizar campos
    update_data = document_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(document, field, value)
    
    document.updated_at = datetime.utcnow()
    
    try:
        db.commit()
        db.refresh(document)
        
        # Obtener información del usuario que subió el documento
        uploader = db.query(User).filter(User.id == document.uploaded_by).first()
        
        response_data = WorkerDocumentResponse(
            id=document.id,
            title=document.title,
            description=document.description,
            category=document.category,
            file_name=document.file_name,
            file_url=document.file_url,
            file_size=document.file_size,
            file_type=document.file_type,
            uploaded_by=document.uploaded_by,
            uploader_name=f"{uploader.first_name} {uploader.last_name}" if uploader else None,
            is_active=document.is_active,
            created_at=document.created_at,
            updated_at=document.updated_at
        )
        
        return response_data
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al actualizar el documento: {str(e)}")


@router.delete("/{worker_id}/documents/{document_id}")
async def delete_worker_document(
    worker_id: int,
    document_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Eliminar un documento (soft delete)"""
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Obtener el documento
    document = db.query(WorkerDocument).filter(
        WorkerDocument.id == document_id,
        WorkerDocument.worker_id == worker_id
    ).first()
    
    if not document:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    if document.file_url and "s3.amazonaws.com" in document.file_url:
        try:
            from app.services.s3_storage import s3_service
            file_key = document.file_url.split('.com/')[-1]
            s3_service.delete_file(file_key)
        except Exception:
            pass
    
    # Soft delete
    document.is_active = False
    document.updated_at = datetime.utcnow()
    
    try:
        db.commit()
        return {"message": "Documento eliminado exitosamente"}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al eliminar el documento: {str(e)}")


@router.get("/{worker_id}/documents/{document_id}/download")
async def download_worker_document(
    worker_id: int,
    document_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Descargar un documento"""
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Obtener el documento
    document = db.query(WorkerDocument).filter(
        WorkerDocument.id == document_id,
        WorkerDocument.worker_id == worker_id,
        WorkerDocument.is_active == True
    ).first()
    
    if not document:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    try:
        if document.file_url and "s3.amazonaws.com" in document.file_url:
            from app.services.s3_storage import s3_service
            file_key = document.file_url.split('.com/')[-1]
            signed_url = s3_service.get_file_url(file_key, expiration=3600)
            if not signed_url:
                raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al generar URL de descarga desde S3")
            import httpx
            async with httpx.AsyncClient() as client:
                resp = await client.get(signed_url)
                resp.raise_for_status()
                file_bytes = resp.content
        else:
            raise HTTPException(status_code=404, detail="URL de archivo no válida para S3")
        return Response(
            content=file_bytes,
            media_type=document.file_type or "application/octet-stream",
            headers={
                "Content-Disposition": f"attachment; filename={document.file_name}"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al descargar el documento: {str(e)}")


@router.get("/{worker_id}/reinduction-history")
async def get_worker_reinduction_history(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
) -> Any:
    """
    Obtener historial de reinducciones de un trabajador específico
    """
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Trabajador no encontrado"
        )
    
    # Verificar permisos: el trabajador solo puede ver su propio historial
    # Los supervisores y admins pueden ver cualquier historial
    user_worker = db.query(Worker).filter(Worker.user_id == current_user.id).first()
    if (current_user.role.value not in ["admin", "supervisor"] and 
        (not user_worker or user_worker.id != worker_id)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tiene permisos para ver este historial"
        )
    
    # Obtener registros de reinducción ordenados por fecha descendente
    reinduction_records = db.query(ReinductionRecord).filter(
        ReinductionRecord.worker_id == worker_id
    ).order_by(ReinductionRecord.created_at.desc()).all()
    
    # Formatear respuesta
    history = []
    for record in reinduction_records:
        history.append({
            "id": record.id,
            "year": record.year,
            "status": record.status.value,
            "due_date": record.due_date.isoformat() if record.due_date else None,
            "completed_date": record.completed_date.isoformat() if record.completed_date else None,
            "created_at": record.created_at.isoformat() if record.created_at else None,
            "updated_at": record.updated_at.isoformat() if record.updated_at else None
        })
    
    return {
        "worker_id": worker_id,
        "worker_name": worker.full_name,
        "total_records": len(history),
        "reinduction_history": history
    }


# ==================== ENDPOINTS DE NOVEDADES ====================

@router.post("/{worker_id}/novedades", response_model=WorkerNovedadResponse)
async def create_worker_novedad(
    worker_id: int,
    novedad_data: WorkerNovedadCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Crear una nueva novedad para un trabajador"""
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Validar que worker_id coincida
    if novedad_data.worker_id != worker_id:
        raise HTTPException(status_code=400, detail="El worker_id no coincide")
    
    # Crear la novedad
    novedad = WorkerNovedad(
        worker_id=worker_id,
        tipo=novedad_data.tipo,
        titulo=novedad_data.titulo,
        descripcion=novedad_data.descripcion,
        fecha_inicio=novedad_data.fecha_inicio,
        fecha_fin=novedad_data.fecha_fin,
        salario_anterior=novedad_data.salario_anterior,
        monto_aumento=novedad_data.monto_aumento,
        cantidad_horas=novedad_data.cantidad_horas,
        valor_hora=novedad_data.valor_hora,
        observaciones=novedad_data.observaciones,
        documento_soporte=novedad_data.documento_soporte,
        registrado_por=current_user.id
    )
    
    # Calcular campos automáticos
    novedad.calcular_dias()
    novedad.calcular_nuevo_salario()
    novedad.calcular_valor_total_horas()
    
    db.add(novedad)
    db.commit()
    db.refresh(novedad)
    
    # Preparar respuesta con información adicional
    response_data = WorkerNovedadResponse.from_orm(novedad)
    response_data.worker_name = worker.full_name
    response_data.worker_document = worker.document_number
    response_data.registrado_por_name = current_user.full_name if hasattr(current_user, 'full_name') else f"{current_user.first_name} {current_user.last_name}"
    
    return response_data


@router.get("/{worker_id}/novedades", response_model=List[WorkerNovedadList])
async def get_worker_novedades(
    worker_id: int,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    tipo: NovedadType = Query(None, description="Filtrar por tipo de novedad"),
    status: NovedadStatus = Query(None, description="Filtrar por estado"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Obtener todas las novedades de un trabajador"""
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Construir query
    query = db.query(WorkerNovedad).filter(
        WorkerNovedad.worker_id == worker_id,
        WorkerNovedad.is_active == True
    )
    
    # Aplicar filtros
    if tipo:
        query = query.filter(WorkerNovedad.tipo == tipo)
    if status:
        query = query.filter(WorkerNovedad.status == status)
    
    # Ordenar por fecha de creación descendente
    query = query.order_by(WorkerNovedad.created_at.desc())
    
    # Aplicar paginación
    novedades = query.offset(skip).limit(limit).all()
    
    # Preparar respuesta con información adicional
    result = []
    for novedad in novedades:
        registrado_por_user = db.query(User).filter(User.id == novedad.registrado_por).first()
        registrado_por_name = "Usuario desconocido"
        if registrado_por_user:
            registrado_por_name = getattr(registrado_por_user, 'full_name', f"{registrado_por_user.first_name} {registrado_por_user.last_name}")
        
        # Obtener información del aprobador si existe
        aprobado_por_name = None
        if novedad.aprobado_por:
            aprobado_por_user = db.query(User).filter(User.id == novedad.aprobado_por).first()
            if aprobado_por_user:
                aprobado_por_name = getattr(aprobado_por_user, 'full_name', f"{aprobado_por_user.first_name} {aprobado_por_user.last_name}")
        
        novedad_data = WorkerNovedadList(
            id=novedad.id,
            worker_id=novedad.worker_id,
            worker_name=worker.full_name,
            worker_document=worker.document_number,
            tipo=novedad.tipo,
            titulo=novedad.titulo,
            descripcion=novedad.descripcion,
            status=novedad.status,
            fecha_inicio=novedad.fecha_inicio,
            fecha_fin=novedad.fecha_fin,
            dias_calculados=novedad.dias_calculados,
            monto_aumento=novedad.monto_aumento,
            valor_total=novedad.valor_total,
            observaciones=novedad.observaciones,
            registrado_por_name=registrado_por_name,
            aprobado_por_name=aprobado_por_name,
            fecha_aprobacion=novedad.fecha_aprobacion,
            created_at=novedad.created_at
        )
        result.append(novedad_data)
    
    return result


@router.get("/novedades", response_model=List[WorkerNovedadList])
async def get_all_novedades(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    tipo: NovedadType = Query(None, description="Filtrar por tipo de novedad"),
    status: NovedadStatus = Query(None, description="Filtrar por estado"),
    search: str = Query(None, description="Buscar por nombre o documento del trabajador"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Obtener todas las novedades del sistema"""
    
    # Construir query con join a workers
    query = db.query(WorkerNovedad).join(Worker).filter(
        WorkerNovedad.is_active == True
    )
    
    # Aplicar filtros
    if tipo:
        query = query.filter(WorkerNovedad.tipo == tipo)
    if status:
        query = query.filter(WorkerNovedad.status == status)
    if search:
        query = query.filter(
            or_(
                Worker.first_name.ilike(f"%{search}%"),
                Worker.last_name.ilike(f"%{search}%"),
                Worker.document_number.ilike(f"%{search}%")
            )
        )
    
    # Ordenar por fecha de creación descendente
    query = query.order_by(WorkerNovedad.created_at.desc())
    
    # Aplicar paginación
    novedades = query.offset(skip).limit(limit).all()
    
    # Preparar respuesta
    result = []
    for novedad in novedades:
        worker = novedad.worker
        registrado_por_user = db.query(User).filter(User.id == novedad.registrado_por).first()
        registrado_por_name = "Usuario desconocido"
        if registrado_por_user:
            registrado_por_name = getattr(registrado_por_user, 'full_name', f"{registrado_por_user.first_name} {registrado_por_user.last_name}")
        
        # Obtener información del aprobador si existe
        aprobado_por_name = None
        if novedad.aprobado_por:
            aprobado_por_user = db.query(User).filter(User.id == novedad.aprobado_por).first()
            if aprobado_por_user:
                aprobado_por_name = getattr(aprobado_por_user, 'full_name', f"{aprobado_por_user.first_name} {aprobado_por_user.last_name}")
        
        novedad_data = WorkerNovedadList(
            id=novedad.id,
            worker_id=novedad.worker_id,
            worker_name=worker.full_name,
            worker_document=worker.document_number,
            tipo=novedad.tipo,
            titulo=novedad.titulo,
            descripcion=novedad.descripcion,
            status=novedad.status,
            fecha_inicio=novedad.fecha_inicio,
            fecha_fin=novedad.fecha_fin,
            dias_calculados=novedad.dias_calculados,
            monto_aumento=novedad.monto_aumento,
            valor_total=novedad.valor_total,
            observaciones=novedad.observaciones,
            registrado_por_name=registrado_por_name,
            aprobado_por_name=aprobado_por_name,
            fecha_aprobacion=novedad.fecha_aprobacion,
            created_at=novedad.created_at
        )
        result.append(novedad_data)
    
    return result


@router.get("/novedades/{novedad_id}", response_model=WorkerNovedadResponse)
async def get_novedad(
    novedad_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Obtener una novedad específica"""
    
    novedad = db.query(WorkerNovedad).filter(
        WorkerNovedad.id == novedad_id,
        WorkerNovedad.is_active == True
    ).first()
    
    if not novedad:
        raise HTTPException(status_code=404, detail="Novedad no encontrada")
    
    # Preparar respuesta con información adicional
    worker = novedad.worker
    registrado_por_user = db.query(User).filter(User.id == novedad.registrado_por).first()
    aprobado_por_user = None
    if novedad.aprobado_por:
        aprobado_por_user = db.query(User).filter(User.id == novedad.aprobado_por).first()
    
    response_data = WorkerNovedadResponse.from_orm(novedad)
    response_data.worker_name = worker.full_name
    response_data.worker_document = worker.document_number
    
    if registrado_por_user:
        response_data.registrado_por_name = getattr(registrado_por_user, 'full_name', f"{registrado_por_user.first_name} {registrado_por_user.last_name}")
    
    if aprobado_por_user:
        response_data.aprobado_por_name = getattr(aprobado_por_user, 'full_name', f"{aprobado_por_user.first_name} {aprobado_por_user.last_name}")
    
    return response_data


@router.put("/novedades/{novedad_id}", response_model=WorkerNovedadResponse)
async def update_novedad(
    novedad_id: int,
    novedad_data: WorkerNovedadUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Actualizar una novedad"""
    
    novedad = db.query(WorkerNovedad).filter(
        WorkerNovedad.id == novedad_id,
        WorkerNovedad.is_active == True
    ).first()
    
    if not novedad:
        raise HTTPException(status_code=404, detail="Novedad no encontrada")
    
    # Verificar que la novedad esté en estado pendiente para poder editarla
    if novedad.status != NovedadStatus.PENDIENTE:
        raise HTTPException(status_code=400, detail="Solo se pueden editar novedades en estado pendiente")
    
    # Actualizar campos
    update_data = novedad_data.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(novedad, field, value)
    
    # Recalcular campos automáticos
    novedad.calcular_dias()
    novedad.calcular_nuevo_salario()
    novedad.calcular_valor_total_horas()
    
    db.commit()
    db.refresh(novedad)
    
    # Preparar respuesta
    worker = novedad.worker
    response_data = WorkerNovedadResponse.from_orm(novedad)
    response_data.worker_name = worker.full_name
    response_data.worker_document = worker.document_number
    
    return response_data


@router.post("/novedades/{novedad_id}/approve", response_model=WorkerNovedadResponse)
async def approve_reject_novedad(
    novedad_id: int,
    approval_data: WorkerNovedadApproval,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Aprobar o rechazar una novedad"""
    
    novedad = db.query(WorkerNovedad).filter(
        WorkerNovedad.id == novedad_id,
        WorkerNovedad.is_active == True
    ).first()
    
    if not novedad:
        raise HTTPException(status_code=404, detail="Novedad no encontrada")
    
    # Verificar que la novedad esté pendiente
    if novedad.status != NovedadStatus.PENDIENTE:
        raise HTTPException(status_code=400, detail="Solo se pueden aprobar/rechazar novedades pendientes")
    
    # Actualizar estado
    novedad.status = approval_data.status
    novedad.aprobado_por = current_user.id
    novedad.fecha_aprobacion = datetime.utcnow()
    
    if approval_data.observaciones:
        novedad.observaciones = approval_data.observaciones
    
    # Si es un aumento de salario aprobado, actualizar el salario del trabajador
    if (approval_data.status == NovedadStatus.APROBADA and 
        novedad.tipo == NovedadType.AUMENTO_SALARIO and 
        novedad.salario_nuevo):
        
        worker = novedad.worker
        worker.salary_ibc = float(novedad.salario_nuevo)
        novedad.status = NovedadStatus.PROCESADA  # Marcar como procesada
    
    db.commit()
    db.refresh(novedad)
    
    # Preparar respuesta
    worker = novedad.worker
    response_data = WorkerNovedadResponse.from_orm(novedad)
    response_data.worker_name = worker.full_name
    response_data.worker_document = worker.document_number
    response_data.aprobado_por_name = getattr(current_user, 'full_name', f"{current_user.first_name} {current_user.last_name}")
    
    return response_data


@router.delete("/novedades/{novedad_id}", response_model=MessageResponse)
async def delete_novedad(
    novedad_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
) -> Any:
    """Eliminar una novedad (soft delete)"""
    
    novedad = db.query(WorkerNovedad).filter(
        WorkerNovedad.id == novedad_id,
        WorkerNovedad.is_active == True
    ).first()
    
    if not novedad:
        raise HTTPException(status_code=404, detail="Novedad no encontrada")
    
    # Soft delete
    novedad.is_active = False
    
    db.commit()
    
    return {"message": "Novedad eliminada exitosamente"}


@router.get("/{worker_id}/novedades/stats", response_model=WorkerNovedadStats)
async def get_worker_novedades_stats(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Obtener estadísticas de novedades de un trabajador específico"""
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Verificar permisos: el trabajador solo puede ver sus propias estadísticas
    # Los supervisores y admins pueden ver cualquier estadística
    user_worker = db.query(Worker).filter(Worker.user_id == current_user.id).first()
    if (current_user.role.value not in ["admin", "supervisor"] and 
        (not user_worker or user_worker.id != worker_id)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tiene permisos para ver estas estadísticas"
        )
    
    # Contar por estado para este trabajador específico
    base_query = db.query(WorkerNovedad).filter(
        WorkerNovedad.worker_id == worker_id,
        WorkerNovedad.is_active == True
    )
    
    total_novedades = base_query.count()
    pendientes = base_query.filter(WorkerNovedad.status == NovedadStatus.PENDIENTE).count()
    aprobadas = base_query.filter(WorkerNovedad.status == NovedadStatus.APROBADA).count()
    rechazadas = base_query.filter(WorkerNovedad.status == NovedadStatus.RECHAZADA).count()
    procesadas = base_query.filter(WorkerNovedad.status == NovedadStatus.PROCESADA).count()
    
    # Contar por tipo para este trabajador específico
    por_tipo = {}
    for tipo in NovedadType:
        count = base_query.filter(WorkerNovedad.tipo == tipo).count()
        por_tipo[tipo.value] = count
    
    return WorkerNovedadStats(
        total_novedades=total_novedades,
        pendientes=pendientes,
        aprobadas=aprobadas,
        rechazadas=rechazadas,
        procesadas=procesadas,
        por_tipo=por_tipo
    )


@router.get("/{worker_id}/novedades/export")
async def export_worker_novedades_to_excel(
    worker_id: int,
    tipo: NovedadType = Query(None, description="Filtrar por tipo de novedad"),
    status: NovedadStatus = Query(None, description="Filtrar por estado"),
    start_date: Optional[str] = Query(None, description="Fecha de inicio (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Fecha de fin (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """
    Exportar las novedades de un trabajador a Excel con filtros de fecha
    """
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Construir query para obtener todas las novedades (sin paginación para exportar)
    query = db.query(WorkerNovedad).filter(
        WorkerNovedad.worker_id == worker_id,
        WorkerNovedad.is_active == True
    )
    
    # Aplicar filtros
    if tipo:
        query = query.filter(WorkerNovedad.tipo == tipo)
    if status:
        query = query.filter(WorkerNovedad.status == status)
    
    # Aplicar filtros de fecha - incluir novedades que se solapen con el período seleccionado
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            # Incluir novedades que terminen después o en la fecha de inicio
            query = query.filter(
                or_(
                    WorkerNovedad.fecha_fin >= start_dt,
                    WorkerNovedad.fecha_fin.is_(None)  # Incluir novedades sin fecha fin
                )
            )
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha de inicio inválido. Use YYYY-MM-DD")
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
            # Incluir novedades que inicien antes o en la fecha de fin
            query = query.filter(
                or_(
                    WorkerNovedad.fecha_inicio <= end_dt,
                    WorkerNovedad.fecha_inicio.is_(None)  # Incluir novedades sin fecha inicio
                )
            )
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha de fin inválido. Use YYYY-MM-DD")
    
    # Ordenar por fecha de creación descendente
    query = query.order_by(WorkerNovedad.created_at.desc())
    
    # Obtener todas las novedades
    novedades = query.all()
    
    # Crear el archivo Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Novedades"
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Definir encabezados
    headers = [
        "ID", "Tipo", "Título", "Descripción", "Estado", 
        "Fecha Inicio", "Fecha Fin", "Días Calculados",
        "Salario Anterior", "Salario Nuevo", "Monto Aumento",
        "Cantidad Horas", "Valor Hora", "Valor Total",
        "Observaciones", "Registrado Por", "Aprobado Por",
        "Fecha Aprobación", "Fecha Registro", "Última Actualización"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Mapeo de tipos de novedad para mostrar nombres legibles
    NOVEDAD_TYPES_MAP = {
        "permiso_dia_familia": "Permiso Día de la Familia",
        "licencia_paternidad": "Licencia de Paternidad", 
        "incapacidad_medica": "Incapacidad Médica",
        "permiso_dia_no_remunerado": "Permiso Día No Remunerado",
        "aumento_salario": "Aumento de Salario",
        "licencia_maternidad": "Licencia de Maternidad",
        "horas_extras": "Horas Extras",
        "recargos": "Recargos",
        "capacitacion": "Capacitación",
        "trabajo_en_casa": "Trabajo en casa",
        "cobertura_en_el_exterior": "Cobertura en el exterior",
    }
    
    # Mapeo de estados
    STATUS_MAP = {
        "pendiente": "Pendiente",
        "aprobada": "Aprobada",
        "rechazada": "Rechazada", 
        "procesada": "Procesada"
    }
    
    # Escribir datos
    for row, novedad in enumerate(novedades, 2):
        # Obtener información del usuario que registró
        registrado_por_user = db.query(User).filter(User.id == novedad.registrado_por).first()
        registrado_por_name = "Usuario desconocido"
        if registrado_por_user:
            registrado_por_name = getattr(registrado_por_user, 'full_name', f"{registrado_por_user.first_name} {registrado_por_user.last_name}")
        
        # Obtener información del aprobador si existe
        aprobado_por_name = ""
        if novedad.aprobado_por:
            aprobado_por_user = db.query(User).filter(User.id == novedad.aprobado_por).first()
            if aprobado_por_user:
                aprobado_por_name = getattr(aprobado_por_user, 'full_name', f"{aprobado_por_user.first_name} {aprobado_por_user.last_name}")
        
        # Escribir datos de la fila
        data = [
            novedad.id,
            NOVEDAD_TYPES_MAP.get(novedad.tipo.value if hasattr(novedad.tipo, 'value') else novedad.tipo, novedad.tipo),
            novedad.titulo,
            novedad.descripcion or "",
            STATUS_MAP.get(novedad.status.value if hasattr(novedad.status, 'value') else novedad.status, novedad.status),
            novedad.fecha_inicio.strftime("%Y-%m-%d") if novedad.fecha_inicio else "",
            novedad.fecha_fin.strftime("%Y-%m-%d") if novedad.fecha_fin else "",
            novedad.dias_calculados or "",
            novedad.salario_anterior or "",
            novedad.salario_nuevo or "",
            novedad.monto_aumento or "",
            novedad.cantidad_horas or "",
            novedad.valor_hora or "",
            novedad.valor_total or "",
            novedad.observaciones or "",
            registrado_por_name,
            aprobado_por_name,
            novedad.fecha_aprobacion.strftime("%Y-%m-%d %H:%M") if novedad.fecha_aprobacion else "",
            novedad.created_at.strftime("%Y-%m-%d %H:%M") if novedad.created_at else "",
            novedad.updated_at.strftime("%Y-%m-%d %H:%M") if novedad.updated_at else ""
        ]
        
        for col, value in enumerate(data, 1):
            worksheet.cell(row=row, column=col, value=value)
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Crear el archivo en memoria
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generar nombre del archivo
    worker_name = f"{worker.first_name}_{worker.last_name}".replace(" ", "_")
    filename = f"novedades_{worker_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Retornar el archivo como respuesta de streaming
    return StreamingResponse(
        BytesIO(excel_buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/vacations/occupied-dates")
async def get_occupied_dates(
    start_date: str = Query(..., description="Fecha de inicio en formato YYYY-MM-DD"),
    end_date: str = Query(..., description="Fecha de fin en formato YYYY-MM-DD"),
    worker_id: int | None = Query(None, description="Opcional: filtra por el área del trabajador indicado"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Obtener fechas ocupadas por vacaciones aprobadas y pendientes activas en un rango de fechas.

    Ahora filtra por área del trabajador:
    - Si se envía `worker_id`, se usan las vacaciones de trabajadores en la misma `area_id`.
    - Si no se envía, se infiere el trabajador a partir de `current_user` y se usa su `area_id`.
    - Si no se puede determinar el `area_id`, se retorna lista vacía para evitar falsos positivos.
    """
    from app.models.worker_vacation import WorkerVacation, VacationStatus
    from datetime import datetime, timedelta
    from sqlalchemy.orm import joinedload
    
    # Validar formato de fechas
    try:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")

    if start_dt > end_dt:
        raise HTTPException(status_code=400, detail="La fecha de inicio debe ser anterior a la fecha de fin")

    # Determinar área objetivo (usar area_id si existe, de lo contrario department)
    target_area_id = None
    target_department = None
    target_worker = None

    if worker_id is not None:
        target_worker = db.query(Worker).filter(Worker.id == worker_id).first()
    else:
        # Inferir worker desde el usuario autenticado
        target_worker = db.query(Worker).filter(Worker.user_id == current_user.id).first()

    if target_worker:
        if target_worker.area_id is not None:
            target_area_id = target_worker.area_id
        if getattr(target_worker, "department", None):
            target_department = target_worker.department

    # Si no hay área ni departamento definida, no mostrar ocupados (respeta el requerimiento por área)
    if target_area_id is None and not target_department:
        return {
            "occupied_dates": [],
            "total_occupied_days": 0,
            "query_range": {
                "start_date": start_date,
                "end_date": end_date
            }
        }

    # Buscar vacaciones aprobadas y pendientes en la misma área/departamento que se solapen con el rango
    conflict_filters = [
        WorkerVacation.status.in_([VacationStatus.APPROVED, VacationStatus.PENDING]),
        WorkerVacation.start_date <= end_dt,
        WorkerVacation.end_date >= start_dt,
        Worker.is_active == True,
    ]

    if target_area_id is not None:
        conflict_filters.append(Worker.area_id == target_area_id)
    elif target_department:
        conflict_filters.append(Worker.department == target_department)

    occupied_vacations = (
        db.query(WorkerVacation)
        .join(Worker, WorkerVacation.worker_id == Worker.id)
        .options(joinedload(WorkerVacation.worker))
        .filter(*conflict_filters)
        .all()
    )

    # Generar lista de fechas ocupadas
    occupied_dates = []
    for vacation in occupied_vacations:
        worker = vacation.worker or db.query(Worker).filter(Worker.id == vacation.worker_id).first()
        worker_name = f"{worker.first_name} {worker.last_name}" if worker else "Trabajador desconocido"

        # Generar todas las fechas del período de vacaciones
        current_date = vacation.start_date
        while current_date <= vacation.end_date:
            occupied_dates.append({
                "date": current_date.strftime('%Y-%m-%d'),
                "worker_name": worker_name,
                "vacation_id": vacation.id,
                "start_date": vacation.start_date.strftime('%Y-%m-%d'),
                "end_date": vacation.end_date.strftime('%Y-%m-%d')
            })
            current_date += timedelta(days=1)

    return {
        "occupied_dates": occupied_dates,
        "total_occupied_days": len(occupied_dates),
        "query_range": {
            "start_date": start_date,
            "end_date": end_date
        }
    }


@router.get("/novedades/stats/summary", response_model=WorkerNovedadStats)
async def get_novedades_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Obtener estadísticas de novedades"""
    
    # Contar por estado
    total_novedades = db.query(WorkerNovedad).filter(WorkerNovedad.is_active == True).count()
    pendientes = db.query(WorkerNovedad).filter(
        WorkerNovedad.is_active == True,
        WorkerNovedad.status == NovedadStatus.PENDIENTE
    ).count()
    aprobadas = db.query(WorkerNovedad).filter(
        WorkerNovedad.is_active == True,
        WorkerNovedad.status == NovedadStatus.APROBADA
    ).count()
    rechazadas = db.query(WorkerNovedad).filter(
        WorkerNovedad.is_active == True,
        WorkerNovedad.status == NovedadStatus.RECHAZADA
    ).count()
    procesadas = db.query(WorkerNovedad).filter(
        WorkerNovedad.is_active == True,
        WorkerNovedad.status == NovedadStatus.PROCESADA
    ).count()
    
    # Contar por tipo
    por_tipo = {}
    for tipo in NovedadType:
        count = db.query(WorkerNovedad).filter(
            WorkerNovedad.is_active == True,
            WorkerNovedad.tipo == tipo
        ).count()
        por_tipo[tipo.value] = count
    
    return WorkerNovedadStats(
        total_novedades=total_novedades,
        pendientes=pendientes,
        aprobadas=aprobadas,
        rechazadas=rechazadas,
        procesadas=procesadas,
        por_tipo=por_tipo
    )




# ==================== ENDPOINTS DE VACACIONES ====================

@router.get("/vacations/all", response_model=List[VacationRequestWithWorker])
async def get_all_vacation_requests(
    status: Optional[VacationStatus] = None,
    year: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Obtener todas las solicitudes de vacaciones (solo para administradores y supervisores)"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus
    from sqlalchemy import and_
    
    # Construir query base - mostrar todas las solicitudes para trazabilidad
    query = db.query(WorkerVacation)
    
    # Filtrar por estado si se especifica
    if status:
        query = query.filter(WorkerVacation.status == status)
    
    # Filtrar por año si se especifica
    if year:
        query = query.filter(
            and_(
                WorkerVacation.start_date >= date(year, 1, 1),
                WorkerVacation.start_date <= date(year, 12, 31)
            )
        )
    
    # Obtener las solicitudes con información del trabajador
    vacations = query.order_by(WorkerVacation.created_at.desc()).offset(skip).limit(limit).all()
    
    # Enriquecer con información del trabajador
    result = []
    for vacation in vacations:
        worker = db.query(Worker).filter(Worker.id == vacation.worker_id).first()
        vacation_data = VacationRequestWithWorker(
            id=vacation.id,
            worker_id=vacation.worker_id,
            worker_name=f"{worker.first_name} {worker.last_name}" if worker else "Trabajador no encontrado",
            start_date=vacation.start_date,
            end_date=vacation.end_date,
            days_requested=vacation.days_requested,
            comments=vacation.comments,
            status=vacation.status,
            created_at=vacation.created_at,
            updated_at=vacation.updated_at,
            approved_by=vacation.approved_by,
            approved_date=vacation.approved_date
        )
        result.append(vacation_data)
    
    return result

@router.get("/vacations/availability", response_model=VacationAvailability)
async def check_vacation_availability(
    start_date: date,
    end_date: date,
    worker_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Verificar disponibilidad de fechas para vacaciones"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus
    from app.schemas.worker_vacation import VacationAvailability, VacationConflict
    from app.models.vacation_balance import VacationBalance
    
    # Calcular días solicitados (solo días laborales)
    requested_days = 0
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() < 5:  # Lunes a Viernes
            requested_days += 1
        current_date = current_date + timedelta(days=1)
    
    # Buscar conflictos con vacaciones aprobadas en el rango
    # Si se especifica worker_id, limitar a trabajadores de la misma área
    if worker_id is not None:
        worker = db.query(Worker).filter(Worker.id == worker_id).first()
    else:
        worker = None

    from sqlalchemy.orm import joinedload
    conflict_query = db.query(WorkerVacation)
    if worker is not None:
        conflict_query = conflict_query.join(Worker, WorkerVacation.worker_id == Worker.id)
        # Filtrar por misma área o mismo departamento (fallback)
        if worker.area_id is not None:
            conflict_query = conflict_query.filter(Worker.area_id == worker.area_id)
        elif getattr(worker, "department", None):
            conflict_query = conflict_query.filter(Worker.department == worker.department)
        else:
            # Si el trabajador no tiene área ni departamento, no reportar conflictos
            return VacationAvailability(
                start_date=start_date,
                end_date=end_date,
                is_available=True,
                conflicts=[],
                requested_days=requested_days,
                available_days=None
            )
        conflict_query = conflict_query.filter(WorkerVacation.worker_id != worker_id)

    conflicts = conflict_query.filter(
        WorkerVacation.status == VacationStatus.APPROVED,
        WorkerVacation.start_date <= end_date,
        WorkerVacation.end_date >= start_date
    ).all()
    
    conflict_list = []
    for conflict in conflicts:
        worker = db.query(Worker).filter(Worker.id == conflict.worker_id).first()
        if worker:
            # Calcular días de solapamiento
            overlap_start = max(start_date, conflict.start_date)
            overlap_end = min(end_date, conflict.end_date)
            
            overlapping_days = 0
            current_date = overlap_start
            while current_date <= overlap_end:
                if current_date.weekday() < 5:  # Solo días laborales
                    overlapping_days += 1
                current_date = current_date + timedelta(days=1)
            
            conflict_list.append(VacationConflict(
                worker_name=f"{worker.first_name} {worker.last_name}",
                start_date=conflict.start_date,
                end_date=conflict.end_date,
                overlapping_days=overlapping_days
            ))
    
    # Verificar días disponibles si se especifica worker_id
    available_days = None
    if worker_id:
        vacation_balance = db.query(VacationBalance).filter(
            VacationBalance.worker_id == worker_id,
            VacationBalance.year == start_date.year
        ).first()
        
        if vacation_balance:
            available_days = vacation_balance.available_days
        else:
            available_days = 15  # Días por defecto
    
    # Determinar disponibilidad
    is_available = len(conflict_list) == 0 and (not worker_id or available_days >= requested_days)
    
    return VacationAvailability(
        start_date=start_date,
        end_date=end_date,
        is_available=is_available,
        conflicts=conflict_list,
        requested_days=requested_days,
        available_days=available_days
    )


@router.get("/vacations/stats", response_model=VacationStats)
async def get_vacation_stats(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Obtener estadísticas de vacaciones"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus
    from app.schemas.worker_vacation import VacationStats
    
    if year is None:
        year = datetime.now().year
    
    # Obtener todas las solicitudes del año
    all_requests = db.query(WorkerVacation).filter(
        WorkerVacation.start_date >= datetime(year, 1, 1).date(),
        WorkerVacation.start_date <= datetime(year, 12, 31).date()
    ).all()
    
    total_requests = len(all_requests)
    pending_requests = len([r for r in all_requests if r.status == VacationStatus.PENDING])
    approved_requests = len([r for r in all_requests if r.status == VacationStatus.APPROVED])
    rejected_requests = len([r for r in all_requests if r.status == VacationStatus.REJECTED])
    
    total_days_requested = sum(r.days_requested for r in all_requests)
    total_days_approved = sum(r.days_requested for r in all_requests if r.status == VacationStatus.APPROVED)
    
    # Trabajadores con solicitudes pendientes
    workers_with_pending = len(set(r.worker_id for r in all_requests if r.status == VacationStatus.PENDING))
    
    return VacationStats(
        total_requests=total_requests,
        pending_requests=pending_requests,
        approved_requests=approved_requests,
        rejected_requests=rejected_requests,
        total_days_requested=total_days_requested,
        total_days_approved=total_days_approved,
        workers_with_pending=workers_with_pending
    )


@router.get("/{worker_id}/vacations/availability", response_model=VacationAvailability)
async def check_worker_vacation_availability(
    worker_id: int,
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Verificar disponibilidad de fechas para vacaciones de un trabajador específico"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus, VacationBalance
    from app.schemas.worker_vacation import VacationAvailability, VacationConflict
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Calcular días solicitados (solo días laborales)
    requested_days = 0
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() < 5:  # Lunes a Viernes
            requested_days += 1
        current_date = current_date + timedelta(days=1)
    
    # Buscar conflictos con vacaciones aprobadas de trabajadores de la misma área
    # (excluyendo al trabajador actual)
    # Conflictos por misma área o mismo departamento (fallback)
    conflict_filters = [
        WorkerVacation.worker_id != worker_id,
        WorkerVacation.status == VacationStatus.APPROVED,
        WorkerVacation.start_date <= end_date,
        WorkerVacation.end_date >= start_date,
    ]
    if worker.area_id is not None:
        conflict_filters.append(Worker.area_id == worker.area_id)
    elif getattr(worker, "department", None):
        conflict_filters.append(Worker.department == worker.department)
    else:
        # Sin área ni departamento, asumimos sin conflictos
        conflicts = []
        conflict_list = []
        requested_days = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:
                requested_days += 1
            current_date = current_date + timedelta(days=1)
        # Balance
        vacation_balance = db.query(VacationBalance).filter(
            VacationBalance.worker_id == worker_id,
            VacationBalance.year == start_date.year
        ).first()
        available_days = vacation_balance.available_days if vacation_balance else 15
        is_available = available_days >= requested_days
        return VacationAvailability(
            start_date=start_date,
            end_date=end_date,
            is_available=is_available,
            conflicts=conflict_list,
            requested_days=requested_days,
            available_days=available_days
        )

    conflicts = db.query(WorkerVacation).join(Worker, WorkerVacation.worker_id == Worker.id).filter(
        *conflict_filters
    ).all()
    
    conflict_list = []
    for conflict in conflicts:
        conflict_worker = db.query(Worker).filter(Worker.id == conflict.worker_id).first()
        if conflict_worker:
            # Calcular días de solapamiento
            overlap_start = max(start_date, conflict.start_date)
            overlap_end = min(end_date, conflict.end_date)
            overlapping_days = (overlap_end - overlap_start).days + 1
            
            conflict_list.append(VacationConflict(
                worker_id=conflict.worker_id,
                worker_name=f"{conflict_worker.first_name} {conflict_worker.last_name}",
                start_date=conflict.start_date,
                end_date=conflict.end_date,
                status=conflict.status,
                overlapping_days=overlapping_days
            ))
    
    # Verificar días disponibles del trabajador
    vacation_balance = db.query(VacationBalance).filter(
        VacationBalance.worker_id == worker_id,
        VacationBalance.year == start_date.year
    ).first()
    
    if not vacation_balance:
        vacation_balance = VacationBalance(
            worker_id=worker_id,
            year=start_date.year,
            total_days=15,
            used_days=0,
            pending_days=0
        )
        from sqlalchemy.exc import IntegrityError
        try:
            db.add(vacation_balance)
            db.commit()
            db.refresh(vacation_balance)
        except IntegrityError:
            db.rollback()
            vacation_balance = db.query(VacationBalance).filter(
                VacationBalance.worker_id == worker_id,
                VacationBalance.year == start_date.year
            ).first()
            if not vacation_balance:
                raise

    
    available_days = vacation_balance.available_days
    
    # Determinar disponibilidad
    is_available = len(conflict_list) == 0 and available_days >= requested_days
    
    return VacationAvailability(
        start_date=start_date,
        end_date=end_date,
        is_available=is_available,
        conflicts=conflict_list,
        requested_days=requested_days,
        available_days=available_days
    )


@router.get("/{worker_id}/vacations", response_model=List[WorkerVacation])
async def get_worker_vacations(
    worker_id: int,
    status: Optional[VacationStatus] = None,
    year: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Obtener vacaciones de un trabajador"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus
    from app.schemas.worker_vacation import WorkerVacation as WorkerVacationSchema
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id, Worker.is_active == True).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Verificar permisos: admin, supervisor o el mismo trabajador
    if not (current_user.role in [UserRole.ADMIN, UserRole.SUPERVISOR] or 
            (hasattr(worker, 'user_id') and worker.user_id == current_user.id)):
        raise HTTPException(status_code=403, detail="No tiene permisos para ver estas vacaciones")
    
    # Construir query
    query = db.query(WorkerVacation).filter(
        WorkerVacation.worker_id == worker_id
    )
    
    if status:
        query = query.filter(WorkerVacation.status == status)
    
    if year:
        from sqlalchemy import extract
        query = query.filter(extract('year', WorkerVacation.start_date) == year)
    
    vacations = query.order_by(WorkerVacation.start_date.desc()).offset(skip).limit(limit).all()
    
    # Enriquecer con información adicional
    result = []
    for vacation in vacations:
        vacation_dict = vacation.__dict__.copy()
        vacation_dict['worker_name'] = f"{worker.first_name} {worker.last_name}"
        
        # Información del aprobador
        if vacation.approved_by:
            approved_user = db.query(User).filter(User.id == vacation.approved_by).first()
            if approved_user:
                vacation_dict['approved_by_name'] = getattr(approved_user, 'full_name', 
                                                          f"{approved_user.first_name} {approved_user.last_name}")
        
        result.append(WorkerVacationSchema(**vacation_dict))
    
    return result


@router.post("/{worker_id}/vacations", response_model=WorkerVacation)
async def create_vacation_request(
    worker_id: int,
    vacation_data: WorkerVacationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Crear solicitud de vacaciones"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus, VacationBalance
    from app.schemas.worker_vacation import WorkerVacation as WorkerVacationSchema
    from datetime import timedelta
    from sqlalchemy import or_
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id, Worker.is_active == True).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Verificar permisos: admin, supervisor o el mismo trabajador
    if not (current_user.role in [UserRole.ADMIN, UserRole.SUPERVISOR] or 
            (hasattr(worker, 'user_id') and worker.user_id == current_user.id)):
        raise HTTPException(status_code=403, detail="No tiene permisos para crear solicitudes de vacaciones")
    
    # Calcular días laborales (excluyendo fines de semana)
    start_date = vacation_data.start_date
    end_date = vacation_data.end_date
    days = 0
    current_date = start_date
    
    while current_date <= end_date:
        if current_date.weekday() < 5:  # Lunes a Viernes (0-4)
            days += 1
        current_date += timedelta(days=1)
    
    if days <= 0:
        raise HTTPException(status_code=400, detail="La solicitud debe incluir al menos un día laboral")

    # Validación por aniversario de ingreso: las vacaciones solo se pueden tomar
    # a partir del primer aniversario y dentro del periodo anual entre aniversarios.
    if not worker.fecha_de_ingreso:
        raise HTTPException(
            status_code=400,
            detail="El trabajador no tiene registrada la fecha de ingreso. Actualice la ficha del trabajador antes de solicitar vacaciones."
        )

    from datetime import date, timedelta as _timedelta

    def _add_years_safe(d: date, years: int) -> date:
        try:
            return d.replace(year=d.year + years)
        except ValueError:
            # Ajuste para fechas como 29 de febrero en años no bisiestos
            return d.replace(year=d.year + years, month=2, day=28)

    hire_date = worker.fecha_de_ingreso
    first_eligible = _add_years_safe(hire_date, 1)

    if start_date < first_eligible:
        raise HTTPException(
            status_code=400,
            detail=f"Aún no cumple un (1) año desde el ingreso ({hire_date}). Puede solicitar vacaciones a partir de {first_eligible}."
        )

    # Calcular el periodo de aniversario correspondiente al rango solicitado
    years_since = start_date.year - hire_date.year
    if (start_date.month, start_date.day) < (hire_date.month, hire_date.day):
        years_since -= 1
    # El inicio del periodo es el aniversario alcanzado más reciente
    period_start = _add_years_safe(hire_date, max(1, years_since))
    period_end = _add_years_safe(period_start, 1) - _timedelta(days=1)

    if start_date < period_start or end_date > period_end:
        raise HTTPException(
            status_code=400,
            detail=(
                "Las fechas deben estar dentro del periodo anual de vacaciones por aniversario: "
                f"desde {period_start} hasta {period_end}."
            )
        )
    
    # Verificar balance de vacaciones
    current_year = start_date.year
    vacation_balance = db.query(VacationBalance).filter(
        VacationBalance.worker_id == worker_id,
        VacationBalance.year == current_year
    ).first()
    
    if not vacation_balance:
        # Crear balance inicial si no existe
        vacation_balance = VacationBalance(
            worker_id=worker_id,
            year=current_year,
            total_days=15,
            used_days=0,
            pending_days=0
        )
        db.add(vacation_balance)
        db.flush()
    
    # Verificar días disponibles
    if vacation_balance.available_days < days:
        raise HTTPException(
            status_code=400, 
            detail=f"No tiene suficientes días disponibles. Disponibles: {vacation_balance.available_days}, Solicitados: {days}"
        )
    
    # Verificar conflictos con solicitudes del mismo trabajador (solo aprobadas)
    self_conflicts = db.query(WorkerVacation).filter(
        WorkerVacation.worker_id == worker_id,
        WorkerVacation.status == VacationStatus.APPROVED,
        WorkerVacation.start_date <= end_date,
        WorkerVacation.end_date >= start_date
    ).all()

    # Validar traslape real en días hábiles (evita falsos positivos por límites contiguos)
    def _business_overlap_days(a_start, a_end, b_start, b_end) -> int:
        from datetime import timedelta as _td
        start = max(a_start, b_start)
        end = min(a_end, b_end)
        # Contigüidad en el mismo día NO cuenta como solape
        if start >= end:
            return 0
        days = 0
        d = start
        while d <= end:
            if d.weekday() < 5:
                days += 1
            d += _td(days=1)
        return days

    for sc in self_conflicts:
        overlap_days = _business_overlap_days(sc.start_date, sc.end_date, start_date, end_date)
        if overlap_days > 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Las fechas seleccionadas se solapan con otra solicitud propia aprobada "
                    f"({sc.start_date}–{sc.end_date}, traslape hábil: {overlap_days} días)"
                )
            )

    # Verificar conflictos con otras solicitudes del mismo área (pendientes o aprobadas)
    # Conflictos por misma área o mismo departamento (fallback)
    area_or_dept_filters = [
        WorkerVacation.worker_id != worker_id,
        or_(
            WorkerVacation.status == VacationStatus.PENDING,
            WorkerVacation.status == VacationStatus.APPROVED
        ),
        WorkerVacation.start_date <= end_date,
        WorkerVacation.end_date >= start_date
    ]
    if worker.area_id is not None:
        area_or_dept_filters.append(Worker.area_id == worker.area_id)
    elif getattr(worker, "department", None):
        area_or_dept_filters.append(Worker.department == worker.department)
    else:
        area_or_dept_filters.append(True)  # sin área/departamento, no filtrar por área

    conflicts = db.query(WorkerVacation).join(Worker, WorkerVacation.worker_id == Worker.id).filter(
        *area_or_dept_filters
    ).all()

    if conflicts:
        first = conflicts[0]
        raise HTTPException(
            status_code=400,
            detail=(
                "Las fechas seleccionadas ya están ocupadas por un trabajador de tu área "
                f"({first.start_date}–{first.end_date})"
            )
        )
    
    # Crear la solicitud
    vacation = WorkerVacation(
        worker_id=worker_id,
        start_date=start_date,
        end_date=end_date,
        days_requested=days,
        comments=vacation_data.comments,
        status=VacationStatus.PENDING
    )
    
    db.add(vacation)
    
    # Actualizar días pendientes en el balance
    vacation_balance.pending_days += days
    
    db.commit()
    db.refresh(vacation)
    
    # Preparar respuesta con información adicional
    vacation_dict = vacation.__dict__.copy()
    vacation_dict['worker_name'] = f"{worker.first_name} {worker.last_name}"
    vacation_dict['requested_by_name'] = getattr(current_user, 'full_name', 
                                                f"{current_user.first_name} {current_user.last_name}")
    
    return WorkerVacationSchema(**vacation_dict)


@router.put("/vacations/{vacation_id}/approve", response_model=WorkerVacation)
async def approve_vacation(
    vacation_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Aprobar solicitud de vacaciones"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus, VacationBalance
    from app.schemas.worker_vacation import WorkerVacation as WorkerVacationSchema
    from datetime import datetime
    
    # Buscar la solicitud
    vacation = db.query(WorkerVacation).filter(
        WorkerVacation.id == vacation_id
    ).first()
    
    if not vacation:
        raise HTTPException(status_code=404, detail="Solicitud de vacaciones no encontrada")
    
    if vacation.status != VacationStatus.PENDING:
        raise HTTPException(status_code=400, detail="Solo se pueden aprobar solicitudes pendientes")
    
    # Obtener información del trabajador
    worker = db.query(Worker).filter(Worker.id == vacation.worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")

    # Validar conflictos de área con otras solicitudes aprobadas
    conflicts = db.query(WorkerVacation).join(Worker, WorkerVacation.worker_id == Worker.id).filter(
        WorkerVacation.worker_id != vacation.worker_id,
        Worker.area_id == worker.area_id,
        WorkerVacation.status == VacationStatus.APPROVED,
        WorkerVacation.start_date <= vacation.end_date,
        WorkerVacation.end_date >= vacation.start_date
    ).all()
    if conflicts:
        raise HTTPException(
            status_code=400,
            detail="No se puede aprobar: las fechas se solapan con una solicitud aprobada en el área"
        )
    
    # Actualizar la solicitud
    vacation.status = VacationStatus.APPROVED
    vacation.approved_by = current_user.id
    vacation.approved_date = datetime.utcnow()
    
    # Actualizar balance de vacaciones
    vacation_balance = db.query(VacationBalance).filter(
        VacationBalance.worker_id == vacation.worker_id,
        VacationBalance.year == vacation.start_date.year
    ).first()
    
    if vacation_balance:
        vacation_balance.pending_days -= vacation.days_requested
        vacation_balance.used_days += vacation.days_requested
    
    db.commit()
    db.refresh(vacation)
    
    # Preparar respuesta
    vacation_dict = vacation.__dict__.copy()
    vacation_dict['worker_name'] = f"{worker.first_name} {worker.last_name}"
    vacation_dict['approved_by_name'] = getattr(current_user, 'full_name', 
                                              f"{current_user.first_name} {current_user.last_name}")
    
    # Información del solicitante
    return WorkerVacationSchema(**vacation_dict)


@router.put("/vacations/{vacation_id}/reject", response_model=WorkerVacation)
async def reject_vacation(
    vacation_id: int,
    rejection_data: WorkerVacationUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> Any:
    """Rechazar solicitud de vacaciones"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus, VacationBalance
    from app.schemas.worker_vacation import WorkerVacation as WorkerVacationSchema
    from datetime import datetime
    
    # Buscar la solicitud
    vacation = db.query(WorkerVacation).filter(
        WorkerVacation.id == vacation_id
    ).first()
    
    if not vacation:
        raise HTTPException(status_code=404, detail="Solicitud de vacaciones no encontrada")
    
    if vacation.status != VacationStatus.PENDING:
        raise HTTPException(status_code=400, detail="Solo se pueden rechazar solicitudes pendientes")
    
    if not rejection_data.comments:
        raise HTTPException(status_code=400, detail="El comentario de rechazo es requerido")
    
    # Obtener información del trabajador
    worker = db.query(Worker).filter(Worker.id == vacation.worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Actualizar la solicitud
    vacation.status = VacationStatus.REJECTED
    vacation.approved_by = current_user.id
    vacation.approved_date = datetime.utcnow()
    vacation.comments = rejection_data.comments
    
    # Liberar días pendientes en el balance
    vacation_balance = db.query(VacationBalance).filter(
        VacationBalance.worker_id == vacation.worker_id,
        VacationBalance.year == vacation.start_date.year
    ).first()
    
    if vacation_balance:
        vacation_balance.pending_days -= vacation.days_requested
    
    db.commit()
    db.refresh(vacation)
    
    # Preparar respuesta
    vacation_dict = vacation.__dict__.copy()
    vacation_dict['worker_name'] = f"{worker.first_name} {worker.last_name}"
    vacation_dict['approved_by_name'] = getattr(current_user, 'full_name', 
                                              f"{current_user.first_name} {current_user.last_name}")
    
    return WorkerVacationSchema(**vacation_dict)


@router.put("/vacations/{vacation_id}/cancel", response_model=WorkerVacation)
async def cancel_vacation(
    vacation_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Cancelar una solicitud de vacaciones"""
    from app.models.worker_vacation import WorkerVacation, VacationStatus, VacationBalance
    from app.schemas.worker_vacation import WorkerVacation as WorkerVacationSchema
    from datetime import datetime
    
    # Buscar la solicitud
    vacation = db.query(WorkerVacation).filter(WorkerVacation.id == vacation_id).first()
    if not vacation:
        raise HTTPException(status_code=404, detail="Solicitud de vacaciones no encontrada")
    
    # Verificar permisos - solo el trabajador o admin/supervisor pueden cancelar
    worker = db.query(Worker).filter(Worker.id == vacation.worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Verificar permisos
    can_cancel = (
        current_user.role in [UserRole.ADMIN, UserRole.SUPERVISOR] or
        (hasattr(worker, 'user_id') and worker.user_id == current_user.id)
    )
    
    if not can_cancel:
        raise HTTPException(status_code=403, detail="No tiene permisos para cancelar esta solicitud")
    
    # Solo se pueden cancelar solicitudes pendientes o aprobadas
    if vacation.status not in [VacationStatus.PENDING, VacationStatus.APPROVED]:
        raise HTTPException(
            status_code=400, 
            detail=f"Solo se pueden cancelar solicitudes pendientes o aprobadas. Estado actual: {vacation.status.value}"
        )
    
    # Guardar el estado original antes de cambiarlo
    original_status = vacation.status
    
    # Actualizar la solicitud
    vacation.status = VacationStatus.CANCELLED
    vacation.approved_by = current_user.id
    vacation.approved_date = datetime.utcnow()
    
    # Liberar días en el balance
    vacation_balance = db.query(VacationBalance).filter(
        VacationBalance.worker_id == vacation.worker_id,
        VacationBalance.year == vacation.start_date.year
    ).first()
    
    if vacation_balance:
        if original_status == VacationStatus.PENDING:
            # Si estaba pendiente, liberar días pendientes
            vacation_balance.pending_days -= vacation.days_requested
        elif original_status == VacationStatus.APPROVED:
            # Si estaba aprobada, liberar días usados
            vacation_balance.used_days -= vacation.days_requested
    
    db.commit()
    db.refresh(vacation)
    
    # Preparar respuesta
    vacation_dict = vacation.__dict__.copy()
    vacation_dict['worker_name'] = f"{worker.first_name} {worker.last_name}"
    vacation_dict['approved_by_name'] = getattr(current_user, 'full_name', 
                                              f"{current_user.first_name} {current_user.last_name}")
    
    return WorkerVacationSchema(**vacation_dict)


@router.get("/{worker_id}/vacation-balance", response_model=VacationBalance)
async def get_vacation_balance(
    worker_id: int,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """Obtener balance de vacaciones de un trabajador"""
    from app.models.worker_vacation import VacationBalance
    from app.schemas.worker_vacation import VacationBalance as VacationBalanceSchema
    from datetime import datetime
    
    # Verificar que el trabajador existe
    worker = db.query(Worker).filter(Worker.id == worker_id, Worker.is_active == True).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado")
    
    # Verificar permisos
    if not (current_user.role in [UserRole.ADMIN, UserRole.SUPERVISOR] or 
            (hasattr(worker, 'user_id') and worker.user_id == current_user.id)):
        raise HTTPException(status_code=403, detail="No tiene permisos para ver este balance")
    
    if not year:
        year = datetime.now().year
    
    # Buscar o crear balance
    vacation_balance = db.query(VacationBalance).filter(
        VacationBalance.worker_id == worker_id,
        VacationBalance.year == year
    ).first()
    
    if not vacation_balance:
        # Crear balance inicial
        vacation_balance = VacationBalance(
            worker_id=worker_id,
            year=year,
            total_days=15,
            used_days=0,
            pending_days=0
        )
        db.add(vacation_balance)
        db.commit()
        db.refresh(vacation_balance)
    
    return VacationBalanceSchema.from_orm(vacation_balance)


@router.get("/vacations/export/excel")
async def export_vacations_to_excel(
    start_date: str = Query(None, description="Fecha de inicio (YYYY-MM-DD)"),
    end_date: str = Query(None, description="Fecha de fin (YYYY-MM-DD)"),
    status: str = Query(None, description="Estado de la solicitud"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_supervisor_or_admin)
) -> StreamingResponse:
    """
    Exportar solicitudes de vacaciones a Excel por período
    """
    from app.models.worker_vacation import WorkerVacation, VacationStatus
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Construir query base
    query = db.query(WorkerVacation, Worker).join(
        Worker, WorkerVacation.worker_id == Worker.id
    )
    
    # Filtros por fecha - lógica inclusiva para capturar vacaciones que se solapan con el período
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            # Incluir vacaciones que terminen después o en la fecha de inicio del período
            query = query.filter(WorkerVacation.end_date >= start_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha de inicio inválido. Use YYYY-MM-DD")
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            # Incluir vacaciones que empiecen antes o en la fecha de fin del período
            query = query.filter(WorkerVacation.start_date <= end_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha de fin inválido. Use YYYY-MM-DD")
    
    # Filtro por estado
    if status:
        if status.lower() == 'pending':
            query = query.filter(WorkerVacation.status == VacationStatus.PENDING)
        elif status.lower() == 'approved':
            query = query.filter(WorkerVacation.status == VacationStatus.APPROVED)
        elif status.lower() == 'rejected':
            query = query.filter(WorkerVacation.status == VacationStatus.REJECTED)
    
    # Ordenar por fecha de solicitud
    query = query.order_by(WorkerVacation.created_at.desc())
    
    results = query.all()
    
    # Crear el archivo Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Solicitudes de Vacaciones"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        "ID Solicitud", "Trabajador", "Documento", "Cargo", "Fecha Inicio", 
        "Fecha Fin", "Días Solicitados", "Motivo", "Estado", "Comentarios Admin",
        "Fecha Solicitud", "Aprobado Por", "Fecha Aprobación"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escribir datos
    for row, (vacation, worker) in enumerate(results, 2):
        # Obtener información del aprobador si existe
        approved_by_name = ""
        if vacation.approved_by:
            approver = db.query(User).filter(User.id == vacation.approved_by).first()
            if approver:
                approved_by_name = f"{approver.first_name} {approver.last_name}"
        
        # Estado en español
        status_text = {
            VacationStatus.PENDING: "Pendiente",
            VacationStatus.APPROVED: "Aprobado", 
            VacationStatus.REJECTED: "Rechazado"
        }.get(vacation.status, vacation.status.value if vacation.status else "")
        
        data = [
            vacation.id,
            f"{worker.first_name} {worker.last_name}",
            worker.document_number,
            worker.position or "",
            vacation.start_date.strftime('%d/%m/%Y') if vacation.start_date else "",
            vacation.end_date.strftime('%d/%m/%Y') if vacation.end_date else "",
            vacation.days_requested,
            vacation.comments or "",
            status_text,
            vacation.comments or "",
            vacation.created_at.strftime('%d/%m/%Y %H:%M') if vacation.created_at else "",
            approved_by_name,
            vacation.approved_date.strftime('%d/%m/%Y %H:%M') if vacation.approved_date else ""
        ]
        
        for col, value in enumerate(data, 1):
            worksheet.cell(row=row, column=col, value=value)
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generar nombre del archivo
    filename = f"solicitudes_vacaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
