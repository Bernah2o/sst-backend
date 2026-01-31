"""
API endpoints para la Matriz Legal SST.

Incluye endpoints para:
- Importación de archivos Excel
- Gestión de normas
- Cumplimiento por empresa
- Dashboard y estadísticas
- Exportación
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, or_, and_, distinct

from app.database import get_db
from app.dependencies import get_current_active_user, require_admin, require_supervisor_or_admin
from app.models.empresa import Empresa
from app.models.sector_economico import SectorEconomico
from app.models.matriz_legal import (
    MatrizLegalNorma, MatrizLegalNormaHistorial,
    MatrizLegalCumplimiento, MatrizLegalCumplimientoHistorial,
    MatrizLegalImportacion,
    EstadoNorma, EstadoCumplimiento, EstadoImportacion
)
from app.models.user import User
from app.schemas.matriz_legal import (
    MatrizLegalNorma as MatrizLegalNormaSchema,
    MatrizLegalNormaUpdate,
    MatrizLegalNormaConCumplimiento,
    MatrizLegalCumplimiento as MatrizLegalCumplimientoSchema,
    MatrizLegalCumplimientoUpdate,
    MatrizLegalCumplimientoBulkUpdate,
    MatrizLegalCumplimientoHistorial as CumplimientoHistorialSchema,
    MatrizLegalImportacionPreview,
    MatrizLegalImportacionResult,
    MatrizLegalEstadisticas,
    MatrizLegalDashboard,
    MatrizLegalEstadisticasPorEstado,
    PaginatedMatrizLegalNormas,
    PaginatedMatrizLegalNormasConCumplimiento,
    PaginatedImportaciones,
)
from app.schemas.common import MessageResponse
from app.services.matriz_legal_service import MatrizLegalService

logger = logging.getLogger(__name__)

router = APIRouter()


# ==================== IMPORTACIÓN ====================

@router.post("/importar/preview", response_model=MatrizLegalImportacionPreview)
async def preview_importacion(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Previsualiza una importación de Excel sin guardar cambios.
    Devuelve estadísticas y errores de validación.
    """
    # Validar tipo de archivo
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se permiten archivos Excel (.xlsx, .xls)"
        )

    # Validar tamaño (máximo 50MB)
    file_content = await file.read()
    max_size = 50 * 1024 * 1024
    if len(file_content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="El archivo es demasiado grande. Máximo 50MB"
        )

    service = MatrizLegalService(db)
    result = service.preview_import(file_content, file.filename)

    return MatrizLegalImportacionPreview(**result)


@router.post("/importar", response_model=MatrizLegalImportacionResult)
async def importar_excel_arl(
    file: UploadFile = File(...),
    sobrescribir_existentes: bool = Query(False, description="Si actualiza normas existentes con cambios"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Importa el archivo Excel mensual de la ARL.
    Detecta normas nuevas vs actualizaciones.
    """
    # Validar tipo de archivo
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se permiten archivos Excel (.xlsx, .xls)"
        )

    # Validar tamaño
    file_content = await file.read()
    max_size = 50 * 1024 * 1024
    if len(file_content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="El archivo es demasiado grande. Máximo 50MB"
        )

    service = MatrizLegalService(db)
    try:
        importacion = service.import_excel(
            file_content,
            file.filename,
            current_user.id,
            sobrescribir_existentes
        )
        return importacion
    except Exception as e:
        logger.error(f"Error en importación: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al importar archivo: {str(e)}"
        )


@router.get("/importaciones", response_model=PaginatedImportaciones)
def list_importaciones(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Historial de importaciones."""
    query = db.query(MatrizLegalImportacion).order_by(
        MatrizLegalImportacion.fecha_importacion.desc()
    )

    total = query.count()
    skip = (page - 1) * size
    items = query.offset(skip).limit(size).all()
    pages = (total + size - 1) // size

    return PaginatedImportaciones(
        items=items,
        total=total,
        page=page,
        size=size,
        pages=pages,
        has_next=page < pages,
        has_prev=page > 1,
    )


# ==================== NORMAS ====================

@router.get("/normas", response_model=PaginatedMatrizLegalNormas)
def list_normas(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    q: Optional[str] = Query(None, description="Búsqueda en descripción, tema, tipo norma"),
    sector_economico_id: Optional[int] = Query(None),
    clasificacion: Optional[str] = Query(None),
    tema_general: Optional[str] = Query(None),
    anio: Optional[int] = Query(None),
    estado: Optional[str] = Query(None),
    activo: bool = Query(True),
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """Lista todas las normas con filtros y paginación."""
    query = db.query(MatrizLegalNorma).options(
        joinedload(MatrizLegalNorma.sector_economico)
    )

    if activo is not None:
        query = query.filter(MatrizLegalNorma.activo == activo)

    if q:
        search = f"%{q}%"
        query = query.filter(
            or_(
                MatrizLegalNorma.descripcion_norma.ilike(search),
                MatrizLegalNorma.tipo_norma.ilike(search),
                MatrizLegalNorma.numero_norma.ilike(search),
                MatrizLegalNorma.tema_general.ilike(search),
                MatrizLegalNorma.clasificacion_norma.ilike(search),
            )
        )

    if sector_economico_id:
        query = query.filter(MatrizLegalNorma.sector_economico_id == sector_economico_id)

    if clasificacion:
        query = query.filter(MatrizLegalNorma.clasificacion_norma == clasificacion)

    if tema_general:
        query = query.filter(MatrizLegalNorma.tema_general == tema_general)

    if anio:
        query = query.filter(MatrizLegalNorma.anio == anio)

    if estado:
        query = query.filter(MatrizLegalNorma.estado == estado)

    total = query.count()
    skip = (page - 1) * size
    items = query.order_by(
        MatrizLegalNorma.anio.desc(),
        MatrizLegalNorma.tipo_norma,
        MatrizLegalNorma.numero_norma
    ).offset(skip).limit(size).all()

    # Agregar campos derivados (sector_economico_nombre no es una propiedad del modelo)
    # identificador_completo es una @property, Pydantic lo lee automáticamente con from_attributes=True
    for item in items:
        item.sector_economico_nombre = item.sector_economico.nombre if item.sector_economico else None

    pages = (total + size - 1) // size

    return PaginatedMatrizLegalNormas(
        items=items,
        total=total,
        page=page,
        size=size,
        pages=pages,
        has_next=page < pages,
        has_prev=page > 1,
    )


@router.get("/normas/{norma_id}", response_model=MatrizLegalNormaSchema)
def get_norma(
    norma_id: int,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """Obtiene una norma por ID."""
    norma = db.query(MatrizLegalNorma).options(
        joinedload(MatrizLegalNorma.sector_economico)
    ).filter(MatrizLegalNorma.id == norma_id).first()

    if not norma:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Norma no encontrada"
        )

    norma.sector_economico_nombre = norma.sector_economico.nombre if norma.sector_economico else None
    # identificador_completo es una @property, Pydantic lo lee automáticamente

    return norma


@router.put("/normas/{norma_id}", response_model=MatrizLegalNormaSchema)
def update_norma(
    norma_id: int,
    payload: MatrizLegalNormaUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Actualiza campos editables de una norma (ej: aplicabilidad)."""
    norma = db.query(MatrizLegalNorma).filter(
        MatrizLegalNorma.id == norma_id
    ).first()

    if not norma:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Norma no encontrada"
        )

    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(norma, key, value)

    norma.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(norma)

    return norma


@router.get("/normas/catalogos/clasificaciones", response_model=List[str])
def list_clasificaciones(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Lista valores únicos de clasificación_norma para filtros."""
    result = db.query(distinct(MatrizLegalNorma.clasificacion_norma)).filter(
        MatrizLegalNorma.activo == True,
        MatrizLegalNorma.clasificacion_norma.isnot(None)
    ).order_by(MatrizLegalNorma.clasificacion_norma).all()

    return [r[0] for r in result if r[0]]


@router.get("/normas/catalogos/temas", response_model=List[str])
def list_temas(
    clasificacion: Optional[str] = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Lista valores únicos de tema_general para filtros."""
    query = db.query(distinct(MatrizLegalNorma.tema_general)).filter(
        MatrizLegalNorma.activo == True,
        MatrizLegalNorma.tema_general.isnot(None)
    )

    if clasificacion:
        query = query.filter(MatrizLegalNorma.clasificacion_norma == clasificacion)

    result = query.order_by(MatrizLegalNorma.tema_general).all()

    return [r[0] for r in result if r[0]]


@router.get("/normas/catalogos/anios", response_model=List[int])
def list_anios(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Lista años disponibles para filtros."""
    result = db.query(distinct(MatrizLegalNorma.anio)).filter(
        MatrizLegalNorma.activo == True
    ).order_by(MatrizLegalNorma.anio.desc()).all()

    return [r[0] for r in result if r[0]]


# ==================== CUMPLIMIENTO POR EMPRESA ====================

@router.get("/empresas/{empresa_id}/normas", response_model=PaginatedMatrizLegalNormasConCumplimiento)
def list_normas_empresa(
    empresa_id: int,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    estado_cumplimiento: Optional[str] = Query(None),
    clasificacion: Optional[str] = Query(None),
    tema_general: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    solo_aplicables: bool = Query(True),
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """
    Lista normas aplicables a una empresa con su estado de cumplimiento.
    """
    # Verificar empresa existe
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada"
        )

    # Query base con left join a cumplimientos
    query = db.query(
        MatrizLegalNorma,
        MatrizLegalCumplimiento
    ).outerjoin(
        MatrizLegalCumplimiento,
        and_(
            MatrizLegalCumplimiento.norma_id == MatrizLegalNorma.id,
            MatrizLegalCumplimiento.empresa_id == empresa_id
        )
    ).filter(
        MatrizLegalNorma.activo == True,
        MatrizLegalNorma.estado == EstadoNorma.VIGENTE.value
    )

    # Si solo aplicables, usar el servicio para filtrar
    if solo_aplicables:
        service = MatrizLegalService(db)
        normas_aplicables = service.get_normas_aplicables_empresa(empresa)
        normas_ids = [n.id for n in normas_aplicables]
        query = query.filter(MatrizLegalNorma.id.in_(normas_ids))

    # Filtros adicionales
    if estado_cumplimiento:
        if estado_cumplimiento == 'pendiente':
            query = query.filter(
                or_(
                    MatrizLegalCumplimiento.estado == EstadoCumplimiento.PENDIENTE.value,
                    MatrizLegalCumplimiento.id.is_(None)
                )
            )
        else:
            query = query.filter(MatrizLegalCumplimiento.estado == estado_cumplimiento)

    if clasificacion:
        query = query.filter(MatrizLegalNorma.clasificacion_norma == clasificacion)

    if tema_general:
        query = query.filter(MatrizLegalNorma.tema_general == tema_general)

    if q:
        search = f"%{q}%"
        query = query.filter(
            or_(
                MatrizLegalNorma.descripcion_norma.ilike(search),
                MatrizLegalNorma.tipo_norma.ilike(search),
                MatrizLegalNorma.numero_norma.ilike(search),
            )
        )

    # Paginación
    total = query.count()
    skip = (page - 1) * size
    results = query.order_by(
        MatrizLegalNorma.anio.desc(),
        MatrizLegalNorma.tipo_norma
    ).offset(skip).limit(size).all()

    # Construir respuesta
    items = []
    for norma, cumplimiento in results:
        item = MatrizLegalNormaConCumplimiento(
            id=norma.id,
            ambito_aplicacion=norma.ambito_aplicacion,
            sector_economico_id=norma.sector_economico_id,
            sector_economico_texto=norma.sector_economico_texto,
            clasificacion_norma=norma.clasificacion_norma,
            tema_general=norma.tema_general,
            subtema_riesgo_especifico=norma.subtema_riesgo_especifico,
            anio=norma.anio,
            tipo_norma=norma.tipo_norma,
            numero_norma=norma.numero_norma,
            fecha_expedicion=norma.fecha_expedicion,
            expedida_por=norma.expedida_por,
            descripcion_norma=norma.descripcion_norma,
            articulo=norma.articulo,
            estado=norma.estado,
            info_adicional=norma.info_adicional,
            descripcion_articulo_exigencias=norma.descripcion_articulo_exigencias,
            aplica_trabajadores_independientes=norma.aplica_trabajadores_independientes,
            aplica_teletrabajo=norma.aplica_teletrabajo,
            aplica_trabajo_alturas=norma.aplica_trabajo_alturas,
            aplica_espacios_confinados=norma.aplica_espacios_confinados,
            aplica_trabajo_caliente=norma.aplica_trabajo_caliente,
            aplica_sustancias_quimicas=norma.aplica_sustancias_quimicas,
            aplica_radiaciones=norma.aplica_radiaciones,
            aplica_trabajo_nocturno=norma.aplica_trabajo_nocturno,
            aplica_menores_edad=norma.aplica_menores_edad,
            aplica_mujeres_embarazadas=norma.aplica_mujeres_embarazadas,
            aplica_conductores=norma.aplica_conductores,
            aplica_manipulacion_alimentos=norma.aplica_manipulacion_alimentos,
            aplica_maquinaria_pesada=norma.aplica_maquinaria_pesada,
            aplica_riesgo_electrico=norma.aplica_riesgo_electrico,
            aplica_riesgo_biologico=norma.aplica_riesgo_biologico,
            aplica_trabajo_excavaciones=norma.aplica_trabajo_excavaciones,
            aplica_general=norma.aplica_general,
            version=norma.version,
            activo=norma.activo,
            created_at=norma.created_at,
            updated_at=norma.updated_at,
            # Campos de cumplimiento
            cumplimiento_id=cumplimiento.id if cumplimiento else None,
            estado_cumplimiento=cumplimiento.estado if cumplimiento else EstadoCumplimiento.PENDIENTE.value,
            aplica_empresa=cumplimiento.aplica_empresa if cumplimiento else True,
            evidencia_cumplimiento=cumplimiento.evidencia_cumplimiento if cumplimiento else None,
            observaciones=cumplimiento.observaciones if cumplimiento else None,
            plan_accion=cumplimiento.plan_accion if cumplimiento else None,
            responsable=cumplimiento.responsable if cumplimiento else None,
            fecha_compromiso=cumplimiento.fecha_compromiso if cumplimiento else None,
            fecha_ultima_evaluacion=cumplimiento.fecha_ultima_evaluacion if cumplimiento else None,
        )
        items.append(item)

    pages = (total + size - 1) // size

    return PaginatedMatrizLegalNormasConCumplimiento(
        items=items,
        total=total,
        page=page,
        size=size,
        pages=pages,
        has_next=page < pages,
        has_prev=page > 1,
    )


@router.get("/empresas/{empresa_id}/dashboard", response_model=MatrizLegalDashboard)
def get_dashboard_empresa(
    empresa_id: int,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """Dashboard de cumplimiento para una empresa."""
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada"
        )

    service = MatrizLegalService(db)
    stats_data = service.get_estadisticas_empresa(empresa_id)

    estadisticas = MatrizLegalEstadisticas(
        empresa_id=stats_data['empresa_id'],
        empresa_nombre=stats_data['empresa_nombre'],
        total_normas_aplicables=stats_data['total_normas_aplicables'],
        por_estado=MatrizLegalEstadisticasPorEstado(**stats_data['por_estado']),
        porcentaje_cumplimiento=stats_data['porcentaje_cumplimiento'],
        normas_con_plan_accion=stats_data['normas_con_plan_accion'],
        normas_vencidas=stats_data['normas_vencidas'],
    )

    # Últimas evaluaciones
    ultimas_evaluaciones = db.query(
        MatrizLegalCumplimiento
    ).filter(
        MatrizLegalCumplimiento.empresa_id == empresa_id,
        MatrizLegalCumplimiento.fecha_ultima_evaluacion.isnot(None)
    ).order_by(
        MatrizLegalCumplimiento.fecha_ultima_evaluacion.desc()
    ).limit(5).all()

    # Normas críticas (no cumple sin plan de acción)
    normas_criticas_query = db.query(
        MatrizLegalNorma,
        MatrizLegalCumplimiento
    ).join(
        MatrizLegalCumplimiento,
        MatrizLegalCumplimiento.norma_id == MatrizLegalNorma.id
    ).filter(
        MatrizLegalCumplimiento.empresa_id == empresa_id,
        MatrizLegalCumplimiento.estado == EstadoCumplimiento.NO_CUMPLE.value,
        or_(
            MatrizLegalCumplimiento.plan_accion.is_(None),
            MatrizLegalCumplimiento.plan_accion == ''
        )
    ).limit(10).all()

    # Importaciones recientes
    importaciones = db.query(MatrizLegalImportacion).order_by(
        MatrizLegalImportacion.fecha_importacion.desc()
    ).limit(5).all()

    return MatrizLegalDashboard(
        estadisticas=estadisticas,
        ultimas_evaluaciones=[
            {
                'id': c.id,
                'norma_id': c.norma_id,
                'estado': c.estado,
                'fecha': c.fecha_ultima_evaluacion.isoformat() if c.fecha_ultima_evaluacion else None
            }
            for c in ultimas_evaluaciones
        ],
        normas_criticas=[],  # Simplificado por ahora
        proximas_revisiones=[],
        importaciones_recientes=importaciones,
    )


@router.get("/empresas/{empresa_id}/estadisticas", response_model=MatrizLegalEstadisticas)
def get_estadisticas_empresa(
    empresa_id: int,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """Estadísticas de cumplimiento para una empresa."""
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada"
        )

    service = MatrizLegalService(db)
    stats_data = service.get_estadisticas_empresa(empresa_id)

    return MatrizLegalEstadisticas(
        empresa_id=stats_data['empresa_id'],
        empresa_nombre=stats_data['empresa_nombre'],
        total_normas_aplicables=stats_data['total_normas_aplicables'],
        por_estado=MatrizLegalEstadisticasPorEstado(**stats_data['por_estado']),
        porcentaje_cumplimiento=stats_data['porcentaje_cumplimiento'],
        normas_con_plan_accion=stats_data['normas_con_plan_accion'],
        normas_vencidas=stats_data['normas_vencidas'],
    )


@router.put("/empresas/{empresa_id}/cumplimiento/{norma_id}", response_model=MatrizLegalCumplimientoSchema)
def update_cumplimiento(
    empresa_id: int,
    norma_id: int,
    payload: MatrizLegalCumplimientoUpdate,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """Actualiza el estado de cumplimiento de una norma para una empresa."""
    # Verificar empresa
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada"
        )

    # Verificar norma
    norma = db.query(MatrizLegalNorma).filter(MatrizLegalNorma.id == norma_id).first()
    if not norma:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Norma no encontrada"
        )

    # Buscar o crear cumplimiento
    cumplimiento = db.query(MatrizLegalCumplimiento).filter(
        MatrizLegalCumplimiento.empresa_id == empresa_id,
        MatrizLegalCumplimiento.norma_id == norma_id
    ).first()

    estado_anterior = cumplimiento.estado if cumplimiento and cumplimiento.estado else None

    if not cumplimiento:
        cumplimiento = MatrizLegalCumplimiento(
            empresa_id=empresa_id,
            norma_id=norma_id,
            estado=EstadoCumplimiento.PENDIENTE.value
        )
        db.add(cumplimiento)
        db.flush()

    # Actualizar campos
    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        # Convertir enums a strings para columnas que ahora son String
        if key == 'estado' and value is not None:
            value = value.value if hasattr(value, 'value') else value
        setattr(cumplimiento, key, value)

    cumplimiento.fecha_ultima_evaluacion = datetime.utcnow()
    cumplimiento.evaluado_por = current_user.id
    cumplimiento.updated_at = datetime.utcnow()

    # Registrar en historial si cambió el estado
    if payload.estado and estado_anterior != payload.estado.value:
        service = MatrizLegalService(db)
        service.registrar_cambio_cumplimiento(
            cumplimiento,
            estado_anterior,
            current_user.id,
            payload.observaciones
        )

    db.commit()
    db.refresh(cumplimiento)

    return cumplimiento


@router.post("/empresas/{empresa_id}/cumplimiento/bulk", response_model=MessageResponse)
def bulk_update_cumplimiento(
    empresa_id: int,
    payload: MatrizLegalCumplimientoBulkUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Actualiza múltiples cumplimientos a la vez."""
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada"
        )

    updated = 0
    for cumplimiento_id in payload.cumplimiento_ids:
        cumplimiento = db.query(MatrizLegalCumplimiento).filter(
            MatrizLegalCumplimiento.id == cumplimiento_id,
            MatrizLegalCumplimiento.empresa_id == empresa_id
        ).first()

        if cumplimiento:
            estado_anterior = cumplimiento.estado
            cumplimiento.estado = payload.estado.value if payload.estado else cumplimiento.estado
            cumplimiento.fecha_ultima_evaluacion = datetime.utcnow()
            cumplimiento.evaluado_por = current_user.id

            # Registrar en historial
            service = MatrizLegalService(db)
            service.registrar_cambio_cumplimiento(
                cumplimiento,
                estado_anterior,
                current_user.id,
                f"Actualización masiva a {payload.estado.value}"
            )
            updated += 1

    db.commit()

    return MessageResponse(
        message=f"Se actualizaron {updated} de {len(payload.cumplimiento_ids)} cumplimientos"
    )


@router.get("/empresas/{empresa_id}/cumplimiento/{cumplimiento_id}/historial",
            response_model=List[CumplimientoHistorialSchema])
def get_historial_cumplimiento(
    empresa_id: int,
    cumplimiento_id: int,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """Historial de cambios de un cumplimiento específico."""
    cumplimiento = db.query(MatrizLegalCumplimiento).filter(
        MatrizLegalCumplimiento.id == cumplimiento_id,
        MatrizLegalCumplimiento.empresa_id == empresa_id
    ).first()

    if not cumplimiento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Cumplimiento no encontrado"
        )

    historial = db.query(MatrizLegalCumplimientoHistorial).options(
        joinedload(MatrizLegalCumplimientoHistorial.usuario)
    ).filter(
        MatrizLegalCumplimientoHistorial.cumplimiento_id == cumplimiento_id
    ).order_by(
        MatrizLegalCumplimientoHistorial.created_at.desc()
    ).all()

    result = []
    for h in historial:
        result.append(CumplimientoHistorialSchema(
            id=h.id,
            cumplimiento_id=h.cumplimiento_id,
            estado_anterior=h.estado_anterior,
            estado_nuevo=h.estado_nuevo,
            observaciones=h.observaciones,
            created_at=h.created_at,
            creado_por=h.creado_por,
            usuario_nombre=f"{h.usuario.first_name} {h.usuario.last_name}" if h.usuario else None
        ))

    return result


# ==================== EXPORTACIÓN ====================

@router.get("/empresas/{empresa_id}/export/excel")
def export_matriz_empresa(
    empresa_id: int,
    incluir_no_aplicables: bool = Query(False),
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """Exporta la matriz legal de una empresa a Excel."""
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa no encontrada"
        )

    # Obtener normas con cumplimiento
    query = db.query(
        MatrizLegalNorma,
        MatrizLegalCumplimiento
    ).outerjoin(
        MatrizLegalCumplimiento,
        and_(
            MatrizLegalCumplimiento.norma_id == MatrizLegalNorma.id,
            MatrizLegalCumplimiento.empresa_id == empresa_id
        )
    ).filter(
        MatrizLegalNorma.activo == True
    )

    if not incluir_no_aplicables:
        service = MatrizLegalService(db)
        normas_aplicables = service.get_normas_aplicables_empresa(empresa)
        normas_ids = [n.id for n in normas_aplicables]
        query = query.filter(MatrizLegalNorma.id.in_(normas_ids))

    results = query.order_by(
        MatrizLegalNorma.clasificacion_norma,
        MatrizLegalNorma.anio.desc()
    ).all()

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz Legal"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Clasificación", "Tema General", "Subtema", "Año",
        "Tipo Norma", "Número", "Artículo", "Descripción",
        "Exigencias", "Estado Norma",
        "Cumple", "Evidencia", "Observaciones",
        "Plan de Acción", "Responsable", "Fecha Compromiso", "Seguimiento"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Datos
    for row_idx, (norma, cumplimiento) in enumerate(results, 2):
        data = [
            norma.clasificacion_norma,
            norma.tema_general,
            norma.subtema_riesgo_especifico,
            norma.anio,
            norma.tipo_norma,
            norma.numero_norma,
            norma.articulo,
            norma.descripcion_norma,
            norma.descripcion_articulo_exigencias,
            norma.estado,
            cumplimiento.estado if cumplimiento and cumplimiento.estado else "pendiente",
            cumplimiento.evidencia_cumplimiento if cumplimiento else None,
            cumplimiento.observaciones if cumplimiento else None,
            cumplimiento.plan_accion if cumplimiento else None,
            cumplimiento.responsable if cumplimiento else None,
            cumplimiento.fecha_compromiso.isoformat() if cumplimiento and cumplimiento.fecha_compromiso else None,
            cumplimiento.seguimiento if cumplimiento else None,
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Ajustar anchos
    column_widths = [20, 25, 25, 8, 15, 15, 15, 50, 50, 12, 12, 30, 30, 30, 20, 15, 30]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"matriz_legal_{empresa.nombre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/normas/export/excel")
def export_todas_normas(
    clasificacion: Optional[str] = Query(None),
    sector_economico_id: Optional[int] = Query(None),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Exporta el catálogo completo de normas a Excel."""
    query = db.query(MatrizLegalNorma).filter(MatrizLegalNorma.activo == True)

    if clasificacion:
        query = query.filter(MatrizLegalNorma.clasificacion_norma == clasificacion)

    if sector_economico_id:
        query = query.filter(MatrizLegalNorma.sector_economico_id == sector_economico_id)

    normas = query.order_by(
        MatrizLegalNorma.clasificacion_norma,
        MatrizLegalNorma.anio.desc()
    ).all()

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo Normas"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")

    # Headers
    headers = [
        "ID", "Clasificación", "Tema General", "Subtema", "Año",
        "Tipo", "Número", "Artículo", "Fecha Expedición", "Expedida Por",
        "Descripción", "Exigencias", "Estado", "Aplica General"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Datos
    for row_idx, norma in enumerate(normas, 2):
        data = [
            norma.id,
            norma.clasificacion_norma,
            norma.tema_general,
            norma.subtema_riesgo_especifico,
            norma.anio,
            norma.tipo_norma,
            norma.numero_norma,
            norma.articulo,
            norma.fecha_expedicion.isoformat() if norma.fecha_expedicion else None,
            norma.expedida_por,
            norma.descripcion_norma,
            norma.descripcion_articulo_exigencias,
            norma.estado,
            "Sí" if norma.aplica_general else "No",
        ]

        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"catalogo_normas_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== SUGERENCIAS IA ====================

@router.post("/normas/{norma_id}/sugerencias-ia")
async def generar_sugerencias_ia(
    norma_id: int,
    current_user: User = Depends(require_supervisor_or_admin),
    db: Session = Depends(get_db),
):
    """
    Genera sugerencias de cumplimiento usando IA (Perplexity).

    Retorna sugerencias para:
    - Evidencia de cumplimiento
    - Observaciones
    - Plan de acción
    """
    from app.services.ai_service import ai_service

    # Verificar que el servicio de IA esté configurado
    if not ai_service.is_configured():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Servicio de IA no configurado. Configure PERPLEXITY_API_KEY en las variables de entorno."
        )

    # Obtener la norma
    norma = db.query(MatrizLegalNorma).filter(MatrizLegalNorma.id == norma_id).first()
    if not norma:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Norma no encontrada"
        )

    try:
        sugerencias = await ai_service.generate_compliance_suggestions(
            tipo_norma=norma.tipo_norma,
            numero_norma=norma.numero_norma,
            anio=norma.anio,
            descripcion_norma=norma.descripcion_norma,
            articulo=norma.articulo,
            exigencias=norma.descripcion_articulo_exigencias,
            tema_general=norma.tema_general,
            clasificacion=norma.clasificacion_norma,
        )

        return {
            "success": True,
            "norma_id": norma_id,
            "sugerencias": sugerencias
        }

    except Exception as e:
        logger.error(f"Error generando sugerencias IA para norma {norma_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar sugerencias: {str(e)}"
        )
